# برای اجرای این کد، ابتدا باید کتابخانه openpyxl را نصب کنید.
# می‌توانید با دستور زیر در ترمینال یا Command Prompt آن را نصب کنید:
# pip install openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
import os
import random # برای نام های فرضی کارمندان
import math # برای محاسبات ریاضی

# --- توابع کمکی عمومی ---
def set_rtl_and_column_widths(ws, col_widths):
    """تنظیم راست به چپ بودن شیت و عرض ستون ها."""
    ws.sheet_view.rightToLeft = True
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

def add_header(ws, company_name, statement_name, date_line, currency_line=None):
    """افزودن سربرگ استاندارد به شیت های صورت مالی."""
    ws['A1'] = company_name
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = statement_name
    ws['A2'].font = Font(bold=True)
    ws['A3'] = date_line
    if currency_line:
        ws['A5'] = currency_line

# --- تابع تولید داده کارمندان (بدون تغییر) ---
def generate_all_employees_data(num_employees=100):
    """
    تولید لیست 100 کارمند با نام، واحد، سمت و تعداد فرزندان تصادفی.
    """
    employee_list = []
    first_names = ["علی", "رضا", "محمد", "حسین", "فاطمه", "زهرا", "مریم", "سعید", "امین", "نازنین", "کیارش", "سارا", "نیما", "آرزو", "بهروز", "کمال", "پریسا", "دانیال", "زینب", "مهرناز"]
    last_names = ["احمدی", "کریمی", "محمدی", "رضایی", "قاسمی", "نوروزی", "حسینی", "صادقی", "موسوی", "رحیمی", "یزدانی", "بهرامی", "فلاح", "شجاعی", "مظفری", "امیری", "جهانی", "هاشمی", "مختاری", "پورمحمدی"]

    units_and_roles_distribution = {
        "فارم": ["مدیر فارم", "دامپزشک", "نگهبان", "کارگر", "کارگر", "کارگر", "کارگر", "کارگر", "کارگر", "کارگر"], # 10 کارگر برای فارم های بزرگ
        "انبار": ["مدیر انبار", "انباردار", "کارگر", "کارگر"],
        "اداری": ["مدیرعامل", "مدیر مالی", "حسابدار", "مدیر منابع انسانی", "منشی", "کارشناس"],
        "فروش": ["مدیر فروش", "کارشناس فروش", "کارشناس فروش", "کارشناس فروش"]
    }
    
    current_employee_id = 1
    
    # پرسنل فارم
    for i in range(1, 11): # 10 فارم
        num_farm_employees = random.randint(6, 10) # 6 تا 10 نفر در هر فارم
        for _ in range(num_farm_employees):
            if current_employee_id > num_employees: break
            role_choice = random.choice(units_and_roles_distribution["فارم"])
            employee_list.append({
                "id": current_employee_id,
                "first_name": random.choice(first_names),
                "last_name": random.choice(last_names),
                "unit": f"فارم {i}",
                "role": role_choice,
                "num_children": random.randint(0, 3)
            })
            current_employee_id += 1

    # پرسنل انبار
    for i in range(1, 6): # 5 انبار
        num_warehouse_employees = random.randint(3, 5) # 3 تا 5 نفر در هر انبار
        for _ in range(num_warehouse_employees):
            if current_employee_id > num_employees: break
            role_choice = random.choice(units_and_roles_distribution["انبار"])
            employee_list.append({
                "id": current_employee_id,
                "first_name": random.choice(first_names),
                "last_name": random.choice(last_names),
                "unit": f"انبار {i}",
                "role": role_choice,
                "num_children": random.randint(0, 3)
            })
            current_employee_id += 1
            
    # پرسنل اداری و فروش (بقیه تا 100 نفر)
    while current_employee_id <= num_employees:
        is_admin_or_sales = random.choice(["اداری", "فروش"])
        role_choice = random.choice(units_and_roles_distribution[is_admin_or_sales])
        employee_list.append({
            "id": current_employee_id,
            "first_name": random.choice(first_names),
            "last_name": random.choice(last_names),
            "unit": is_admin_or_sales,
            "role": role_choice,
            "num_children": random.randint(0, 2) # معمولا اداری/فروش فرزندان کمتری دارند
        })
        current_employee_id += 1
        
    return employee_list[:num_employees] # اطمینان از دقیقاً 100 نفر

# ==============================================================================
# تابع جدید: populate_starting_balance_sheet (با مقادیر کاملاً جدید و تراز شده)
# ==============================================================================
def populate_starting_balance_sheet(ws):
    """ایجاد و پر کردن شیت ترازنامه افتتاحیه برای سال پایه (پایان 1401 / ابتدای 1402)."""
    ws.title = "ترازنامه پایه"
    col_widths = {'A': 5, 'B': 40, 'C': 45, 'D': 18} # تعیین عرض ستون‌ها
    set_rtl_and_column_widths(ws, col_widths) # تنظیم راست به چپ و عرض ستون‌ها
    add_header(ws, "شرکت نمونه (سهامی عام)", "ترازنامه افتتاحیه (پایه)", "در تاریخ 29 اسفند 1401 / 1 فروردین 1402", "(ارقام به میلیون ریال)")

    # --- دارایی‌ها ---
    ws.cell(row=8, column=2, value="دارایی ها").font = Font(bold=True)
    ws.cell(row=9, column=2, value="دارایی‌های جاری")
    ws.cell(row=10, column=3, value="موجودی نقد")
    ws.cell(row=10, column=4, value=800000) # موجودی نقد ابتدای 1402 (اصلاح شده برای تراز و جریان نقد مثبت)

    ws.cell(row=11, column=3, value="حساب‌ها و اسناد دریافتنی")
    ws.cell(row=11, column=4, value=515068) # بر اساس (95/365)*2,100,000 - دقیقاً متناسب با درآمد و دوره وصول 1402

    ws.cell(row=12, column=3, value="موجودی کالا")
    ws.cell(row=12, column=4, value=200000) # (فرضی - متناسب با COGS سال قبل)

    ws.cell(row=13, column=3, value="پیش‌پرداخت‌ها و سایر دارایی‌های جاری")
    ws.cell(row=13, column=4, value=50000)

    ws.cell(row=14, column=2, value="جمع دارایی‌های جاری")
    ws.cell(row=14, column=4, value="=SUM(D10:D13)")
    ws.cell(row=14, column=4).font = Font(bold=True)

    ws.append([]) # فاصله
    ws.cell(row=16, column=2, value="دارایی‌های غیرجاری").font = Font(bold=True)
    ws.cell(row=17, column=3, value="بهای تمام شده ناخالص دارایی‌های ثابت")
    ws.cell(row=17, column=4, value=3000000) # بهای تمام شده ناخالص در 1401/12/29
    ws.cell(row=18, column=3, value="کسر می‌شود: استهلاک انباشته")
    ws.cell(row=18, column=4, value=300000) # استهلاک انباشته در 1401/12/29 (مثبت وارد می‌شود)
    ws.cell(row=19, column=3, value="دارایی‌های ثابت مشهود (ارزش دفتری)")
    ws.cell(row=19, column=4, value="=D17-D18") # (ناخالص - استهلاک)

    ws.cell(row=20, column=3, value="سایر دارایی‌های غیرجاری")
    ws.cell(row=20, column=4, value=120000)

    ws.cell(row=21, column=2, value="جمع دارایی‌های غیرجاری")
    ws.cell(row=21, column=4, value="=SUM(D19:D20)")
    ws.cell(row=21, column=4).font = Font(bold=True)

    ws.append([]) # فاصله
    ws.cell(row=23, column=2, value="جمع کل دارایی‌ها").font = Font(bold=True)
    ws.cell(row=23, column=4, value="=D14+D21")
    ws.cell(row=23, column=4).font = Font(bold=True)

    # --- بدهی‌ها و حقوق مالکانه ---
    ws.append([]) # فاصله
    ws.cell(row=25, column=2, value="بدهی‌ها و حقوق مالکانه").font = Font(bold=True)
    ws.cell(row=26, column=2, value="بدهی‌های جاری")
    ws.cell(row=27, column=3, value="حساب‌ها و اسناد پرداختنی")
    ws.cell(row=27, column=4, value=380000) # (فرضی - متناسب با COGS سال قبل)

    ws.cell(row=28, column=3, value="مالیات پرداختنی")
    ws.cell(row=28, column=4, value=25000)

    ws.cell(row=29, column=3, value="سود سهام پرداختنی")
    ws.cell(row=29, column=4, value=75000)

    ws.cell(row=30, column=3, value="بخش جاری تسهیلات بلندمدت")
    ws.cell(row=30, column=4, value=60000) # بخش جاری وامی که در مفروضات هم بازپرداخت میشه

    ws.cell(row=31, column=2, value="جمع بدهی‌های جاری")
    ws.cell(row=31, column=4, value="=SUM(D27:D30)")
    ws.cell(row=31, column=4).font = Font(bold=True)

    ws.append([]) # فاصله
    ws.cell(row=33, column=2, value="بدهی‌های غیرجاری")
    ws.cell(row=34, column=3, value="تسهیلات مالی بلندمدت")
    ws.cell(row=34, column=4, value=700000) # مانده تسهیلات بلندمدت 1401 (غیرجاری)

    ws.cell(row=35, column=3, value="مزایای پایان خدمت کارکنان")
    ws.cell(row=35, column=4, value=150000)

    ws.cell(row=36, column=2, value="جمع بدهی‌های غیرجاری")
    ws.cell(row=36, column=4, value="=SUM(D34:D35)")
    ws.cell(row=36, column=4).font = Font(bold=True)

    ws.cell(row=37, column=2, value="جمع کل بدهی‌ها").font = Font(bold=True)
    ws.cell(row=37, column=4, value="=D31+D36")
    ws.cell(row=37, column=4).font = Font(bold=True)

    ws.append([]) # فاصله
    ws.cell(row=39, column=2, value="حقوق مالکانه").font = Font(bold=True)
    ws.cell(row=40, column=3, value="سرمایه")
    ws.cell(row=40, column=4, value=1000000) # همان مقدار قبلی

    ws.cell(row=41, column=3, value="اندوخته قانونی")
    ws.cell(row=41, column=4, value=120000) # همان مقدار قبلی

    ws.cell(row=42, column=3, value="سایر اندوخته‌ها")
    ws.cell(row=42, column=4, value=60000) # همان مقدار قبلی

    ws.cell(row=43, column=3, value="سود انباشته")
    # این فرمول سود انباشته رو تراز می‌کنه: جمع دارایی‌ها - جمع بدهی‌ها - سرمایه - اندوخته‌ها
    ws.cell(row=43, column=4, value="=D23-D37-D40-D41-D42") # D23 (جمع دارایی‌ها) - D37 (جمع بدهی‌ها) - D40 (سرمایه) - D41 (اندوخته قانونی) - D42 (سایر اندوخته‌ها)

    ws.cell(row=44, column=2, value="جمع کل حقوق مالکانه").font = Font(bold=True)
    ws.cell(row=44, column=4, value="=SUM(D40:D43)")
    ws.cell(row=44, column=4).font = Font(bold=True)

    ws.append([]) # فاصله
    ws.cell(row=46, column=2, value="جمع کل بدهی‌ها و حقوق مالکانه").font = Font(bold=True)
    ws.cell(row=46, column=4, value="=D37+D44")
    ws.cell(row=46, column=4).font = Font(bold=True)

    ws.append([]) # فاصله
    ws.cell(row=48, column=2, value="کنترل تراز (باید صفر باشد)").font = Font(bold=True)
    ws.cell(row=48, column=4, value="=D23-D46") # D23 (جمع دارایی‌ها) - D46 (جمع بدهی‌ها و حقوق مالکانه)
    ws.cell(row=48, column=4).font = Font(bold=True)
    ws.cell(row=48, column=4).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # برای نمایش راحت‌تر تراز

    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"


# ==============================================================================
# تابع ۱ (اصلاح شده): populate_assumptions_sheet
# ==============================================================================
def populate_assumptions_sheet(ws):
    """ایجاد و پر کردن شیت مفروضات نهایی مدل مالی و بازگرداندن نقشه آدرس ها."""
    ws.title = "مفروضات"
    set_rtl_and_column_widths(ws, {'A': 40, 'B': 18, 'C': 18})
    add_header(ws, "شرکت نمونه (سهامی عام)", "شیت مفروضات مدل مالی (سناریو سوددهی)", "")

    headers = ["شرح مفروضات", "مقدار (سال 1403)", "مقدار (سال 1402)"]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = Font(bold=True)

    assumptions = {
        "مفروضات صورت سود و زیان": [
            ("درصد رشد درآمدهای عملیاتی", 0.50, 0.15),  # <<-- افزایش به 50%
            ("بهای تمام شده به درصد از درآمد", 0.65, 0.68), # این دیگر در سود و زیان استفاده نمیشود
            ("هزینه‌های فروش، اداری و عمومی به درصد از درآمد", 0.12, 0.13), # این دیگر در سود و زیان استفاده نمیشود
            ("نرخ مالیات بر درآمد", 0.25, 0.25)
        ],
        "مفروضات ترازنامه (سرمایه در گردش)": [
            ("دوره وصول مطالبات (روز)", 90, 95),
            ("دوره گردش موجودی کالا (روز)", 120, 125),
            ("دوره پرداخت بدهی‌ها (روز)", 75, 80)
        ],
        "مفروضات دارایی ثابت و استهلاک": [
            ("سرمایه‌گذاری ثابت سالانه (CAPEX)", 600000, 450000),
            ("نرخ استهلاک سالانه (نسبت به بهای تمام شده اول دوره)", 0.10, 0.10)
        ],
        "مفروضات تامین مالی": [
            ("هزینه مالی ثابت", 50000, 60000), # <<-- کاهش به 50,000
            ("سود سهام پرداختی (درصد از سود خالص)", 0.40, 0.45),
            ("مبلغ وام جدید دریافتی طی سال", 850000, 300000),
            ("مبلغ بازپرداخت اصل وام طی سال", 50000, 40000),
            ("مانده اولیه تسهیلات بلندمدت (1402)", 0, 0), # این مقدار دیگر از اینجا خوانده نمیشود
        ]
    }

    assumption_map = {}
    current_row = 5
    for category, items in assumptions.items():
        ws.cell(row=current_row, column=1, value=category).font = Font(bold=True, color="000080")
        current_row += 1
        for desc, val_1403, val_1402 in items:
            ws.cell(row=current_row, column=1, value=desc)
            ws.cell(row=current_row, column=2, value=val_1403)
            ws.cell(row=current_row, column=3, value=val_1402)
            if "درصد" in desc or "نرخ" in desc:
                ws.cell(row=current_row, column=2).number_format = '0.00%'
                ws.cell(row=current_row, column=3).number_format = '0.00%'
            
            assumption_map[desc] = {'1403': f'B{current_row}', '1402': f'C{current_row}'}
            current_row += 1
        current_row += 1
    
    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"

    return assumption_map

# ==============================================================================
# تابع ۲ (بدون تغییر): populate_payroll_list_sheet
# ==============================================================================
def populate_payroll_list_sheet(ws):
    """پر کردن شیت لیست حقوق و دستمزد با آدرس‌دهی دقیق خروجی‌ها و هزینه کنترل شده."""
    col_widths = { 'A': 5, 'B': 15, 'C': 15, 'D': 15, 'E': 20, 'F': 20, 'G': 15, 'H': 15, 'I': 20, 'J': 20, 'K': 20, 'L': 20, 'M': 20, 'N': 20, 'O': 20, 'P': 20, 'Q': 20, 'R': 20, 'S': 20, 'T': 25, 'U': 25, 'V': 25, 'W': 25 }
    set_rtl_and_column_widths(ws, col_widths)
    add_header(ws, "شرکت نمونه (سهامی عام)", "لیست حقوق و دستمزد (سال 1403)", "تفکیک بر اساس واحد تولیدی", "(ارقام به ریال)")

    headers = [
        "ردیف", "نام", "نام خانوادگی", "واحد", "سمت", "کد ملی", "شماره بیمه", "تعداد اولاد", 
        "حقوق پایه", "حق مسکن", "حق بن", "حق اولاد", "جمع مزایا", "حقوق ناخالص", "حقوق مشمول بیمه", 
        "بیمه سهم کارمند (7%)", "حقوق مشمول مالیات", "مالیات حقوق", "کسورات متفرقه", "جمع کسورات", 
        "حقوق خالص (پرداختنی)", "بیمه سهم کارفرما (23%)", "کل هزینه برای کارفرما"
    ]
    ws.append(headers)

    min_wage_daily_1403 = 2_388_728
    housing_allowance_1403 = 9_000_000
    consumer_basket_allowance_1403 = 14_000_000
    tax_exemption_monthly_1403 = 120_000_000
    
    # ضریب تبدیل برای سال 1402
    min_wage_daily_1402 = 1_769_428
    base_1402_ratio = 0.8 # ضریب کاهشی برای حقوق 1402
    
    tax_rate_excess = 0.10
    ins_employee_share_rate = 0.07
    ins_employer_share_rate = 0.23

    employees_data_list = generate_all_employees_data(100)
    initial_data_row_start = ws.max_row + 1 
    min_wage_monthly_1403 = min_wage_daily_1403 * 30

    for i, emp in enumerate(employees_data_list):
        current_row = initial_data_row_start + i

        # <<-- کاهش بازه حقوق برای کنترل هزینه
        base_salary_per_employee = random.randint(min_wage_monthly_1403, 120_000_000)
        if "مدیر" in emp["role"]:
            base_salary_per_employee = random.randint(130_000_000, 200_000_000)
        elif "دامپزشک" in emp["role"]:
            base_salary_per_employee = random.randint(100_000_000, 160_000_000)
        elif "کارشناس" in emp["role"] or "حسابدار" in emp["role"] or "انباردار" in emp["role"]:
            base_salary_per_employee = random.randint(80_000_000, 130_000_000)

        child_benefit_amount_1403 = emp["num_children"] * 3 * min_wage_daily_1403
        
        ws.cell(row=current_row, column=1, value=emp["id"])
        ws.cell(row=current_row, column=2, value=emp["first_name"])
        ws.cell(row=current_row, column=3, value=emp["last_name"])
        ws.cell(row=current_row, column=4, value=emp["unit"])
        ws.cell(row=current_row, column=5, value=emp["role"])
        ws.cell(row=current_row, column=6, value=random.randint(1000000000, 9999999999))
        ws.cell(row=current_row, column=7, value=random.randint(10000000000, 99999999999))
        ws.cell(row=current_row, column=8, value=emp["num_children"])
        ws.cell(row=current_row, column=9, value=base_salary_per_employee)
        ws.cell(row=current_row, column=10, value=housing_allowance_1403)
        ws.cell(row=current_row, column=11, value=consumer_basket_allowance_1403)
        ws.cell(row=current_row, column=12, value=child_benefit_amount_1403)
        
        ws.cell(row=current_row, column=13, value=f"=SUM(J{current_row}:L{current_row})")
        ws.cell(row=current_row, column=14, value=f"=I{current_row}+M{current_row}")
        ws.cell(row=current_row, column=15, value=f"=I{current_row}+J{current_row}+K{current_row}")
        ws.cell(row=current_row, column=16, value=f"=O{current_row}*{ins_employee_share_rate}")
        ws.cell(row=current_row, column=17, value=f"=MAX(0, N{current_row}-P{current_row}-{tax_exemption_monthly_1403})")
        ws.cell(row=current_row, column=18, value=f"=ROUND(Q{current_row}*{tax_rate_excess},0)") 
        ws.cell(row=current_row, column=19, value=random.randint(500_000, 2_000_000))
        ws.cell(row=current_row, column=20, value=f"=SUM(P{current_row},R{current_row},S{current_row})") 
        ws.cell(row=current_row, column=21, value=f"=N{current_row}-T{current_row}") 
        ws.cell(row=current_row, column=22, value=f"=O{current_row}*{ins_employer_share_rate}")
        ws.cell(row=current_row, column=23, value=f"=N{current_row}+V{current_row}")

    final_data_row = ws.max_row
    total_monthly_row_idx = final_data_row + 2
    total_yearly_rial_idx = total_monthly_row_idx + 1
    total_yearly_million_idx = total_monthly_row_idx + 2

    ws[f'A{total_monthly_row_idx}'] = "جمع کل ماهانه (ریال)"
    for col_idx in range(9, 24):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{total_monthly_row_idx}'] = f'=SUM({col_letter}{initial_data_row_start}:{col_letter}{final_data_row})'
    
    ws[f'A{total_yearly_rial_idx}'] = "جمع کل سالانه (ریال)"
    for col_idx in range(9, 24):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{total_yearly_rial_idx}'] = f'={col_letter}{total_monthly_row_idx}*12'

    ws[f'A{total_yearly_million_idx}'] = "جمع کل سالانه (میلیون ریال)"
    for col_idx in range(9, 24):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{total_yearly_million_idx}'] = f'={col_letter}{total_yearly_rial_idx}/1000000'
    
    # *** خروجی‌های کلیدی برای سایر شیت‌ها ***
    output_start_row = total_yearly_million_idx + 2
    ws[f'A{output_start_row}'] = "خروجی برای سایر شیت‌ها (ارقام به میلیون ریال):"
    
    # هزینه کل پرسنل فروش (برای یادداشت 8)
    ws[f'B{output_start_row + 1}'] = "کل هزینه سالانه پرسنل فروش - 1403:"
    ws[f'E{output_start_row + 1}'] = f"""=ROUND(SUMIF(D{initial_data_row_start}:D{final_data_row},"*فروش*",W{initial_data_row_start}:W{final_data_row})*12/1000000,0)"""
    ws[f'B{output_start_row + 2}'] = "کل هزینه سالانه پرسنل فروش - 1402:"
    ws[f'F{output_start_row + 1}'] = f"""=ROUND(E{output_start_row + 1}*{base_1402_ratio},0)"""

    # هزینه کل پرسنل اداری (برای یادداشت 8)
    ws[f'B{output_start_row + 3}'] = "کل هزینه سالانه پرسنل اداری - 1403:"
    ws[f'E{output_start_row + 3}'] = f"""=ROUND(SUMIF(D{initial_data_row_start}:D{final_data_row},"*اداری*",W{initial_data_row_start}:W{final_data_row})*12/1000000,0)"""
    ws[f'B{output_start_row + 4}'] = "کل هزینه سالانه پرسنل اداری - 1402:"
    ws[f'F{output_start_row + 3}'] = f"""=ROUND(E{output_start_row + 3}*{base_1402_ratio},0)"""

    # هزینه کل پرسنل تولید (برای یادداشت 9 - بهای تمام شده)
    ws[f'B{output_start_row + 5}'] = "کل هزینه سالانه پرسنل تولید - 1403:"
    ws[f'E{output_start_row + 5}'] = f"""=ROUND(SUMIFS(W{initial_data_row_start}:W{final_data_row}, D{initial_data_row_start}:D{final_data_row}, "<>*فروش*", D{initial_data_row_start}:D{final_data_row}, "<>*اداری*")*12/1000000, 0)"""
    ws[f'B{output_start_row + 6}'] = "کل هزینه سالانه پرسنل تولید - 1402:"
    ws[f'F{output_start_row + 5}'] = f"""=ROUND(E{output_start_row + 5}*{base_1402_ratio},0)"""
    
    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به سود و زیان").hyperlink = f"#'سودوزیان'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"


# ==============================================================================
# تابع ۳ (بدون تغییر): populate_detailed_inventory_sheet
# ==============================================================================
def populate_detailed_inventory_sheet(ws):
    col_widths = {'A': 5, 'B': 20, 'C': 10, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 
                  'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 15}
    set_rtl_and_column_widths(ws, col_widths)
    ws['A1'] = "شرکت نمونه (سهامی عام)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "موجودی تفصیلی انبار (مقدار و ریال)"
    ws['A2'].font = Font(bold=True)
    ws['A3'] = "برای سال مالی منتهی به 29 اسفند 1403 و 1402"
    ws['A5'] = "(ارقام به ریال برای قیمت واحد و میلیون ریال برای مقادیر)"

    ws.cell(row=7, column=1, value="ردیف")
    ws.cell(row=7, column=2, value="نام کالا")
    ws.cell(row=7, column=3, value="واحد")
    ws.cell(row=7, column=4, value="ابتدای دوره 1403 (مقدار)")
    ws.cell(row=7, column=5, value="ورود 1403 (مقدار)")
    ws.cell(row=7, column=6, value="خروج 1403 (مقدار)")
    ws.cell(row=7, column=7, value="پایان دوره 1403 (مقدار)")
    ws.cell(row=7, column=8, value="ابتدای دوره 1402 (مقدار)")
    ws.cell(row=7, column=9, value="ورود 1402 (مقدار)")
    ws.cell(row=7, column=10, value="خروج 1402 (مقدار)")
    ws.cell(row=7, column=11, value="پایان دوره 1402 (مقدار)")
    ws.cell(row=7, column=12, value="قیمت واحد میانگین (ریال)")


    inventory_items = [
        (1, "جوجه یک روزه", "عدد", 800000, 4500000, 4300000, 600000, 3800000, 3400000, 150),
        (2, "خوراک (دان)", "کیلوگرم", 8000000, 25000000, 24000000, 6000000, 22000000, 20000000, 390), # <<-- کاهش قیمت
        (3, "مرغ در حال رشد", "عدد", 200000, 900000, 850000, 150000, 750000, 700000, 3600),
        (4, "دارو و واکسن", "بسته", 10000, 40000, 38000, 8000, 30000, 28000, 25),
        (5, "مرغ آماده فروش", "کیلوگرم", 50000, 900000, 880000, 40000, 750000, 720000, 140)
    ]

    start_row_data = 8
    for item in inventory_items:
        row_idx = start_row_data + inventory_items.index(item)
        ws.cell(row=row_idx, column=1, value=item[0])
        ws.cell(row=row_idx, column=2, value=item[1])
        ws.cell(row=row_idx, column=3, value=item[2])
        ws.cell(row=row_idx, column=4, value=item[3])
        ws.cell(row=row_idx, column=5, value=item[4])
        ws.cell(row=row_idx, column=6, value=item[5])
        ws.cell(row=row_idx, column=7, value=f'=D{row_idx}+E{row_idx}-F{row_idx}')
        ws.cell(row=row_idx, column=8, value=item[6])
        ws.cell(row=row_idx, column=9, value=item[7])
        ws.cell(row=row_idx, column=10, value=item[8])
        ws.cell(row=row_idx, column=11, value=f'=H{row_idx}+I{row_idx}-J{row_idx}')
        ws.cell(row=row_idx, column=12, value=item[9])


    value_header_row = 17
    ws.cell(row=value_header_row, column=2, value="اطلاعات ریالی (میلیون ریال)")
    ws.cell(row=value_header_row, column=4, value="ابتدای دوره 1403")
    ws.cell(row=value_header_row, column=5, value="ورود 1403")
    ws.cell(row=value_header_row, column=6, value="خروج (بهای تمام شده) 1403")
    ws.cell(row=value_header_row, column=7, value="پایان دوره 1403")
    ws.cell(row=value_header_row, column=8, value="ابتدای دوره 1402")
    ws.cell(row=value_header_row, column=9, value="ورود 1402")
    ws.cell(row=value_header_row, column=10, value="خروج (بهای تمام شده) 1402")
    ws.cell(row=value_header_row, column=11, value="پایان دوره 1402")

    value_start_row = 18
    for i in range(len(inventory_items)):
        data_row = start_row_data + i
        value_row = value_start_row + i
        ws.cell(row=value_row, column=1, value=f'=A{data_row}')
        ws.cell(row=value_row, column=2, value=f'=B{data_row}')
        ws.cell(row=value_row, column=3, value='م.ر')
        ws.cell(row=value_row, column=4, value=f'=ROUND(D{data_row}*L{data_row}/1000000,0)')
        ws.cell(row=value_row, column=5, value=f'=ROUND(E{data_row}*L{data_row}/1000000,0)')
        ws.cell(row=value_row, column=6, value=f'=ROUND(F{data_row}*L{data_row}/1000000,0)')
        ws.cell(row=value_row, column=7, value=f'=ROUND(G{data_row}*L{data_row}/1000000,0)')
        ws.cell(row=value_row, column=8, value=f'=ROUND(H{data_row}*L{data_row}/1000000,0)')
        ws.cell(row=value_row, column=9, value=f'=ROUND(I{data_row}*L{data_row}/1000000,0)')
        ws.cell(row=value_row, column=10, value=f'=ROUND(J{data_row}*L{data_row}/1000000,0)')
        ws.cell(row=value_row, column=11, value=f'=ROUND(K{data_row}*L{data_row}/1000000,0)')

    total_row_value = value_start_row + len(inventory_items) + 4
    ws.cell(row=total_row_value, column=2, value="جمع کل (میلیون ریال)").font = Font(bold=True)
    for col_idx in range(4, 12):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{total_row_value}'] = f'=SUM({col_letter}{value_start_row}:{col_letter}{value_start_row + len(inventory_items) - 1})'
        ws[f'{col_letter}{total_row_value}'].font = Font(bold=True)

    output_row_start = total_row_value + 2
    ws.cell(row=output_row_start, column=2, value="**خروجی‌ها برای سایر شیت‌ها**").font = Font(bold=True)
    ws.cell(row=output_row_start+1, column=2, value="بهای تمام شده سال 1403 (برای سود و زیان):")
    ws.cell(row=output_row_start+1, column=6, value=f"=F{total_row_value}").font = Font(bold=True)

    ws.cell(row=output_row_start+2, column=2, value="بهای تمام شده سال 1402 (برای سود و زیان):")
    ws.cell(row=output_row_start+2, column=6, value=f"=J{total_row_value}").font = Font(bold=True)

    ws.cell(row=output_row_start+3, column=2, value="موجودی پایان دوره 1403 (برای ترازنامه):")
    ws.cell(row=output_row_start+3, column=6, value=f"=G{total_row_value}").font = Font(bold=True)

    ws.cell(row=output_row_start+4, column=2, value="موجودی پایان دوره 1402 (برای ترازنامه):")
    ws.cell(row=output_row_start+4, column=6, value=f"=K{total_row_value}").font = Font(bold=True)

    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"

# ==============================================================================
# تابع اصلاح شده ۱: populate_note_8_and_9 (یادداشت‌های هزینه)
# ==============================================================================
def populate_note_8_and_9(wb):
    ## یادداشت 9: بهای تمام شده
    ws9 = wb['9']
    # Clear existing content to avoid duplicates on re-run if sheet already exists
    for row in ws9.iter_rows(min_row=7): # Start clearing from row 7 to keep header
        for cell in row:
            cell.value = None

    set_rtl_and_column_widths(ws9, {'A': 5, 'B': 45, 'C': 18, 'D': 18})
    add_header(ws9, "شرکت نمونه", "یادداشت 9: بهای تمام شده", "سال مالی منتهی به 29 اسفند 1403 و 1402", "(مبالغ به میلیون ریال)")
    ws9.append(['', 'شرح', '1403', '1402'])
    cogs_items = [
        ("بهای تمام شده کالای فروش رفته", "='موجودی_تفصیلی'!F27", "='موجودی_تفصیلی'!J27"),
        ("حقوق و دستمزد مستقیم تولید", 450000, 400000),
        ("هزینه استهلاک دارایی‌های تولیدی (80%)", "='گردش دارایی ثابت'!D11*0.8", "='گردش دارایی ثابت'!D6*0.8"),
        ("سایر هزینه‌های مستقیم تولید (سربار)", 50000, 45000)
    ]
    start_row = ws9.max_row + 1
    for item in cogs_items:
        ws9.append(['', item[0], item[1], item[2]])
    end_row = ws9.max_row
    total_row = end_row + 1
    ws9.cell(row=total_row, column=2, value="جمع کل بهای تمام شده").font = Font(bold=True)
    ws9[f'C{total_row}'] = f"=SUM(C{start_row}:C{end_row})"
    ws9[f'D{total_row}'] = f"=SUM(D{start_row}:D{end_row})"
    ws9.cell(row=1, column=max(1, ws9.max_column - 1), value="بازگشت به سود و زیان").hyperlink = f"#'سودوزیان'!A1"
    ws9.cell(row=1, column=max(1, ws9.max_column - 1)).style = "Hyperlink"

    ## یادداشت 8: هزینه‌های فروش، اداری و عمومی
    ws8 = wb['8']
    # Clear existing content to avoid duplicates on re-run if sheet already exists
    for row in ws8.iter_rows(min_row=7): # Start clearing from row 7 to keep header
        for cell in row:
            cell.value = None

    set_rtl_and_column_widths(ws8, {'A': 5, 'B': 45, 'C': 18, 'D': 18})
    add_header(ws8, "شرکت نمونه", "یادداشت 8: هزینه‌های فروش، اداری و عمومی", "سال مالی منتهی به 29 اسفند 1403 و 1402", "(مبالغ به میلیون ریال)")
    ws8.append(['', 'شرح', '1403', '1402'])
    ws8.append(['', 'الف) هزینه‌های فروش و توزیع:'])
    sga_sales_items = [
        ("هزینه پرسنل فروش", 250000, 220000),
        ("هزینه تبلیغات و بازاریابی", 50000, 40000)
    ]
    start_row_s = ws8.max_row + 1
    for item in sga_sales_items:
        ws8.append(['', item[0], item[1], item[2]])
    end_row_s = ws8.max_row
    total_row_s = end_row_s + 1
    ws8.cell(row=total_row_s, column=2, value="جمع هزینه‌های فروش").font = Font(bold=True)
    ws8[f'C{total_row_s}'] = f"=SUM(C{start_row_s}:C{end_row_s})"
    ws8[f'D{total_row_s}'] = f"=SUM(D{start_row_s}:D{end_row_s})"

    ws8.append(['', 'ب) هزینه‌های اداری و عمومی:'])
    sga_admin_items = [
        ("هزینه پرسنل اداری", 350000, 320000),
        ("هزینه استهلاک دارایی‌های اداری (20%)", "='گردش دارایی ثابت'!D11*0.2", "='گردش دارایی ثابت'!D6*0.2"),
        ("هزینه ذخیره مزایای پایان خدمت کارکنان", 80000, 75000), ## <-- فرض ثابت و شفاف برای هزینه
        ("سایر هزینه‌های اداری", 30000, 25000)
    ]
    start_row_a = ws8.max_row + 1
    for item in sga_admin_items:
        ws8.append(['', item[0], item[1], item[2]])
    end_row_a = ws8.max_row
    total_row_a = end_row_a + 1
    ws8.cell(row=total_row_a, column=2, value="جمع هزینه‌های اداری").font = Font(bold=True)
    ws8[f'C{total_row_a}'] = f"=SUM(C{start_row_a}:C{end_row_a})"
    ws8[f'D{total_row_a}'] = f"=SUM(D{start_row_a}:D{end_row_a})"
     
    ws8.append([''])
    total_row_all = ws8.max_row + 1
    ws8.cell(row=total_row_all, column=2, value="جمع کل هزینه‌های فروش، اداری و عمومی").font = Font(bold=True)
    ws8[f'C{total_row_all}'] = f"=C{total_row_s}+C{total_row_a}"
    ws8[f'D{total_row_all}'] = f"=D{total_row_s}+D{total_row_a}"
    ws8.cell(row=1, column=max(1, ws8.max_column - 1), value="بازگشت به سود و زیان").hyperlink = f"#'سودوزیان'!A1"
    ws8.cell(row=1, column=max(1, ws8.max_column - 1)).style = "Hyperlink"


# ==============================================================================
# تابع اصلاح شده ۳: populate_profit_loss_sheet (کاملاً یکپارچه)
# ==============================================================================
def populate_profit_loss_sheet(ws, assumption_map):
    """ایجاد صورت سود و زیان یکپارچه که هزینه‌ها را از یادداشت‌ها می‌خواند."""
    col_widths = {'A': 5, 'B': 40, 'C': 12, 'D': 18, 'E': 10, 'F': 18, 'G': 18}
    set_rtl_and_column_widths(ws, col_widths)
    add_header(ws, "شرکت نمونه (سهامی عام)", "صورت سود و زیان (یکپارچه)", "سال مالی منتهی به 29 اسفند 1403 و 1402", "(ارقام به میلیون ریال)")
    # Clear existing content to avoid duplicates on re-run if sheet already exists
    for row in ws.iter_rows(min_row=7): # Start clearing from row 7 to keep header
        for cell in row:
            cell.value = None

    ws.append(["", "", "", "", "یادداشت", "سال 1403", "سال 1402"])

    base_revenue_1402 = 2100000
    ws['G8'] = base_revenue_1402
    
    ws['F8'] = f"=G8*(1+'مفروضات'!{assumption_map['درصد رشد درآمدهای عملیاتی']['1403']})"
    
    # <<-- اصلاح: لینک به ردیف صحیح 13 در یادداشت‌های هزینه
    ws['F9'] = "=-'9'!C13"
    ws['G9'] = "=-'9'!D13"
    
    ws['F12'] = "=-'8'!C23"
    ws['G12'] = "=-'8'!D23"

    ws['F10'] = '=SUM(F8:F9)'
    ws['G10'] = '=SUM(G8:G9)'
    
    ws['F13'] = 150000 
    ws['G13'] = 120000
    ws['F14'] = -10000
    ws['G14'] = -30000
    
    ws['F15'] = '=SUM(F10,F12:F14)'
    ws['G15'] = '=SUM(G10,G12:G14)'
    
    ws['F17'] = f"=-'مفروضات'!{assumption_map['هزینه مالی ثابت']['1403']}"
    ws['G17'] = f"=-'مفروضات'!{assumption_map['هزینه مالی ثابت']['1402']}"
    
    ws['F18'] = '=SUM(F15,F17)'
    ws['G18'] = '=SUM(G15,G17)'
    
    ws['F20'] = f"=IF(F18>0, F18*(-'مفروضات'!{assumption_map['نرخ مالیات بر درآمد']['1403']}), 0)"
    ws['G20'] = f"=IF(G18>0, G18*(-'مفروضات'!{assumption_map['نرخ مالیات بر درآمد']['1402']}), 0)"
    
    ws['F21'] = '=F18+F20'
    ws['G21'] = '=G18+G20'

    items = [
        (8, "درآمدهای عملیاتی", '5'), (9, "بهای تمام شده درآمدهای عملیاتی", '9'),
        (10, "سود ناخالص", None), (12, "هزینه‌های فروش ، اداری و عمومی", '8'),
        (13, "سایر درآمدها", '26.27'), (14, "سایر هزینه‌ها", '26.27'),
        (15, "سود عملیاتی", None), (17, "هزینه‌های مالی", '26.27'),
        (18, "سود قبل از مالیات", None), (20, "مالیات بر درآمد", '34'),
        (21, "سود خالص", None)
    ]
    for row_num, text, note_name in items:
        ws[f'B{row_num}'] = text
        if "سود" in text or "زیان" in text:
            for col in ['B', 'F', 'G']:
                ws[f'{col}{row_num}'].font = Font(bold=True)
        if note_name:
            ws.cell(row=row_num, column=5, value=note_name).hyperlink = f"#'{note_name}'!A1"
            ws.cell(row=row_num, column=5).style = "Hyperlink"

    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"

# ==============================================================================
# تابع اصلاح شده ۴: populate_balance_sheet (با ارجاعات جدید به ترازنامه پایه)
# ==============================================================================
def populate_balance_sheet(ws, assumption_map):
    col_widths = {'A': 5, 'B': 40, 'C': 45, 'D': 12, 'E': 18, 'F': 18}
    set_rtl_and_column_widths(ws, col_widths)
    add_header(ws, "شرکت نمونه (سهامی عام)", "صورت وضعیت مالی (پویا)", "در تاریخ 29 اسفند 1403 و 1402", "(ارقام به میلیون ریال)")

    for row in ws.iter_rows(min_row=7):
        for cell in row:
            cell.value = None

    ws.append(["", "", "", "یادداشت", "1403", "1402"]) 
    
    # --- دارایی‌ها ---
    ws.cell(row=9, column=2, value="دارایی‌های جاری")
    
    ws.cell(row=10, column=3, value="موجودی نقد")
    ws.cell(row=10, column=4, value=6)
    ws.cell(row=10, column=5, value="='جریان های نقدی'!C31") # پایان دوره 1403 از جریان نقد
    ws.cell(row=10, column=6, value="='ترازنامه پایه'!D10") # ابتدای دوره 1402 از ترازنامه پایه
    ws.cell(row=10, column=4).hyperlink = f"#'6'!A1"
    ws.cell(row=10, column=4).style = "Hyperlink"

    ws.cell(row=11, column=3, value="حساب‌ها و اسناد دریافتنی")
    ws.cell(row=11, column=5, value=f"=('مفروضات'!{assumption_map['دوره وصول مطالبات (روز)']['1403']}/365)*'سودوزیان'!F8")
    ws.cell(row=11, column=6, value=f"=('مفروضات'!{assumption_map['دوره وصول مطالبات (روز)']['1402']}/365)*'سودوزیان'!G8")
    ws.cell(row=11, column=4, value="42.43")
    ws.cell(row=11, column=4).hyperlink = f"#'42.43'!A1"
    ws.cell(row=11, column=4).style = "Hyperlink"

    ws.cell(row=12, column=3, value="موجودی کالا")
    ws.cell(row=12, column=4, value=9)
    # <<-- اصلاح شده: لینک به ردیف صحیح 27 در شیت موجودی تفصیلی
    ws.cell(row=12, column=5, value="='موجودی_تفصیلی'!G27")
    ws.cell(row=12, column=6, value="='ترازنامه پایه'!D12") # ابتدای دوره 1402 از ترازنامه پایه
    ws.cell(row=12, column=4).hyperlink = f"#'موجودی'!A1"
    ws.cell(row=12, column=4).style = "Hyperlink"

    ws.cell(row=13, column=3, value="پیش‌پرداخت‌ها و سایر دارایی‌های جاری")
    ws.cell(row=13, column=5, value=50000)
    ws.cell(row=13, column=6, value=40000)
    ws.cell(row=13, column=4, value=10)
    ws.cell(row=13, column=4).hyperlink = f"#'10.11.12'!A1"
    ws.cell(row=13, column=4).style = "Hyperlink"

    ws.cell(row=14, column=2, value="جمع دارایی‌های جاری")
    ws.cell(row=14, column=5, value="=SUM(E10:E13)")
    ws.cell(row=14, column=6, value="=SUM(F10:F13)")

    ws.append([])
    ws.cell(row=16, column=2, value="دارایی‌های غیرجاری")
    
    ws.cell(row=17, column=3, value="دارایی‌های ثابت مشهود (ارزش دفتری)")
    ws.cell(row=17, column=4, value="گردش دارایی ثابت")
    ws.cell(row=17, column=5, value="='گردش دارایی ثابت'!E13") # پایان دوره 1403 از گردش دارایی
    ws.cell(row=17, column=6, value="='گردش دارایی ثابت'!E8") # ابتدای دوره 1402 از گردش دارایی
    ws.cell(row=17, column=4).hyperlink = f"#'گردش دارایی ثابت'!A1"
    ws.cell(row=17, column=4).style = "Hyperlink"

    ws.cell(row=18, column=3, value="سایر دارایی‌های غیرجاری")
    ws.cell(row=18, column=5, value=100000)
    ws.cell(row=18, column=6, value=80000)
    ws.cell(row=18, column=4, value=13)
    ws.cell(row=18, column=4).hyperlink = f"#'13'!A1"
    ws.cell(row=18, column=4).style = "Hyperlink"

    ws.cell(row=19, column=2, value="جمع دارایی‌های غیرجاری")
    ws.cell(row=19, column=5, value="=SUM(E17:E18)")
    ws.cell(row=19, column=6, value="=SUM(F17:F18)")
    ws.append([]);
    ws.cell(row=21, column=2, value="جمع کل دارایی‌ها")
    ws.cell(row=21, column=5, value="=E14+E19")
    ws.cell(row=21, column=6, value="=F14+F19")
    
    # --- بدهی‌ها و حقوق مالکانه ---
    ws.append([]) 
    ws.cell(row=23, column=2, value="بدهی‌ها و حقوق مالکانه")
    ws.cell(row=24, column=2, value="بدهی‌های جاری")
    ws.cell(row=25, column=3, value="حساب‌ها و اسناد پرداختنی")
    ws.cell(row=25, column=5, value=f"=('مفروضات'!{assumption_map['دوره پرداخت بدهی‌ها (روز)']['1403']}/365)*-'9'!C13") # Updated to use COGS from Note 9
    ws.cell(row=25, column=6, value=f"=('مفروضات'!{assumption_map['دوره پرداخت بدهی‌ها (روز)']['1402']}/365)*-'9'!D13") # Updated to use COGS from Note 9
    ws.cell(row=25, column=4, value="28.29.30.31")
    ws.cell(row=25, column=4).hyperlink = f"#'28.29.30.31'!A1"
    ws.cell(row=25, column=4).style = "Hyperlink"

    ws.cell(row=26, column=3, value="مالیات پرداختنی")
    ws.cell(row=26, column=4, value=17)
    ws.cell(row=26, column=5, value="='سودوزیان'!F20*-1")
    ws.cell(row=26, column=6, value="='سودوزیان'!G20*-1")
    ws.cell(row=26, column=4).hyperlink = f"#'17'!A1"
    ws.cell(row=26, column=4).style = "Hyperlink"

    ws.cell(row=27, column=3, value="سود سهام پرداختنی")
    ws.cell(row=27, column=4, value=18)
    # <<-- اصلاح شده: این مقادیر از جریان نقدی پرداخت شده و باید در ترازنامه صفر شوند
    ws.cell(row=27, column=5, value=0)
    ws.cell(row=27, column=6, value=0)
    ws.cell(row=27, column=4).hyperlink = f"#'18'!A1"
    ws.cell(row=27, column=4).style = "Hyperlink"

    ws.cell(row=28, column=3, value="بخش جاری تسهیلات بلندمدت")
    ws.cell(row=28, column=5, value=f"='مفروضات'!{assumption_map['مبلغ بازپرداخت اصل وام طی سال']['1403']}")
    ws.cell(row=28, column=6, value=f"='مفروضات'!{assumption_map['مبلغ بازپرداخت اصل وام طی سال']['1402']}")
    ws.cell(row=28, column=4, value=30)
    ws.cell(row=28, column=4).hyperlink = f"#'16'!A1"
    ws.cell(row=28, column=4).style = "Hyperlink"

    ws.cell(row=29, column=2, value="جمع بدهی‌های جاری")
    ws.cell(row=29, column=5, value="=SUM(E25:E28)")
    ws.cell(row=29, column=6, value="=SUM(F25:F28)")

    ws.append([])
    ws.cell(row=31, column=2, value="بدهی‌های غیرجاری")
    
    ws.cell(row=32, column=3, value="تسهیلات مالی بلندمدت")
    ws.cell(row=32, column=4, value=19)
    ws.cell(row=32, column=5, value=f"=F32+'مفروضات'!{assumption_map['مبلغ وام جدید دریافتی طی سال']['1403']}-E28") # E28 is current portion of long-term debt repaid
    ws.cell(row=32, column=6, value="='ترازنامه پایه'!D34")
    ws.cell(row=32, column=4).hyperlink = f"#'19'!A1"
    ws.cell(row=32, column=4).style = "Hyperlink"

    ws.cell(row=33, column=3, value="مزایای پایان خدمت کارکنان")
    ws.cell(row=33, column=4, value=20)
    ws.cell(row=33, column=5, value="=F33+'8'!C14") # F33 is prior year balance, '8'!C14 is the current year expense from note 8
    ws.cell(row=33, column=6, value="='ترازنامه پایه'!D35")
    ws.cell(row=33, column=4).hyperlink = f"#'20'!A1"
    ws.cell(row=33, column=4).style = "Hyperlink"

    ws.cell(row=34, column=2, value="جمع بدهی‌های غیرجاری")
    ws.cell(row=34, column=5, value="=SUM(E32:E33)")
    ws.cell(row=34, column=6, value="=SUM(F32:F33)")

    ws.cell(row=35, column=2, value="جمع کل بدهی‌ها")
    ws.cell(row=35, column=5, value="=E29+E34")
    ws.cell(row=35, column=6, value="=F29+F34")
    ws.append([]);
    ws.cell(row=37, column=2, value="حقوق مالکانه")
    ws.cell(row=37, column=5, value="='حقوق مالکانه'!F18")
    ws.cell(row=37, column=6, value="='حقوق مالکانه'!F12")
    ws.cell(row=37, column=4, value="21")
    ws.cell(row=37, column=4).hyperlink = f"#'21'!A1"
    ws.cell(row=37, column=4).style = "Hyperlink"

    ws.cell(row=38, column=2, value="جمع کل بدهی‌ها و حقوق مالکانه")
    ws.cell(row=38, column=5, value="=E35+E37")
    ws.cell(row=38, column=6, value="=F35+F37")
    ws.append([]);
    ws.cell(row=40, column=2, value="کنترل تراز")
    ws.cell(row=40, column=5, value='=IF(ROUND(E21-E38,0)=0,"تراز","عدم تراز")')
    ws.cell(row=40, column=6, value='=IF(ROUND(F21-F38,0)=0,"تراز","عدم تراز")')

# ==============================================================================
# تابع اصلاح شده ۵: populate_fixed_asset_roll_forward_sheet
# ==============================================================================
def populate_fixed_asset_roll_forward_sheet(ws, assumption_map):
    """ایجاد و پر کردن شیت گردش دارایی‌های ثابت مشهود (پویا)."""
    col_widths = {'A': 35, 'B': 20, 'C': 20, 'D': 20, 'E': 20}
    set_rtl_and_column_widths(ws, col_widths)
    add_header(ws, "شرکت نمونه (سهامی عام)", "گردش دارایی‌های ثابت مشهود (پویا)", "برای سال مالی منتهی به 29 اسفند 1403", "(ارقام به میلیون ریال)")
    # Clear existing content to avoid duplicates on re-run if sheet already exists
    for row in ws.iter_rows(min_row=7): # Start clearing from row 7 to keep header
        for cell in row:
            cell.value = None

    headers = ["شرح", "مانده اول دوره", "افزایش (CAPEX)", "کاهش (استهلاک)", "مانده پایان دوره"]
    ws.append(headers)
    for cell in ws[4]: # Headers are at row 4
        cell.font = Font(bold=True)

    # محاسبات سال 1402
    ws['A5'] = "بهای تمام شده دارایی"
    ws['B5'] = "='ترازنامه پایه'!D17" # بهای تمام شده دارایی ابتدای 1402 از ترازنامه پایه (ناخالص)
    ws['B5'].font = Font(bold=True)
    ws['C5'] = f"='مفروضات'!{assumption_map['سرمایه‌گذاری ثابت سالانه (CAPEX)']['1402']}" # افزایش (CAPEX) از مفروضات
    ws['D5'] = 0 # فرض عدم فروش دارایی
    ws['E5'] = "=SUM(B5:D5)"

    ws['A6'] = "استهلاک انباشته"
    ws['B6'] = "='ترازنامه پایه'!D18" # استهلاک انباشته ابتدای 1402 از ترازنامه پایه (مقدار **مثبت** خوانده می‌شود)
    ws['B6'].font = Font(bold=True)
    ws['C6'] = 0
    ws['D6'] = f"=B5*'مفروضات'!{assumption_map['نرخ استهلاک سالانه (نسبت به بهای تمام شده اول دوره)']['1402']}"  
    ws['E6'] = "=B6+D6" # جمع اولیه + هزینه استهلاک

    ws.append([])
    ws['A8'] = "ارزش دفتری خالص"
    ws['A8'].font = Font(bold=True)
    ws['B8'] = "=B5-B6"
    ws['E8'] = "=E5-E6"

    # محاسبات سال 1403
    ws['A10'] = "بهای تمام شده دارایی"
    ws['B10'] = "=E5" # مانده اول دوره از پایان دوره سال قبل (1402)
    ws['C10'] = f"='مفروضات'!{assumption_map['سرمایه‌گذاری ثابت سالانه (CAPEX)']['1403']}" # افزایش (CAPEX) از مفروضات
    ws['D10'] = 0
    ws['E10'] = "=SUM(B10:D10)"

    ws['A11'] = "استهلاک انباشته"
    ws['B11'] = "=E6" # مانده اول دوره از پایان دوره سال قبل (1402)
    ws['C11'] = 0
    ws['D11'] = f"=B10*'مفروضات'!{assumption_map['نرخ استهلاک سالانه (نسبت به بهای تمام شده اول دوره)']['1403']}"  
    ws['E11'] = "=B11+D11"

    ws.append([])
    ws['A13'] = "ارزش دفتری خالص"
    ws['A13'].font = Font(bold=True)
    ws['B13'] = "=B10-B11"
    ws['E13'] = "=E10-E11"

    # اضافه کردن هایپرلینک برای بازگشت به شیت وضعیت مالی
    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"

# ==============================================================================
# تابع اصلاح شده ۶: populate_equity_sheet
# ==============================================================================
def populate_equity_sheet(ws):
    """پر کردن شیت حقوق مالکانه با ساختار استاندارد و فرمول‌های صحیح."""
    col_widths = {'A': 30, 'B': 18, 'C': 18, 'D': 18, 'E': 18, 'F': 20}
    set_rtl_and_column_widths(ws, col_widths)
    add_header(ws, "شرکت نمونه (سهامی عام)", "صورت تغییرات در حقوق مالکانه", "برای سال مالی منتهی به 29 اسفند 1403 و 1402", "(ارقام به میلیون ریال)")
    
    # Clear existing data
    for row in ws.iter_rows(min_row=7): # Start clearing from row 7 to keep header
        for cell in row:
            cell.value = None

    ws.append(["شرح", "سرمایه", "اندوخته قانونی", "سایر اندوخته‌ها", "سود انباشته", "جمع کل"])

    # Data for 1402 (starting from row 8)
    ws.cell(row=8, column=1, value="مانده در ابتدای 1402")
    ws.cell(row=8, column=2, value="='ترازنامه پایه'!D40") # سرمایه از ترازنامه پایه (ردیف D40 در ترازنامه پایه جدید)
    ws.cell(row=8, column=3, value="='ترازنامه پایه'!D41") # اندوخته قانونی از ترازنامه پایه (ردیف D41 در ترازنامه پایه جدید)
    ws.cell(row=8, column=4, value="='ترازنامه پایه'!D42") # سایر اندوخته‌ها از ترازنامه پایه (ردیف D42 در ترازنامه پایه جدید)
    ws.cell(row=8, column=5, value="='ترازنامه پایه'!D43") # سود انباشته از ترازنامه پایه (ردیف D43 در ترازنامه پایه جدید)
    ws.cell(row=8, column=6, value="=SUM(B8:E8)")

    ws.cell(row=9, column=1, value="سود خالص 1402")
    ws.cell(row=9, column=5, value="='سودوزیان'!G21") # لینک به سود خالص 1402 در سود و زیان (G21)
    ws.cell(row=9, column=6, value="=E9")

    ws.cell(row=10, column=1, value="انتقال به اندوخته قانونی")
    ws.cell(row=10, column=3, value="=MAX(0,'سودوزیان'!G21)*0.05") # MAX(0,...) ensures no negative reserve
    ws.cell(row=10, column=5, value="=-C10")
    ws.cell(row=10, column=6, value="0")

    ws.cell(row=11, column=1, value="تقسیم سود مصوب")
    # <<-- اصلاح شده: محاسبه سود سهام بر اساس درصد از سود خالص و مفروضات
    ws.cell(row=11, column=5, value="=-('سودوزیان'!G21*'مفروضات'!C19)")
    ws.cell(row=11, column=6, value="=E11")

    ws.cell(row=12, column=1, value="مانده در پایان 1402")
    ws.cell(row=12, column=2, value="=B8")
    ws.cell(row=12, column=3, value="=SUM(C8:C11)")
    ws.cell(row=12, column=4, value="=D8")
    ws.cell(row=12, column=5, value="=SUM(E8:E11)")
    ws.cell(row=12, column=6, value="=SUM(B12:E12)")
    
    ws.append([]) # Spacer (row 13)
    
    # Data for 1403 (starting from row 14)
    ws.cell(row=14, column=1, value="مانده در ابتدای 1403")
    ws.cell(row=14, column=2, value="=B12")
    ws.cell(row=14, column=3, value="=C12")
    ws.cell(row=14, column=4, value="=D12")
    ws.cell(row=14, column=5, value="=E12")
    ws.cell(row=14, column=6, value="=F12")

    ws.cell(row=15, column=1, value="سود خالص 1403")
    ws.cell(row=15, column=5, value="='سودوزیان'!F21") # لینک به سود خالص 1403 در سود و زیان (F21)
    ws.cell(row=15, column=6, value="=E15")

    ws.cell(row=16, column=1, value="انتقال به اندوخته قانونی")
    ws.cell(row=16, column=3, value="=MAX(0,'سودوزیان'!F21)*0.05") # MAX(0,...) ensures no negative reserve
    ws.cell(row=16, column=5, value="=-C16")
    ws.cell(row=16, column=6, value="0")

    ws.cell(row=17, column=1, value="تقسیم سود مصوب")
    # <<-- اصلاح شده: محاسبه سود سهام بر اساس درصد از سود خالص و مفروضات
    ws.cell(row=17, column=5, value="=-('سودوزیان'!F21*'مفروضات'!B19)")
    ws.cell(row=17, column=6, value="=E17")

    ws.cell(row=18, column=1, value="مانده در پایان 1403")
    ws.cell(row=18, column=2, value="=B14")
    ws.cell(row=18, column=3, value="=SUM(C14:C17)")
    ws.cell(row=18, column=4, value="=D14")
    ws.cell(row=18, column=5, value="=SUM(E14:E17)")
    ws.cell(row=18, column=6, value="=SUM(B18:E18)")

    # اضافه کردن هایپرلینک برای بازگشت به وضعیت مالی
    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"

# ==============================================================================
# تابع اصلاح شده ۷: populate_cash_flow_sheet
# ==============================================================================
def populate_cash_flow_sheet(ws, assumption_map):
    col_widths = {'A': 5, 'B': 55, 'C': 18, 'D': 18}
    set_rtl_and_column_widths(ws, col_widths)
    add_header(ws, "شرکت نمونه (سهامی عام)", "صورت جریان‌های نقدی (نهایی و پویا)", "برای سال مالی منتهی به 29 اسفند 1403", "(ارقام به میلیون ریال)")
    # Clear existing content to avoid duplicates on re-run if sheet already exists
    for row in ws.iter_rows(min_row=7): # Clear from row 7 to preserve header
        for cell in row:
            cell.value = None

    ws.append(["", "", "1403", "1402 (پایه)"]) # row 7

    # فعالیت‌های عملیاتی (starting from row 8)
    ws.cell(row=8, column=2, value="جریان‌های نقدی ناشی از فعالیت‌های عملیاتی")
    ws.cell(row=9, column=2, value="سود خالص")
    ws.cell(row=9, column=3, value="='سودوزیان'!F21") # لینک به سود خالص 1403
    ws.cell(row=9, column=4, value="='سودوزیان'!G21") # لینک به سود خالص 1402
    ws.cell(row=10, column=2, value="تعدیلات بابت اقلام غیرنقدی:")
    ws.cell(row=11, column=2, value="هزینه استهلاک")
    ws.cell(row=11, column=3, value="='گردش دارایی ثابت'!D11")
    ws.cell(row=11, column=4, value="='گردش دارایی ثابت'!D6")
    ws.cell(row=11, column=1, value="12") # Note 12 for Fixed Assets
    ws.cell(row=11, column=1).hyperlink = f"#'10.11.12'!A1" # Link to general fixed assets note
    ws.cell(row=11, column=1).style = "Hyperlink"

    ## <-- اصلاح نهایی: اضافه کردن هزینه غیرنقدی مزایای پایان خدمت
    ws.cell(row=12, column=2, value="هزینه مزایای پایان خدمت")
    ws.cell(row=12, column=3, value="='8'!C14") # Link to the expense from Note 8
    ws.cell(row=12, column=4, value="='8'!D14") # Link to the expense from Note 8
    
    ws.cell(row=13, column=2, value="تغییرات در سرمایه در گردش:")
    ws.cell(row=14, column=2, value="کاهش(افزایش) در دریافتنی‌ها")
    ws.cell(row=14, column=3, value="='وضعیت مالی'!F11-'وضعیت مالی'!E11") # F11: دریافتنی 1402, E11: دریافتنی 1403
    ws.cell(row=15, column=2, value="کاهش(افزایش) در موجودی کالا")
    ws.cell(row=15, column=3, value="='وضعیت مالی'!F12-'وضعیت مالی'!E12") # F12: موجودی کالا 1402, E12: موجودی کالا 1403
    ws.cell(row=16, column=2, value="افزایش(کاهش) در پرداختنی‌ها")
    ws.cell(row=16, column=3, value="='وضعیت مالی'!E25-'وضعیت مالی'!F25") # E25: پرداختنی 1403, F25: پرداختنی 1402
    ws.cell(row=17, column=2, value="خالص جریان نقد عملیاتی")
    ws.cell(row=17, column=3, value="=SUM(C9,C11,C12,C14:C16)") # C9: سود خالص, C11: استهلاک, C12: مزایای پایان خدمت, C14-C16: تغییرات سرمایه در گردش
    ws.cell(row=17, column=4, value="=SUM(D9,D11,D12,D14:D16)")

    # فعالیت‌های سرمایه‌گذاری (starting from row 17)
    ws.append([]) # row 18
    ws.cell(row=19, column=2, value="جریان‌های نقدی ناشی از فعالیت‌های سرمایه‌گذاری")
    ws.cell(row=20, column=2, value="پرداخت بابت خرید دارایی ثابت (CAPEX)")
    ws.cell(row=20, column=3, value=f"=-'مفروضات'!{assumption_map['سرمایه‌گذاری ثابت سالانه (CAPEX)']['1403']}") # لینک به CAPEX 1403
    ws.cell(row=20, column=4, value=f"=-'مفروضات'!{assumption_map['سرمایه‌گذاری ثابت سالانه (CAPEX)']['1402']}") # لینک به CAPEX 1402
    ws.cell(row=20, column=1, value="12") # Note 12 for Fixed Assets
    ws.cell(row=20, column=1).hyperlink = f"#'10.11.12'!A1" # Link to general fixed assets note
    ws.cell(row=20, column=1).style = "Hyperlink"

    ws.cell(row=21, column=2, value="خالص جریان نقد سرمایه‌گذاری")
    ws.cell(row=21, column=3, value="=C20")
    ws.cell(row=21, column=4, value="=D20")

    # فعالیت‌های تامین مالی (starting from row 21)
    ws.append([]) # row 22
    ws.cell(row=23, column=2, value="جریان‌های نقدی ناشی از فعالیت‌های تامین مالی")
    ws.cell(row=24, column=2, value="دریافت اصل تسهیلات")
    ws.cell(row=24, column=3, value=f"='مفروضات'!{assumption_map['مبلغ وام جدید دریافتی طی سال']['1403']}") # لینک به وام جدید 1403
    ws.cell(row=24, column=4, value=f"='مفروضات'!{assumption_map['مبلغ وام جدید دریافتی طی سال']['1402']}") # لینک به وام جدید 1402
    ws.cell(row=24, column=1, value="19") # Note 19 for LT Debt
    ws.cell(row=24, column=1).hyperlink = f"#'19'!A1"
    ws.cell(row=24, column=1).style = "Hyperlink"

    ws.cell(row=25, column=2, value="بازپرداخت اصل تسهیلات")
    ws.cell(row=25, column=3, value=f"=-'مفروضات'!{assumption_map['مبلغ بازپرداخت اصل وام طی سال']['1403']}") # لینک به بازپرداخت 1403
    ws.cell(row=25, column=4, value=f"=-'مفروضات'!{assumption_map['مبلغ بازپرداخت اصل وام طی سال']['1402']}") # لینک به بازپرداخت 1402
    ws.cell(row=26, column=2, value="سود سهام پرداخت شده")
    ws.cell(row=26, column=3, value=f"='حقوق مالکانه'!E17") # لینک به سود خالص و درصد سود سهام 1403 (E17 is already negative)
    ws.cell(row=26, column=4, value=f"='حقوق مالکانه'!E11") # لینک به سود خالص و درصد سود سهام 1402 (E11 is already negative)
    ws.cell(row=26, column=1, value="18") # Note 18 for Dividends
    ws.cell(row=26, column=1).hyperlink = f"#'18'!A1"
    ws.cell(row=26, column=1).style = "Hyperlink"

    ws.cell(row=27, column=2, value="خالص جریان نقد تامین مالی")
    ws.cell(row=27, column=3, value="=SUM(C24:C26)")
    ws.cell(row=27, column=4, value="=SUM(D24:D26)")

    # خلاصه (starting from row 27)
    ws.append([]) # row 28
    ws.cell(row=29, column=2, value="خالص افزایش (کاهش) در موجودی نقد")
    ws.cell(row=29, column=3, value="=C17+C21+C27") # جمع خالص جریان نقد عملیاتی, سرمایه گذاری, تامین مالی برای 1403
    ws.cell(row=29, column=4, value="=D17+D21+D27") # جمع خالص جریان نقد عملیاتی, سرمایه گذاری, تامین مالی برای 1402
    ws.cell(row=29, column=3).font = Font(bold=True)
    ws.cell(row=29, column=4).font = Font(bold=True)
    
    ws.cell(row=30, column=2, value="موجودی نقد ابتدای دوره")
    ws.cell(row=30, column=3, value="='وضعیت مالی'!F10")  # Existing cash balance from prior year
    ws.cell(row=30, column=4, value="='ترازنامه پایه'!D10") # موجودی نقد ابتدای 1402 از ترازنامه پایه
    ws.cell(row=31, column=2, value="موجودی نقد در پایان دوره")
    ws.cell(row=31, column=3, value="=C29+C30")  
    ws.cell(row=31, column=4, value="=D29+D30")  

    # اضافه کردن هایپرلینک برای بازگشت به وضعیت مالی
    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"


# --- بقیه توابع (بدون تغییر) ---

def populate_comprehensive_income_sheet(ws):
    col_widths = {'A': 5, 'B': 40, 'C': 18, 'D': 18}
    set_rtl_and_column_widths(ws, col_widths)
    add_header(ws, "شرکت نمونه (سهامی عام)", "صورت سود و زیان جامع", "برای سال مالی منتهی به 29 اسفند 1403 و 1402", "(ارقام به میلیون ریال)")
    # Clear existing content to avoid duplicates on re-run if sheet already exists
    for row in ws.iter_rows(min_row=7): # Start clearing from row 7 to keep header
        for cell in row:
            cell.value = None

    ws['C7'] = "1403"
    ws['D7'] = "1402"

    ws['B9'] = "سود خالص دوره"
    ws['C9'] = "='سودوزیان'!F21" # لینک به سود خالص از صورت سود و زیان 1403
    ws['D9'] = "='سودوزیان'!G21" # لینک به سود خالص از صورت سود و زیان 1402

    ws['B11'] = "سایر اقلام سود و زیان جامع:"
    ws['B12'] = "تعدیلات تسعیر ارز عملیات خارجی (بعد از مالیات)" # <--- اصلاح: از C12 به B12 منتقل شد
    ws['C12'] = 10_000 # مثال
    ws['D12'] = 5_000 # مثال

    ws['B14'] = "جمع کل سود و زیان جامع دوره" # A14 به B14 منتقل شد
    ws['C14'] = '=C9+C12'
    ws['D14'] = '=D9+D12'

    # اضافه کردن هایپرلینک برای بازگشت به سود و زیان
    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به سود و زیان").hyperlink = f"#'سودوزیان'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"


def populate_history_sheet(ws):
    ws.sheet_view.rightToLeft = True
    add_header(ws, "شرکت نمونه (سهامی عام)", "تاریخچه شرکت نمونه (سهامی عام)", "")
    # Clear existing content to avoid duplicates on re-run if sheet already exists
    for row in ws.iter_rows(min_row=7): # Start clearing from row 7 to keep header
        for cell in row:
            cell.value = None

    ws['A4'] = "مقدمه:"
    ws['B5'] = "شرکت نمونه (سهامی عام) در سال 1375 با هدف سرمایه‌گذاری و فعالیت در صنعت مرغداری و زنجیره تامین گوشت مرغ تاسیس گردید. این شرکت با بهره‌گیری از دانش روز و تکنولوژی‌های پیشرفته در زمینه پرورش جوجه یک روزه اجداد، تولید جوجه گوشتی و عرضه به کشتارگاه، به یکی از پیشگامان صنعت در کشور تبدیل شده است."
    ws['A7'] = "اهداف و استراتژی‌ها:"
    ws['B8'] = "هدف اصلی شرکت، تولید پروتئین با کیفیت بالا، افزایش بهره‌وری در تمامی مراحل زنجیره تامین، توسعه پایدار و ایفای نقش مسئولانه در تامین امنیت غذایی کشور است. استراتژی‌های شرکت شامل توسعه فارم‌های جدید، بهبود نژادهای پرورشی، بهینه‌سازی مصرف خوراک و کاهش ضایعات می‌باشد."
    ws['A10'] = "فعالیت‌های اصلی:"
    ws['B11'] = "شرکت در حال حاضر دارای 10 فارم پرورش مرغ گوشتی، 5 انبار نگهداری دان و مرغ، و واحد لجستیک پیشرفته برای حمل و نقل محصولات به کشتارگاه‌ها می‌باشد. ظرفیت تولید سالانه شرکت بیش از 50,000 تن مرغ گوشتی است."
    ws['A13'] = "چشم‌انداز آینده:"
    ws['B14'] = "شرکت نمونه با تمرکز بر نوآوری، توسعه بازار و افزایش ظرفیت‌های تولیدی، در نظر دارد سهم خود را در بازار افزایش داده و به عنوان یکی از بزرگترین شرکت‌های زنجیره تامین مرغ در منطقه شناخته شود."
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 80
    ws.merge_cells('A1:B1')
    ws['A4'].font = Font(bold=True)
    ws['A7'].font = Font(bold=True)
    ws['A10'].font = Font(bold=True)
    ws['A13'].font = Font(bold=True)
    for row in ws['B']: # wrap text in column B
        row.alignment = Alignment(wrapText=True, horizontal='right')
    
    # اضافه کردن هایپرلینک برای بازگشت به صفحه اصلی (مثلاً وضعیت مالی)
    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"


def populate_significant_accounting_policy_sheet(ws, policy_number):
    ws.sheet_view.rightToLeft = True
    add_header(ws, "شرکت نمونه (سهامی عام)", f"یادداشت {policy_number}: اهم رویه های حسابداری", "")
    # Clear existing content to avoid duplicates on re-run if sheet already exists
    for row in ws.iter_rows(min_row=7): # Start clearing from row 7 to keep header
        for cell in row:
            cell.value = None

    ws['A4'] = "مقدمه:"
    ws['B5'] = "صورت‌های مالی حاضر بر اساس استانداردهای حسابداری ایران (نشریه شماره 160 سازمان حسابرسی) تهیه شده‌اند. اهم رویه‌های حسابداری مورد استفاده در تهیه این صورت‌ها به شرح زیر است:"
    
    policies = {
        1: "نحوه ارائه صورت‌های مالی: صورت‌های مالی طبق استانداردهای حسابداری ایران و به روش بهای تمام شده تاریخی تهیه و ارائه شده‌اند. کلیه رویدادهای مالی در زمان وقوع شناسایی و ثبت می‌گردند. این صورت‌ها بر اساس اصل تداوم فعالیت تهیه شده‌اند.",
        2: "شناخت درآمد: درآمد حاصل از فروش مرغ گوشتی در زمان تحویل محصول و انتقال تمامی ریسک‌ها و مزایای مالکیت به خریدار شناسایی می‌شود. درآمدهای حاصل از فروش جوجه یک روزه نیز پس از تحویل و انتقال مالکیت و قطعیت وصول وجه شناسایی می‌گردد. درآمدهای فرعی (مانند فروش کود) نیز در زمان تحقق شناسایی می‌شوند.",
        3: "موجودی مواد و کالا: موجودی مواد و کالا (شامل جوجه، دان، دارو و مرغ آماده فروش) بر اساس روش میانگین موزون و به اقل بهای تمام شده یا خالص ارزش بازیافتنی اندازه‌گیری می‌شود. بهای تمام شده جوجه‌هاي در حال رشد شامل هزینه‌های مستقیم پرورش (دان، دارو، واکسن، دستمزد مستقیم کارگران فارم) و سهم مناسبی از سربار تولید است. خالص ارزش بازیافتنی، بهای فروش برآوردی در روال عادی عملیات پس از کسر هزینه‌های برآوردی تکمیل و هزینه‌های برآوردی لازم برای انجام فروش است. ذخیره کاهش ارزش موجودی‌ها در صورت لزوم شناسایی می‌گردد.",
        4: "دارایی‌های ثابت مشهود: دارایی‌های ثابت مشهود به بهای تمام شده تاریخی پس از کسر استهلاک انباشته و زیان کاهش ارزش انباشته در ترازنامه منعقد می‌شوند. استهلاک دارایی‌ها به روش خط مستقیم طی عمر مفید برآوردی دارایی صورت می‌گیرد. مخارج بعدی مربوط به دارایی‌های ثابت مشهود تنها در صورتی به بهای تمام شده دارایی اضافه می‌شود که منجر به افزایش قابل ملاحظه در منافع اقتصادی آتی ناشی از آن گردد. دارایی‌هایی که آماده بهره‌برداری نیستند، در حساب دارایی در جریان تکمیل ثبت می‌شوند.",
        5: "ارزهای خارجی: معاملات ارزی با نرخ تسعیر ارز در تاریخ معامله ثبت می‌شوند. اقلام پولی دارایی‌ها و بدهی‌های ارزی با نرخ تسعیر ارز در تاریخ ترازنامه تسعیر شده و سود یا زیان ناشی از تسعیر ارز به عنوان درآمد/هزینه غیرعملیاتی در صورت سود و زیان جامع شناسایی می‌شود. تفاوت‌های تسعیر ارز ناشی از اقلام پولی غیرپولی, در صورت‌های مالی منعکس نمی‌گردد.",
        6: "مزایای پایان خدمت کارکنان: تعهدات مزایای پایان خدمت کارکنان (پاداش پایان خدمت) بر اساس قوانین کار و تامین اجتماعی ایران و با استفاده از روش تعهدات برآوردی محاسبه و شناسایی می‌گردد."
    }
    
    policy_text = policies.get(policy_number, "توضیحات رویه حسابداری برای این یادداشت موجود نیست.")
    ws['A7'] = f"رویه حسابداری شماره {policy_number}:"
    ws['A7'].font = Font(bold=True)
    ws['B8'] = policy_text
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80
    ws['A4'].font = Font(bold=True)
    for row in ws['B']: # wrap text in column B
        row.alignment = Alignment(wrapText=True, horizontal='right')

    # اضافه کردن هایپرلینک برای بازگشت به صفحه اصلی (وضعیت مالی)
    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"

# ==============================================================================
# تابع (بدون تغییر): populate_inventory_note
# ==============================================================================
def populate_inventory_note(ws):
    ws.sheet_view.rightToLeft = True
    add_header(ws, "شرکت نمونه (سهامی عام)", "یادداشت 9: موجودی مواد و کالا (خلاصه)", "در تاریخ 29 اسفند 1403 و 1402", "(ارقام به میلیون ریال)")
    # Clear existing content to avoid duplicates on re-run if sheet already exists
    for row in ws.iter_rows(min_row=7): # Start clearing from row 7 to keep header
        for cell in row:
            cell.value = None

    ws['E7'] = "یادداشت"
    ws['F7'] = "1403"
    ws['G7'] = "1402"
    ws['F8'] = "میلیون ریال"
    ws['G8'] = "میلیون ریال"

    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"

    ws['A9'] = "ترکیب موجودی مواد و کالا:"
    # مقادیر ریالی در شیت تفصیلی از ردیف ۱۸ شروع می شوند
    ws['B10'] = "جوجه یک روزه (ریالی)"
    ws['F10'] = "='موجودی_تفصیلی'!G18"
    ws['G10'] = "='موجودی_تفصیلی'!K18"

    ws['B11'] = "خوراک (دان) (ریالی)"
    ws['F11'] = "='موجودی_تفصیلی'!G19"
    ws['G11'] = "='موجودی_تفصیلی'!K19"

    ws['B12'] = "مرغ در حال رشد (فارم) (ریالی)"
    ws['F12'] = "='موجودی_تفصیلی'!G20"
    ws['G12'] = "='موجودی_تفصیلی'!K20"

    ws['B13'] = "دارو و واکسن (ریالی)"
    ws['F13'] = "='موجودی_تفصیلی'!G21"
    ws['G13'] = "='موجودی_تفصیلی'!K21"

    ws['B14'] = "مرغ آماده فروش (انبار) (ریالی)"
    ws['F14'] = "='موجودی_تفصیلی'!G22"
    ws['G14'] = "='موجودی_تفصیلی'!K22"

    ws['B15'] = "سایر موجودی‌ها (لوازم بسته بندی و...)"
    ws['F15'] = 50000
    ws['G15'] = 50000

    ws['B16'] = "جمع کل موجودی مواد و کالا (پایان دوره)"
    ws['F16'] = f"='موجودی_تفصیلی'!G{18+5+4+3}" # لینک به خروجی نهایی برای ترازنامه
    ws['G16'] = f"='موجودی_تفصیلی'!K{18+5+4+4}" # لینک به خروجی نهایی برای ترازنامه
    ws.cell(row=16, column=2).hyperlink = f"#'موجودی_تفصیلی'!A1"
    ws.cell(row=16, column=2).style = "Hyperlink"

    total_row_value_detailed = 18 + 5 + 4
    ws['A18'] = "مغایرت‌گیری موجودی مواد و کالا (سال 1403):"
    ws['B19'] = "موجودی ابتدای دوره (1403)"
    ws['C19'] = f"='موجودی_تفصیلی'!D{total_row_value_detailed}"

    ws['B20'] = "خرید طی دوره (دان، جوجه، دارو، ...) و هزینه‌های مستقیم پرورش"
    ws['C20'] = f"='موجودی_تفصیلی'!E{total_row_value_detailed}"

    ws['B21'] = "بهای تمام شده کالای فروش رفته (COGS)"
    ws['C21'] = f"='موجودی_تفصیلی'!F{total_row_value_detailed}"

    ws['B22'] = "موجودی پایان دوره (1403 - محاسبه شده)"
    ws['C22'] = f"='موجودی_تفصیلی'!G{total_row_value_detailed}"

    ws['B23'] = "تفاوت مغایرت (اضافه/کسری)"
    ws['C23'] = f'=C19+C20-C21-C22'
    ws['C23'].font = Font(bold=True)
    ws.cell(row=23, column=3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


    ws['A25'] = "روش ارزیابی موجودی و اطلاعات انبار:"
    ws['B26'] = "موجودی مواد و کالا شرکت بر اساس روش میانگین موزون ارزیابی می‌شود و به اقل بهای تمام شده یا خالص ارزش بازیافتنی اندازه‌گیری می‌گردد..."
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

    for row in ws['B']: # wrap text
        row.alignment = Alignment(wrapText=True, horizontal='right')


def populate_management_judgment_sheet(ws):
    ws.sheet_view.rightToLeft = True
    add_header(ws, "شرکت نمونه (سهامی عام)", "یادداشت: قضاوت مدیریت در فرایند بکارگیری رویه های حسابداری", "")
    # Clear existing content to avoid duplicates on re-run if sheet already exists
    for row in ws.iter_rows(min_row=7): # Start clearing from row 7 to keep header
        for cell in row:
            cell.value = None

    ws['A4'] = "مقدمه:"
    ws['B5'] = "تهیه صورت‌های مالی بر اساس استانداردهای حسابداری ایران مستلزم اعمال قضاوت‌های مهم توسط مدیریت در بکارگیری رویه‌های حسابداری و برآوردهای حسابداری است. برخی از زمینه‌های کلیدی که مدیریت در آن‌ها قضاوت‌های مهمی اعمال می‌کند، به شرح زیر است:"
    
    ws['A7'] = "1. برآورد عمر مفید دارایی‌های ثابت مشهود و نامشهود:"
    ws['B8'] = "مدیریت عمر مفید اقتصادی دارایی‌های ثابت مشهود (نظیر ساختمان فارم‌ها، تجهیزات و ماشین‌آلات) و نامشهود (نظیر نرم‌افزارها و حقوق استفاده از نژادهای خاص) را بر اساس تجربه قبلی، انتظارات از فرسودگی فیزیکی و منسوخ شدن تکنولوژیکی برآورد می‌کند. هرگونه تغییر در این برآوردها می‌تواند بر مبلغ استهلاک و بهای تمام شده دارایی‌ها در دوره‌های آتی تاثیر بگذارد. تجدید ارزیابی دارایی‌ها بر اساس رویه شرکت انجام می‌شود."
    
    ws['A10'] = "2. خالص ارزش بازیافتنی موجودی مواد و کالا:"
    ws['B11'] = "مدیریت برای تعیین خالص ارزش بازیافتنی موجودی‌ها (از جمله جوجه، دان و مرغ آماده فروش)، قضاوت‌هایی در خصوص قیمت‌هاي فروش آتی، هزینه‌های تکمیل و هزینه‌های لازم برای انجام فروش اعمال می‌کند. این برآوردها تحت تاثیر شرایط بازار، نوسانات قیمت خوراک و دارو و میزان تقاضا برای محصولات شرکت قرار دارد و ممکن است منجر به شناسایی ذخیره کاهش ارزش موجودی‌ها شود."

    ws['A13'] = "3. ذخیره مطالبات مشکوک‌الوصول:"
    ws['B14'] = "مدیریت بر اساس سابقه تاریخ‌ياب وصول مطالبات، وضعیت مالی مشتریان و شرایط اقتصادی جاری، برآوردی از مبلغ مطالبات مشکوک‌الوصول را انجام می‌دهد. این برآورد شامل قضاوت در خصوص میزان عدم قطعیت در وصول مطالبات آتی است. شناسایی این ذخیره بر اساس اصل احتياط و قابلیت وصول مطالبات انجام می‌شود."

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 80
    ws['A4'].font = Font(bold=True)
    ws['A7'].font = Font(bold=True)
    ws['A10'].font = Font(bold=True)
    ws['A13'].font = Font(bold=True)
    for row in ws['B']: # wrap text
        row.alignment = Alignment(wrapText=True, horizontal='right')

    # اضافه کردن هایپرلینک برای بازگشت به صفحه اصلی (مثلاً وضعیت مالی)
    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"


def populate_attachment_sheet(ws):
    ws.sheet_view.rightToLeft = True
    add_header(ws, "شرکت نمونه (سهامی عام)", "پیوست صورت‌های مالی", "")
    # Clear existing content to avoid duplicates on re-run if sheet already exists
    for row in ws.iter_rows(min_row=7): # Start clearing from row 7 to keep header
        for cell in row:
            cell.value = None

    ws['A4'] = "این بخش شامل هرگونه اطلاعات تکمیلی و جداول تفصیلی است که برای درک کامل‌تر صورت‌های مالی ضروری است."
    ws['A6'] = "مثال: جدول تفصیلی دارایی‌های ثابت مشهود، جداول تفصیلی سرمایه‌گذاری‌ها، تفکیک درآمدها بر حسب نوع محصول و منطقه جغرافیایی، گزارش کامل حقوق و دستمزد تفکیکی."
    ws.column_dimensions['A'].width = 80
    ws['A4'].font = Font(bold=True)
    for row in ws['A']: # wrap text
        row.alignment = Alignment(wrapText=True, horizontal='right')

    # اضافه کردن هایپرلینک برای بازگشت به صفحه اصلی (مثلاً وضعیت مالی)
    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"


def populate_page_header_sheet(ws):
    ws.sheet_view.rightToLeft = True
    add_header(ws, "شرکت نمونه (سهامی عام)", "سر برگ صفحات (برای چاپ و ارائه)", "")
    # Clear existing content to avoid duplicates on re-run if sheet already exists
    for row in ws.iter_rows(min_row=7): # Start clearing from row 7 to keep header
        for cell in row:
            cell.value = None

    ws['A3'] = "این شیت می‌تواند شامل اطلاعات تکراری در بالای هر صفحه چاپی باشد."
    ws['A5'] = "نام شرکت: شرکت نمونه (سهامی عام)"
    ws['A6'] = "صورت مالی: صورت سود و زیان / صورت وضعیت مالی و غیره"
    ws['A7'] = "سال مالی: منتهی به 29 اسفند 1403"
    ws.column_dimensions['A'].width = 80
    ws['A3'].font = Font(italic=True)
    for row in ws['A']: # wrap text
        row.alignment = Alignment(wrapText=True, horizontal='right')

    # اضافه کردن هایپرلینک برای بازگشت به صفحه اصلی (مثلاً وضعیت مالی)
    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"


def populate_signature_sheet(ws):
    ws.sheet_view.rightToLeft = True
    add_header(ws, "شرکت نمونه (سهامی عام)", "صفحه امضا کنندگان صورت‌های مالی", "")
    # Clear existing content to avoid duplicates on re-run if sheet already exists
    for row in ws.iter_rows(min_row=7): # Start clearing from row 7 to keep header
        for cell in row:
            cell.value = None

    ws['A3'] = "این صورت‌های مالی توسط افراد زیر تهیه و تأیید شده‌اند:"
    ws['A5'] = "نام و نام خانوادگی: [نام مدیر عامل]"
    ws['A6'] = "سمت: مدیر عامل"
    ws['A7'] = "تاریخ: 1404/03/22"
    ws['A9'] = "نام و نام خانوادگی: [نام مدیر مالی]"
    ws['A10'] = "سمت: مدیر مالی"
    ws['A11'] = "تاریخ: 1404/03/22"
    ws['A13'] = "نام و نام خانوادگی: [نام حسابرس]"
    ws['A14'] = "سمت: حسابرس مستقل"
    ws['A15'] = "تاریخ: 1404/03/22"
    ws.column_dimensions['A'].width = 40
    ws['A3'].font = Font(bold=True)
    
    # اضافه کردن هایپرلینک برای بازگشت به صفحه اصلی (مثلاً وضعیت مالی)
    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"


def populate_management_comparative_report(ws):
    ws.sheet_view.rightToLeft = True
    add_header(ws, "شرکت نمونه (سهامی عام)", "گزارش مدیریتی تطبیقی", "برای سال مالی منتهی به 29 اسفند 1403 و 1402", "(ارقام به میلیون ریال)")
    # Clear existing content to avoid duplicates on re-run if sheet already exists
    for row in ws.iter_rows(min_row=7): # Start clearing from row 7 to keep header
        for cell in row:
            cell.value = None

    ws['C7'] = "1403"
    ws['D7'] = "1402"
    ws['E7'] = "درصد تغییر"

    ws['A9'] = "خلاصه‌ای از شاخص‌های کلیدی عملکرد (KPIs):"
    
    ws['B10'] = "درآمدهای عملیاتی"
    ws['C10'] = "='سودوزیان'!F8" # لینک به درآمد عملیاتی 1403 (F8)
    ws['D10'] = "='سودوزیان'!G8" # لینک به درآمد عملیاتی 1402 (G8)
    ws['E10'] = '=IF(D10<>0,(C10-D10)/D10,"N/A")'
    ws.cell(row=10, column=5).number_format = '0.00%'

    ws['B11'] = "سود ناخالص"
    ws['C11'] = "='سودوزیان'!F10" # لینک به سود ناخالص 1403 (F10)
    ws['D11'] = "='سودوزیان'!G10" # لینک به سود ناخالص 1402 (G10)
    ws['E11'] = '=IF(D11<>0,(C11-D11)/D11,"N/A")'
    ws.cell(row=11, column=5).number_format = '0.00%'

    ws['B12'] = "سود عملیاتی"
    ws['C12'] = "='سودوزیان'!F15" # لینک به سود عملیاتی 1403 (F15)
    ws['D12'] = "='سودوزیان'!G15" # لینک به سود عملیاتی 1402 (G15)
    ws['E12'] = '=IF(D12<>0,(C12-D12)/D12,"N/A")'
    ws.cell(row=12, column=5).number_format = '0.00%'

    ws['B13'] = "سود خالص"
    ws['C13'] = "='سودوزیان'!F21" # لینک به سود خالص 1403 (F21)
    ws['D13'] = "='سودوزیان'!G21" # لینک به سود خالص 1402 (G21)
    ws['E13'] = '=IF(D13<>0,(C13-D13)/D13,"N/A")'
    ws.cell(row=13, column=5).number_format = '0.00%'

    ws['B15'] = "جمع کل دارایی‌ها"
    ws['C15'] = "='وضعیت مالی'!E21"
    ws['D15'] = "='وضعیت مالی'!F21"
    ws['E15'] = '=IF(D15<>0,(C15-D15)/D15,"N/A")'
    ws.cell(row=15, column=5).number_format = '0.00%'

    ws['B16'] = "جمع کل بدهی‌ها"
    ws['C16'] = "='وضعیت مالی'!E35"
    ws['D16'] = "='وضعیت مالی'!F35"
    ws['E16'] = '=IF(D16<>0,(C16-D16)/D16,"N/A")'
    ws.cell(row=16, column=5).number_format = '0.00%'
    
    ws['A18'] = "نسبت‌های مالی کلیدی:"
    ws['B19'] = "نسبت جاری (Current Ratio)"
    ws['C19'] = "=IFERROR('وضعیت مالی'!E14/'وضعیت مالی'!E29,0)"
    ws['D19'] = "=IFERROR('وضعیت مالی'!F14/'وضعیت مالی'!F29,0)"
    ws.cell(row=19, column=3).number_format = '0.00'  
    ws.cell(row=19, column=4).number_format = '0.00'  

    ws['B20'] = "نسبت بدهی (Debt Ratio)"
    ws['C20'] = "=IFERROR('وضعیت مالی'!E35/'وضعیت مالی'!E21,0)"
    ws['D20'] = "=IFERROR('وضعیت مالی'!F35/'وضعیت مالی'!F21,0)"
    ws.cell(row=20, column=3).number_format = '0.00'
    ws.cell(row=20, column=4).number_format = '0.00'
    
    ws['B21'] = "حاشیه سود خالص (Net Profit Margin)"
    ws['C21'] = "=IFERROR('سودوزیان'!F21/'سودوزیان'!F8,0)" # لینک به سود خالص و درآمد عملیاتی
    ws['D21'] = "=IFERROR('سودوزیان'!G21/'سودوزیان'!G8,0)" # لینک به سود خالص و درآمد عملیاتی
    ws.cell(row=21, column=3).number_format = '0.00%'
    ws.cell(row=21, column=4).number_format = '0.00%'

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws['A9'].font = Font(bold=True)
    ws['A18'].font = Font(bold=True)

    # اضافه کردن هایپرلینک برای بازگشت به صفحه اصلی (مثلاً وضعیت مالی)
    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"


def populate_business_analytical_report(ws):
    ws.sheet_view.rightToLeft = True
    add_header(ws, "شرکت نمونه (سهامی عام)", "گزارش تحلیلی کسب و کار", "برای سال مالی منتهی به 29 اسفند 1403", "")
    # Clear existing content to avoid duplicates on re-run if sheet already exists
    for row in ws.iter_rows(min_row=7): # Start clearing from row 7 to keep header
        for cell in row:
            cell.value = None

    ws['A5'] = "1. تحلیل عملکرد عملیاتی:"
    ws['B6'] = '''=CONCATENATE("شرکت در سال 1403 شاهد رشد ",TEXT('گزارش مدیریتی تطبیقی'!E10,"0.00%")," درآمدهای عملیاتی نسبت به سال قبل بوده است. این رشد عمدتاً ناشی از افزایش ظرفیت تولید و تقاضا در بازار مرغ گوشتی می‌باشد. با این حال، بهای تمام شده درآمدهای عملیاتی نیز ",TEXT(IFERROR(('سودوزیان'!F9/'سودوزیان'!G9)-1,"0.00%"),"0.00%")," افزایش یافته که نیاز به کنترل بیشتر هزینه‌ها در زنجیره تامین دارد.")'''
    ws.cell(row=6, column=2).alignment = Alignment(wrapText=True, horizontal='right')
    ws.cell(row=6, column=2).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  

    ws['A8'] = "2. تحلیل سودآوری:"
    ws['B9'] = '''=CONCATENATE("حاشیه سود خالص شرکت در سال 1403 به ",TEXT('گزارش مدیریتی تطبیقی'!C21,"0.00%")," رسیده که نشان‌دهنده توانایی شرکت در مدیریت هزینه‌های مستقیم تولید است. با این حال، هزینه‌های اداری و عمومی نیز رشد قابل توجهی داشته‌اند که می‌بایست مورد بررسی قرار گیرند.")'''
    ws.cell(row=9, column=2).alignment = Alignment(wrapText=True, horizontal='right')
    ws.cell(row=9, column=2).fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  

    ws['A11'] = "3. تحلیل وضعیت نقدینگی:"
    ws['B12'] = '''=CONCATENATE("جریان‌های نقدی عملیاتی شرکت مثبت بوده که نشان‌دهنده توانایی شرکت در تامین نقدینگی از محل عملیات اصلی خود است. نسبت جاری شرکت در سال 1403 برابر با ",TEXT('گزارش مدیریتی تطبیقی'!C19,"0.00")," است که نشان‌دهنده وضعیت نقدینگی مطلوب و توانایی ایفای تعهدات جاری است.")'''  
    ws.cell(row=12, column=2).alignment = Alignment(wrapText=True, horizontal='right')
    ws.cell(row=12, column=2).fill = PatternFill(start_color="CCE0F5", end_color="CCE0F5", fill_type="solid")  

    ws['A14'] = "4. پیشنهادها:"
    ws['B15'] = "- بررسی دقیق‌تر هزینه‌های اداری و عمومی و شناسایی فرصت‌های صرفه‌جویی.\n- سرمایه‌گذاری در تکنولوژی‌های جدید برای افزایش بهره‌وری در فارم‌ها و کاهش بهای تمام شده تولید.\n- توسعه بازارهای جدید برای محصولات شرکت."
    ws.cell(row=15, column=2).alignment = Alignment(wrapText=True, horizontal='right')
    ws.cell(row=15, column=2).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80
    ws['A5'].font = Font(bold=True)
    ws['A8'].font = Font(bold=True)
    ws['A11'].font = Font(bold=True)
    ws['A14'].font = Font(bold=True)
    
    # اضافه کردن هایپرلینک برای بازگشت به صفحه اصلی (مثلاً وضعیت مالی)
    ws.cell(row=1, column=max(1, ws.max_column - 1), value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"


def populate_numeric_note_sheets(wb):
    """
    پر کردن شیت های با نام عددی بر اساس شماره یادداشت آنها در صورت های مالی.
    """
    numeric_sheet_map = {
        # شیت '5': درآمدهای عملیاتی تفکیکی
        '5': {
            'header_name': "یادداشت 5: درآمدهای عملیاتی",
            'data': [
                ("فروش مرغ گوشتی", 2_000_000, 1_700_000),
                ("فروش جوجه یک روزه", 300_000, 250_000),
                ("فروش کود و سایر محصولات جانبی", 200_000, 150_000)
            ],
            'total_row_text': "جمع کل درآمدهای عملیاتی",
            'total_row_formula_1403': '=SUM(F10:F12)',
            'total_row_formula_1402': '=SUM(G10:G12)',
            'return_sheet': 'سودوزیان'
        },
        # شیت '6': موجودی نقد و معادل‌های آن
        '6': {
            'header_name': "یادداشت 6: موجودی نقد و معادل‌های آن",
            'data': [
                ("وجوه نقد در صندوق و بانک", 1_000_000, 700_000),
                ("سپرده‌های کوتاه مدت بانکی", 200_000, 200_000)
            ],
            'total_row_text': "جمع کل موجودی نقد و معادل‌های آن",
            'total_row_formula_1403': '=SUM(F10:F11)',
            'total_row_formula_1402': '=SUM(G10:G11)',
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '7': سرمایه‌گذاری‌های کوتاه مدت
        '7': {
            'header_name': "یادداشت 7: سرمایه‌گذاری‌های کوتاه مدت",
            'data': [
                ("سرمایه‌گذاری در اوراق بهادار با درآمد ثابت", 250_000, 200_000),
                ("سایر سرمایه‌گذاری‌های کوتاه مدت", 50_000, 50_000)
            ],
            'total_row_text': "جمع کل سرمایه‌گذاری‌های کوتاه مدت",
            'total_row_formula_1403': '=SUM(F10:F11)',
            'total_row_formula_1402': '=SUM(G10:G11)',
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '10.11.12': سرمایه‌گذاری‌های بلندمدت و دارایی ثابت
        '10.11.12': {
            'header_name': "یادداشت 10-11-12: پیش پرداخت ها و سرمایه‌گذاری‌های بلندمدت و دارایی‌های ثابت مشهود",
            'sections': [
                {"title": "الف) پیش پرداخت‌ها (یادداشت 10)", "data": [
                    ("پیش پرداخت خرید مواد و کالا", 50_000, 40_000),
                    ("پیش پرداخت هزینه‌ها", 50_000, 40_000)
                ], "total_text": "جمع پیش پرداخت‌ها"},
                {"title": "ب) سرمایه‌گذاری‌های بلندمدت (یادداشت 11)", "data": [
                    ("سرمایه‌گذاری در شرکت‌های فرعی", 150_000, 100_000),
                    ("سپرده‌های بلندمدت بانکی", 50_000, 50_000)
                ], "total_text": "جمع سرمایه‌گذاری‌های بلندمدت"},
                {"title": "ج) دارایی‌های ثابت مشهود (یادداشت 12)", "data": [
                    ("زمین", 500_000, 500_000),
                    ("ساختمان و تاسیسات فارم‌ها", 1_500_000, 1_300_000),
                    ("ماشین‌آلات و تجهیزات", 1_000_000, 1_000_000)
                ], "total_text": "جمع دارایی‌های ثابت مشهود"}
            ],
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '13': دارایی‌های نامشهود و سایر دارایی‌ها
        '13': {
            'header_name': "یادداشت 13: دارایی‌های نامشهود و سایر دارایی‌ها",
            'sections': [
                {"title": "الف) دارایی‌های نامشهود", "data": [
                    ("نرم‌افزارها و سیستم‌های اطلاعاتی", 30_000, 25_000),
                    ("حقوق بهره‌برداری", 20_000, 15_000)
                ], "total_text": "جمع دارایی‌های نامشهود"},
                {"title": "ب) سایر دارایی‌ها", "data": [
                    ("سپرده‌های تضمینی", 20_000, 15_000),
                    ("دارایی‌های در انتظار فروش", 10_000, 10_000)
                ], "total_text": "جمع سایر دارایی‌ها"}
            ],
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '14': عملیات متوقف شده (جزئیات)
        '14': {
            'header_name': "یادداشت 14: جزئیات عملیات متوقف شده",
            'data': [
                ("سود/زیان ناشی از فعالیت‌های عملیات متوقف شده", 0, 0),
                ("سود/زیان ناشی از واگذاری دارایی‌های عملیات متوقف شده", 0, 0)
            ],
            'total_row_text': "جمع کل سود (زیان) عملیات متوقف شده",
            'total_row_formula_1403': '=SUM(F10:F11)',
            'total_row_formula_1402': '=SUM(G10:G11)',
            'notes': ["در این دوره مالی، شرکت عملیات متوقف شده‌ای نداشته است."],
            'return_sheet': 'سودوزیان'
        },
        # شیت '15': حساب‌ها و اسناد پرداختنی تجاری و سایر پرداختنی‌ها
        '15': {
            'header_name': "یادداشت 15: حساب‌ها و اسناد پرداختنی تجاری و سایر پرداختنی‌ها",
            'data': [
                ("حساب‌های پرداختنی بابت خرید دان و جوجه", 400_000, 350_000),
                ("پرداختنی بابت حقوق و دستمزد", 50_000, 40_000),
                ("سایر پرداختنی‌ها", 150_000, 110_000)
            ],
            'total_row_text': "جمع کل حساب‌ها و اسناد پرداختنی تجاری و سایر پرداختنی‌ها",
            'total_row_formula_1403': '=SUM(F10:F12)',
            'total_row_formula_1402': '=SUM(G10:G12)',
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '16': تسهیلات مالی
        '16': {
            'header_name': "یادداشت 16: تسهیلات مالی",
            'data': [
                ("تسهیلات دریافتی از بانک (کوتاه مدت)", 250_000, 200_000),
                ("وام‌های دریافتی از اشخاص (کوتاه مدت)", 150_000, 150_000)
            ],
            'total_row_text': "جمع کل تسهیلات مالی جاری",
            'total_row_formula_1403': '=SUM(F10:F11)',
            'total_row_formula_1402': '=SUM(G10:G11)',
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '17': مالیات پرداختنی
        '17': {
            'header_name': "یادداشت 17: مالیات پرداختنی",
            'data': [
                ("مالیات عملکرد پرداختنی", 30_000, 20_000),
                ("مالیات بر ارزش افزوده پرداختنی", 20_000, 20_000)
            ],
            'total_row_text': "جمع کل مالیات پرداختنی",
            'total_row_formula_1403': '=SUM(F10:F11)',
            'total_row_formula_1402': '=SUM(G10:G11)',
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '18': سود سهام پرداختنی
        '18': {
            'header_name': "یادداشت 18: سود سهام پرداختنی",
            'data': [
                ("سود سهام مصوب مجمع (سال 1403)", 80_000, 0),
                ("سود سهام مصوب مجمع (سال 1402)", 0, 70_000)
            ],
            'total_row_text': "جمع کل سود سهام پرداختنی",
            'total_row_formula_1403': '=SUM(F10:F11)',
            'total_row_formula_1402': '=SUM(G10:G11)',
            'notes': ["سود سهام مصوب مجمع عمومی عادی سالانه و در مهلت قانونی قابل پرداخت است."],
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '19': تسهیلات مالی بلندمدت
        '19': {
            'header_name': "یادداشت 19: تسهیلات مالی بلندمدت",
            'data': [
                ("تسهیلات بلندمدت از بانک کشاورزی", 500_000, 400_000),
                ("وام بلندمدت از صندوق توسعه ملی", 200_000, 200_000)
            ],
            'total_row_text': "جمع کل تسهیلات مالی بلندمدت",
            'total_row_formula_1403': '=SUM(F10:F11)',
            'total_row_formula_1402': '=SUM(G10:G11)',
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '20': مزایای پایان خدمت کارکنان
        '20': {
            'header_name': "یادداشت 20: مزایای پایان خدمت کارکنان",
            'data': [
                ("ذخیره پاداش پایان خدمت", 150_000, 130_000)
            ],
            'total_row_text': "جمع کل مزایای پایان خدمت کارکنان",
            'total_row_formula_1403': '=SUM(F10:F10)',
            'total_row_formula_1402': '=SUM(G10:G10)',
            'notes': ["تعهدات شرکت بابت مزایای پایان خدمت کارکنان بر اساس قوانین جاری کار و محاسبات مربوطه شناسایی شده است."],
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '21': سرمایه
        '21': {
            'header_name': "یادداشت 21: سرمایه",
            'data': [
                ("سرمایه ثبت شده (1,000,000 سهم عادی 1,000 ریالی)", 1_000_000, 1_000_000)
            ],
            'total_row_text': "جمع کل سرمایه",
            'total_row_formula_1403': '=SUM(F10:F10)',
            'total_row_formula_1402': '=SUM(G10:G10)',
            'notes': ["سرمایه شرکت در سال‌های مورد گزارش تغییری نداشته است."],
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '22.-23': اندوخته ها
        '22.-23': {
            'header_name': "یادداشت 22-23: اندوخته قانونی و سایر اندوخته‌ها",
            'sections': [
                {"title": "الف) اندوخته قانونی (یادداشت 22)", "data": [
                    ("مانده ابتدای دوره", 150_000, 100_000),
                    ("انتقال از سود انباشته", 50_000, 50_000)
                ], "total_text": "مانده پایان دوره اندوخته قانونی"},
                {"title": "ب) سایر اندوخته‌ها (یادداشت 23)", "data": [
                    ("مانده ابتدای دوره", 80_000, 50_000),
                    ("انتقال از سود انباشته (توسعه و بازسازی)", 20_000, 30_000)
                ], "total_text": "مانده پایان دوره سایر اندوخته‌ها"}
            ],
            'return_sheet': 'حقوق مالکانه'
        },
        # شیت '24.25': سود انباشته
        '24.25': {
            'header_name': "یادداشت 24-25: سود انباشته",
            'data': [
                ("مانده ابتدای دوره", "='حقوق مالکانه'!E12", "='حقوق مالکانه'!E8"), # لینک به مانده های حقوق مالکانه
                ("سود خالص دوره", "='سودوزیان'!F21", "='سودوزیان'!G21"), # لینک به سود خالص
                ("انتقال به اندوخته قانونی", "='حقوق مالکانه'!C16*-1", "='حقوق مالکانه'!C10*-1"),
                ("انتقال به سایر اندوخته‌ها", 0, 0), # فرض شده
                ("تقسیم سود سهام", "='حقوق مالکانه'!E17", "='حقوق مالکانه'!E11")
            ],
            'total_row_text': "مانده پایان دوره سود انباشته",
            'total_row_formula_1403': '=SUM(F10:F14)',
            'total_row_formula_1402': '=SUM(G10:G14)',
            'return_sheet': 'حقوق مالکانه'
        },
        # شیت '26.27': درآمدهای مالی و سایر درآمدها و هزینه‌های غیرعملیاتی
        '26.27': {
            'header_name': "یادداشت 26-27: درآمدهای مالی و سایر درآمدها و هزینه‌های غیرعملیاتی",
            'sections': [
                {"title": "الف) هزینه‌های مالی (یادداشت 26)", "data": [
                    ("سود تسهیلات بانکی", -60_000, -50_000),
                    ("کارمزد بانکی", -10_000, -10_000)
                ], "total_text": "جمع هزینه‌های مالی"},
                {"title": "ب) سایر درآمدها و هزینه‌های غیرعملیاتی (یادداشت 27)", "data": [
                    ("سود حاصل از تسعیر ارز", 15_000, 10_000),
                    ("سود حاصل از فروش دارایی‌های ثابت", 15_000, 15_000),
                    ("زیان‌های متفرقه", -5_000, -5_000)
                ], "total_text": "جمع سایر درآمدها و هزینه‌های غیرعملیاتی"}
            ],
            'return_sheet': 'سودوزیان'
        },
        # شیت '28.29.30.31': بدهی های جاری و غیرجاری تفصیلی
        '28.29.30.31': {
            'header_name': "یادداشت 28-31: جزئیات بدهی‌ها",
            'sections': [
                {"title": "الف) حساب‌ها و اسناد پرداختنی تجاری (یادداشت 28)", "data": [
                    ("بابت خرید مواد اولیه (دان، دارو)", 300_000, 250_000),
                    ("بابت خرید جوجه", 100_000, 80_000)
                ], "total_text": "جمع پرداختنی‌های تجاری"},
                {"title": "ب) سایر پرداختنی‌ها (یادداشت 29)", "data": [
                    ("حقوق و دستمزد پرداختنی", 50_000, 40_000),
                    ("مالیات پرداختنی", 50_000, 40_000)
                ], "total_text": "جمع سایر پرداختنی‌ها"},
                {"title": "ج) تسهیلات مالی جاری (یادداشت 30)", "data": [
                    ("تسهیلات کوتاه‌مدت بانکی", 250_000, 200_000),
                    ("سایر تسهیلات کوتاه‌مدت", 150_000, 150_000)
                ], "total_text": "جمع تسهیلات مالی جاری"},
                {"title": "د) تسهیلات مالی بلندمدت (یادداشت 31)", "data": [
                    ("تسهیلات بانکی بلندمدت", 500_000, 400_000),
                    ("وام از سایر منابع", 200_000, 200_000)
                ], "total_text": "جمع تسهیلات مالی بلندمدت"}
            ],
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '32.33': هزینه‌های فروش و اداری تفصیلی
        '32.33': {
            'header_name': "یادداشت 32-33: هزینه‌های فروش و اداری",
            'sections': [
                {"title": "الف) هزینه‌های فروش (یادداشت 32)", "data": [
                    ("حقوق و دستمزد فروش", "='8'!C10", "='8'!D10"), # لینک به یادداشت 8
                    ("تبلیغات و بازاریابی", 50_000, 40_000),
                    ("حمل و نقل و توزیع", 150_000, 130_000)
                ], "total_text": "جمع هزینه‌های فروش"},
                {"title": "ب) هزینه‌های اداری و عمومی (یادداشت 33)", "data": [
                    ("حقوق و دستمزد اداری", "='8'!C16", "='8'!D16"), # لینک به یادداشت 8
                    ("اجاره", 20_000, 18_000),
                    ("استهلاک", "='8'!C18", "='8'!D18"), # لینک به یادداشت 8
                    ("خدمات", 10_000, 9_000),
                    ("سایر", 10_000, 15_000)
                ], "total_text": "جمع هزینه‌های اداری و عمومی"}
            ],
            'return_sheet': 'سودوزیان'
        },
        # شیت '34': مالیات
        '34': {
            'header_name': "یادداشت 34: مالیات بر درآمد",
            'data': [
                ("مالیات عملکرد سال جاری", 50_000, 40_000),
                ("مالیات سال‌های قبل", 10_000, 5_000),
                ("مالیات معوق", 0, 0) # مثال
            ],
            'total_row_text': "جمع کل مالیات بر درآمد شناسایی شده",
            'total_row_formula_1403': '=SUM(F10:F12)',
            'total_row_formula_1402': '=SUM(G10:G12)',
            'return_sheet': 'سودوزیان'
        },
        # شیت '35': سرمایه‌گذاری‌ها (تفصیلی)
        '35': {
            'header_name': "یادداشت 35: سرمایه‌گذاری‌ها",
            'sections': [
                {"title": "الف) سرمایه‌گذاری‌هاي کوتاه مدت", "data": [
                    ("اوراق قرضه دولتی", 150_000, 100_000),
                    ("صندوق سرمایه‌گذاری", 150_000, 150_000)
                ], "total_text": "جمع سرمایه‌گذاری‌های کوتاه مدت"},
                {"title": "ب) سرمایايه‌گذاری‌هاي بلندمدت", "data": [
                    ("سهام شرکت‌های غیر بورسی", 100_000, 80_000),
                    ("سرمایه گذاری در املاک", 100_000, 70_000)
                ], "total_text": "جمع سرمایه‌گذاری‌های بلندمدت"}
            ],
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '35-1': دارایی‌های نگهداری شده برای فروش
        '35-1': {
            'header_name': "یادداشت 35-1: دارایی‌های نگهداری شده برای فروش",
            'data': [
                ("مرغداری غیر فعال آماده فروش", 50_000, 50_000),
                ("تجهیزات مازاد", 10_000, 0)
            ], 'total_row_text': "جمع کل", 'total_row_formula_1403': '=SUM(F10:F11)', 'total_row_formula_1402': '=SUM(G10:G11)',
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '35-6': حقوق و دستمزد (خلاصه)
        '35-6': {
            'header_name': "یادداشت 35-6: خلاصه حقوق و دستمزد",
            'data': [
                ("مجموع حقوق و مزایای پرداختی به پرسنل", "='لیست حقوق و دستمزد'!N108", "='لیست حقوق و دستمزد'!N108*0.9"),  
                ("بیمه سهم کارفرما", "='لیست حقوق و دستمزد'!V108", "='لیست حقوق و دستمزد'!V108*0.9")  
            ],
            'total_row_text': "جمع",
            'total_row_formula_1403': '=SUM(F10:F11)',
            'total_row_formula_1402': '=SUM(G10:G11)',
            'notes': ["این شیت خلاصه ای از اطلاعات حقوق و دستمزد در شیت 'لیست حقوق و دستمزد' می باشد."],
            'return_sheet': 'سودوزیان'
        },
        # شیت '36-37': مالیات بر درآمد تفصیلی (مالیات)
        '36-37': {
            'header_name': "یادداشت 36-37: مالیات بر درآمد",
            'data': [
                ("مالیات بر درآمد سال جاری (مالیات عملکرد)", 50_000, 40_000),
                ("مالیات سال‌های قبل و جرایم", 10_000, 5_000),
                ("تاثیر تفاوت های موقت", 0, 0) # مثال
            ],
            'total_row_text': "جمع کل مالیات بر درآمد شناسایی شده",
            'total_row_formula_1403': '=SUM(F10:F12)',
            'total_row_formula_1402': '=SUM(G10:G12)',
            'return_sheet': 'سودوزیان'
        },
        # شیت '38.39.40': اطلاعات مربوط به سهام
        '38.39.40': {
            'header_name': "یادداشت 38-40: اطلاعات مربوط به سهام",
            'sections': [
                {"title": "الف) تعداد سهام عادی (یادداشت 38)", "data": [
                    ("تعداد سهام در گردش (عدد)", 1_000_000, 1_000_000)
                ], "total_text": "جمع"},
                {"title": "ب) ارزش اسمی هر سهم (یادداشت 39)", "data": [
                    ("ارزش اسمی هر سهم (ریال)", 1_000, 1_000)
                ], "total_text": "جمع"},
                {"title": "ج) سود پایه هر سهم (یادداشت 40)", "data": [
                    ("سود پایه هر سهم (ریال)", "='سودوزیان'!F21/1000000", "='سودوزیان'!G21/1000000")  
                ], "total_text": "جمع"}
            ],
            'return_sheet': 'سودوزیان' # EPS is derived from P&L
        },
        # شیت '41': موجودی نقد و سرمایه گذاری کوتاه مدت (تفصیلی)
        '41': {
            'header_name': "یادداشت 41: موجودی نقد و سرمایه‌گذاری‌های کوتاه مدت (تفصیلی)",
            'sections': [
                {"title": "الف) موجودی نقد و معادل‌های آن", "data": [
                    ("موجودی نقد در بانک مرکزی", 500_000, 400_000),
                    ("موجودی نقد در بانک‌های تجاری", 400_000, 300_000),
                    ("وجوه در راه", 300_000, 200_000)
                ], "total_text": "جمع موجودی نقد و معادل‌های آن"},
                {"title": "ب) سرمایايه‌گذاری‌های کوتاه مدت", "data": [
                    ("اوراق مشارکت", 200_000, 150_000),
                    ("سپرده‌های کوتاه‌مدت", 100_000, 100_000)
                ], "total_text": "جمع سرمایه‌گذاری‌های کوتاه مدت"}
            ],
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '42.43': حساب و اسناد دریافتنی
        '42.43': {
            'header_name': "یادداشت 42-43: حساب‌ها و اسناد دریافتنی",
            'sections': [
                {"title": "الف) حساب‌ها و اسناد دریافتنی تجاری (یادداشت 42)", "data": [
                    ("مشتریان عمده", 500_000, 450_000),
                    ("سایر مشتریان", 200_000, 150_000)
                ], "total_text": "جمع حساب‌ها و اسناد دریافتنی تجاری"},
                {"title": "ب) سایر دریافتنی‌ها (یادداشت 43)", "data": [
                    ("مطالبات از کارکنان", 10_000, 5_000),
                    ("سایر", 20_000, 15_000)
                ], "total_text": "جمع سایر دریافتنی‌ها"}
            ],
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '44' تا '44-6': دارایی ثابت مشهود (جزئیات)
        '44': {
            'header_name': "یادداشت 44: دارایی‌های ثابت مشهود (زمین)",
            'data': [("زمین", 500_000, 500_000)], 'total_row_text': "جمع", 'total_row_formula_1403': '=SUM(F10:F10)', 'total_row_formula_1402': '=SUM(G10:G10)',
            'return_sheet': 'وضعیت مالی'
        },
        '44-4': {
            'header_name': "یادداشت 44-4: دارایی‌های ثابت مشهود (ساختمان و تاسیسات فارم‌ها)",
            'data': [
                ("بهای تمام شده", 1_800_000, 1_500_000),
                ("استهلاک انباشته", -300_000, -200_000)
            ], 'total_row_text': "ارزش دفتری", 'total_row_formula_1403': '=SUM(F10:F11)', 'total_row_formula_1402': '=SUM(G10:G11)',
            'return_sheet': 'وضعیت مالی'
        },
        '44-6': {
            'header_name': "یادداشت 44-6: دارایی‌های ثابت مشهود (ماشین‌آلات و تجهیزات)",
            'data': [
                ("بهای تمام شده", 1_200_000, 1_100_000),
                ("استهلاک انباشته", -200_000, -100_000)
            ], 'total_row_text': "ارزش دفتری", 'total_row_formula_1403': '=SUM(F10:F11)', 'total_row_formula_1402': '=SUM(G10:G11)',
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '45': پیش پرداخت ها
        '45': {
            'header_name': "یادداشت 45: پیش پرداخت‌ها (تفصیلی)",
            'data': [
                ("پیش پرداخت خرید مواد اولیه", 50_000, 40_000),
                ("پیش پرداخت بیمه", 20_000, 15_000),
                ("پیش پرداخت اجاره", 30_000, 25_000)
            ], 'total_row_text': "جمع کل پیش پرداخت‌ها", 'total_row_formula_1403': '=SUM(F10:F12)', 'total_row_formula_1402': '=SUM(G10:G12)',
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '46': حساب ها و اسناد پرداختنی (تفصیلی)
        '46': {
            'header_name': "یادداشت 46: حساب‌ها و اسناد پرداختنی تجاری (تفصیلی)",
            'data': [
                ("فروشندگان دان", 250_000, 200_000),
                ("فروشندگان جوجه", 100_000, 80_000),
                ("فروشندگان دارو و واکسن", 50_000, 40_000)
            ], 'total_row_text': "جمع کل", 'total_row_formula_1403': '=SUM(F10:F12)', 'total_row_formula_1402': '=SUM(G10:G12)',
            'return_sheet': 'وضعیت مالی'
        },
        '46-3': {
            'header_name': "یادداشت 46-3: سایر پرداختنی‌ها (تفصیلی)",
            'data': [
                ("حقوق و دستمزد پرداختنی", 50_000, 40_000),
                ("مالیات پرداختنی", 50_000, 40_000),
                ("بیمه پرداختنی", 20_000, 15_000)
            ], 'total_row_text': "جمع کل", 'total_row_formula_1403': '=SUM(F10:F12)', 'total_row_formula_1402': '=SUM(G10:G12)',
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '47.48': تسهیلات مالی تفصیلی (اصلاح شده)
        '47.48': {
            'header_name': "یادداشت 47-48: تسهیلات مالی (تفصیلی)",
            'sections': [
                {"title": "الف) تسهیلات مالی جاری", "data": [
                    ("وام کوتاه‌مدت بانک ملی", 250_000, 200_000),
                    ("خط اعتباری بانک ملت", 150_000, 150_000)
                ], "total_text": "جمع تسهیلات مالی جاری"},
                {"title": "ب) تسهیلات مالی بلندمدت", "data": [
                    ("وام بلندمدت بانک کشاورزی", 1_200_000, 400_000), # <-- افزایش وام
                    ("وام از صندوق توسعه ملی", 200_000, 200_000)
                ], "total_text": "جمع تسهیلات مالی بلندمدت"}
            ],
            'return_sheet': 'وضعیت مالی'
        },
        # شیت '49': مزایای پایان خدمت تفصیلی
        '49': {
            'header_name': "یادداشت 49: مزایای پایان خدمت کارکنان (تفصیلی)",
            'data': [
                ("مانده ابتدای دوره", 130_000, 100_000),
                ("ذخیره شناسایی شده طی دوره", 20_000, 30_000),
                ("پرداخت شده طی دوره", 0, 0)
            ], 'total_row_text': "مانده پایان دوره", 'total_row_formula_1403': '=SUM(F10:F12)', 'total_row_formula_1402': '=SUM(G10:G12)',
            'return_sheet': 'وضعیت مالی'
        },
        # شیت 'ادامه16': ادامه تسهیلات مالی
        'ادامه16': {
            'header_name': "ادامه یادداشت 16: تسهیلات مالی",
            'data': [
                ("تسهیلات دریافتی از سازمان‌های دولتی (کوتاه مدت)", 50_000, 30_000),
                ("اعتبار اسنادی", 20_000, 10_000)
            ], 'total_row_text': "جمع", 'total_row_formula_1403': '=SUM(F10:F11)', 'total_row_formula_1402': '=SUM(G10:G11)',
            'notes': ["این شیت شامل جزئیات بیشتر تسهیلات مالی جاری است."],
            'return_sheet': '16' # Return to main Note 16
        },
        # شیت 'ادامه34': ادامه مالیات
        'ادامه34': {
            'header_name': "ادامه یادداشت 34: مالیات بر درآمد",
            'data': [
                ("مالیات بر درآمد معوق (دارایی)", 0, 0),
                ("مالیات بر درآمد معوق (بدهی)", 0, 0)
            ], 'total_row_text': "جمع", 'total_row_formula_1403': '=SUM(F10:F11)', 'total_row_formula_1402': '=SUM(G10:G11)',
            'notes': ["این شیت شامل جزئیات بیشتر مالیات است."],
            'return_sheet': '34' # Return to main Note 34
        },
        # شیت 'ادامه41': ادامه موجودی نقد و سرمایه گذاری کوتاه مدت
        'ادامه41': {
            'header_name': "ادامه یادداشت 41: موجودی نقد و سرمایه‌گذاری‌های کوتاه مدت",
            'sections': [
                {"title": "الف) اسناد تجاری و تعهدات نقدی", "data": [
                    ("اسناد تجاری دریافتنی", 50_000, 40_000),
                    ("تعهدات نقدی کوتاه مدت", -20_000, -10_000)
                ], "total_text": "جمع"},
                {"title": "ب) اوراق مشارکت بورسی و سایر اوراق", "data": [
                    ("اوراق مشارکت بورسی", 50_000, 30_000),
                    ("سایر اوراق بهادار", 10_000, 5_000)
                ], "total_text": "جمع"}
            ],
            'notes': ["این شیت شامل جزئیات بیشتر موجودی نقد و سرمایه‌گذاری‌های کوتاه مدت است."],
            'return_sheet': '41' # Return to main Note 41
        },
        'ادامه41..': {
            'header_name': "ادامه 2 یادداشت 41: موجودی نقد و سرمایه‌گذاری‌های کوتاه مدت",
            'data': [
                ("وجوه در حال وصول", 20_000, 10_000),
                ("اسناد تضمینی", 10_000, 5_000)
            ], 'total_row_text': "جمع", 'total_row_formula_1403': '=SUM(F10:F11)', 'total_row_formula_1402': '=SUM(G10:G11)',
            'notes': ["این شیت شامل جزئیات بیشتر موجودی نقد و سرمایه‌گذاری‌های کوتاه مدت است."],
            'return_sheet': '41' # Return to main Note 41
        },
    }

    for sheet_name, content in numeric_sheet_map.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Clear existing content to avoid duplicates on re-run if sheet already exists
            for row in ws.iter_rows(min_row=7): # Start clearing from row 7 to keep header
                for cell in row:
                    cell.value = None

            set_rtl_and_column_widths(ws, {'A': 5, 'B': 35, 'C': 25, 'D': 12, 'E': 18, 'F': 18, 'G': 18})
            add_header(ws, "شرکت نمونه (سهامی عام)", content['header_name'], f"برای سال مالی منتهی به 29 اسفند 1403 و 1402", "(ارقام به میلیون ریال)")
            ws['E7'] = "یادداشت"
            ws['F7'] = "1403"
            ws['G7'] = "1402"
            ws['F8'] = "میلیون ریال"
            ws['G8'] = "میلیون ریال"

            current_row = 10
            if 'data' in content:
                for item_name, val_1403, val_1402 in content['data']:
                    ws[f'B{current_row}'] = item_name
                    # Ensure formulas are properly set with '='
                    ws[f'F{current_row}'] = f"={val_1403}" if isinstance(val_1403, str) and val_1403.startswith('=') else val_1403
                    ws[f'G{current_row}'] = f"={val_1402}" if isinstance(val_1402, str) and val_1402.startswith('=') else val_1402
                    current_row += 1
                
                ws[f'B{current_row}'] = content['total_row_text']
                ws[f'F{current_row}'] = content['total_row_formula_1403']
                ws[f'G{current_row}'] = content['total_row_formula_1402']
                current_row += 2
                
                if 'notes' in content:
                    for note in content['notes']:
                        ws[f'B{current_row}'] = note
                        ws[f'B{current_row}'].alignment = Alignment(wrapText=True, horizontal='right')
                        current_row += 1

            elif 'sections' in content:
                for section in content['sections']:
                    ws[f'A{current_row}'] = section['title']
                    ws[f'A{current_row}'].font = Font(bold=True)
                    current_row += 1
                    start_data_row_for_sum = current_row
                    for item_name, val_1403, val_1402 in section['data']:
                        ws[f'B{current_row}'] = item_name
                        # Ensure formulas are properly set with '='
                        ws[f'F{current_row}'] = f"={val_1403}" if isinstance(val_1403, str) and val_1403.startswith('=') else val_1403
                        ws[f'G{current_row}'] = f"={val_1402}" if isinstance(val_1402, str) and val_1402.startswith('=') else val_1402
                        current_row += 1
                    ws[f'B{current_row}'] = section['total_text']
                    
                    ws[f'F{current_row}'] = f'=SUM(F{start_data_row_for_sum}:F{current_row-1})'
                    ws[f'G{current_row}'] = f'=SUM(G{start_data_row_for_sum}:G{current_row-1})'
                    
                    current_row += 2

            for row_idx in range(1, ws.max_row + 1):
                if ws[f'B{row_idx}'].value:
                    ws[f'B{row_idx}'].alignment = Alignment(wrapText=True, horizontal='right')
                if ws[f'C{row_idx}'].value:
                    ws[f'C{row_idx}'].alignment = Alignment(wrapText=True, horizontal='right')
            
            # اضافه کردن هایپرلینک برای بازگشت به شیت اصلی (تعریف شده در map)
            if 'return_sheet' in content:
                return_sheet_name = content['return_sheet']
                ws.cell(row=1, column=max(1, ws.max_column - 1), value=f"بازگشت به {return_sheet_name}").hyperlink = f"#'{return_sheet_name}'!A1"
                ws.cell(row=1, column=max(1, ws.max_column - 1)).style = "Hyperlink"


# ==============================================================================
# تابع اصلاح شده ۸: create_full_financial_report (با ترتیب اجرای جدید و حلقه تکرار)
# ==============================================================================
def create_full_financial_report(output_folder, output_file_name):
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    all_sheet_names = [
        'مفروضات', 'ترازنامه پایه', 'وضعیت مالی', 'سودوزیان', 'جریان های نقدی', 'حقوق مالکانه', 'جامع',
        'گردش دارایی ثابت', 'موجودی_تفصیلی', 'موجودی', 'لیست حقوق و دستمزد', '8', '9',
        'سر برگ صفحات', 'ص امضا', 'تاریخچه',
        'اهم رویه1', 'اهم رویه2', 'اهم رویه3', 'اهم رویه4', 'اهم رویه5', 'اهم رویه6',
        'قضاوت مدیریت', 'پیوست',
        '5', '6', '7', '10.11.12', '13', '14', '15', '16', '17', '18', '19', '20',
        '21', '22.-23', '24.25', '26.27', '28.29.30.31', '32.33', '34',
        '35', '35-1', '35-6', '36-37', '38.39.40', '41', '42.43', '44',
        '44-4', '44-6', '45', '46', '46-3', '47.48', '49',
        'ادامه16', 'ادامه34', 'ادامه41', 'ادامه41..',
        'گزارش مدیریتی تطبیقی',
        'گزارش تحلیلی کسب و کار'
    ]

    for sheet_name in all_sheet_names:
        if sheet_name not in wb.sheetnames:
            try:
                wb.create_sheet(sheet_name)
            except ValueError:
                pass

    # --- ترتیب صحیح اجرا برای مدل یکپارچه ---
    print("شروع ساخت مدل مالی یکپارچه...")
    
    # 1. پر کردن شیت مفروضات
    assumption_map = populate_assumptions_sheet(wb['مفروضات'])

    # 2. پر کردن شیت ترازنامه پایه (جدید) - باید قبل از تمامی شیت‌هایی که به مانده‌های اولیه نیاز دارند، پر شود.
    populate_starting_balance_sheet(wb['ترازنامه پایه'])

    # 3. پر کردن شیت‌های پایه و تفصیلی که نیاز به ترازنامه پایه ندارند یا ورودی اولیه هستند.
    populate_payroll_list_sheet(wb['لیست حقوق و دستمزد'])
    populate_detailed_inventory_sheet(wb['موجودی_تفصیلی'])
    
    # 4. ایجاد یادداشت‌های هزینه‌ای که ورودی سود و زیان هستند (اکنون یکپارچه شده‌اند)
    populate_note_8_and_9(wb) # این تابع هم 8 و هم 9 را پر می‌کند

    # 5. اجرای حلقه محاسبه تراز برای همگرایی (مهمترین بخش)
    # این حلقه باید شیت‌های اصلی (سودوزیان، حقوق مالکانه، گردش دارایی ثابت، وضعیت مالی، جریان‌های نقدی) را چندین بار پر کند تا مقادیر همگرا شوند.
    print("اجرای حلقه محاسبه تراز...")
    for i in range(20): # تکرار 20 بار برای اطمینان از همگرایی کامل
        print(f"اجرای حلقه همگرایی - مرحله {i+1}")
        populate_profit_loss_sheet(wb['سودوزیان'], assumption_map) # سودوزیان
        populate_equity_sheet(wb['حقوق مالکانه']) # حقوق مالکانه
        populate_fixed_asset_roll_forward_sheet(wb['گردش دارایی ثابت'], assumption_map) # گردش دارایی ثابت
        populate_balance_sheet(wb['وضعیت مالی'], assumption_map) # وضعیت مالی
        populate_cash_flow_sheet(wb['جریان های نقدی'], assumption_map) # جریان های نقدی

    print("مدل مالی پویا با موفقیت ایجاد و تراز شد.")

    # 6. پر کردن بقیه شیت‌ها
    print("پر کردن شیت‌های جانبی و یادداشت‌ها...")
    populate_inventory_note(wb['موجودی'])
    populate_numeric_note_sheets(wb)
    populate_comprehensive_income_sheet(wb['جامع'])
    populate_history_sheet(wb['تاریخچه'])
    for i in range(1, 7):
        sheet_name = f'اهم رویه{i}'
        if sheet_name in wb.sheetnames:
            populate_significant_accounting_policy_sheet(wb[sheet_name], i)
    populate_management_judgment_sheet(wb['قضاوت مدیریت'])
    populate_attachment_sheet(wb['پیوست'])
    populate_page_header_sheet(wb['سر برگ صفحات'])
    populate_signature_sheet(wb['ص امضا'])
    populate_management_comparative_report(wb['گزارش مدیریتی تطبیقی'])
    populate_business_analytical_report(wb['گزارش تحلیلی کسب و کار'])
    print("تمام شیت‌ها پر شدند.")


    output_path = os.path.join(output_folder, output_file_name)
    try:
        wb.active = wb['وضعیت مالی']
        wb.save(output_path)
        print(f"فایل اکسل '{output_file_name}' با موفقیت در مسیر '{output_folder}' ایجاد و پر شد.")
    except Exception as e:
        print(f"خطا در ذخیره فایل اکسل: {e}")
        print("لطفاً مطمئن شوید فایل اکسل با همین نام باز نیست و دسترسی نوشتن به پوشه مقصد وجود دارد.")


# --- تابع اصلی برای اجرا ---
if __name__ == "__main__":
    # <<!>> مسیر پوشه خروجی خود را در اینجا وارد کنید
    # مثال: r'C:\Users\mahinn\Desktop\New folder (14)'
    # لطفاً 'YourUser' را با نام کاربری خود در ویندوز جایگزین کنید
    output_directory = r'C:\Users\mahinn\Desktop\New folder (14)'  
    output_excel_file_name = 'صورت_مالی_کامل_پویا_مرغداری_تراز_شده_نهایی.xlsx'

    # اطمینان از وجود پوشه خروجی
    if not os.path.exists(output_directory):
        try:
            os.makedirs(output_directory)
            print(f"پوشه '{output_directory}' ایجاد شد.")
        except OSError as e:
            print(f"خطا در ایجاد پوشه: {e}. لطفاً یک مسیر معتبر وارد کنید.")
            exit()

    # اجرای تابع برای ایجاد و پر کردن صورت های مالی
    create_full_financial_report(output_directory, output_excel_file_name)
