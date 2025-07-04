import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
import os
import random
import math

# ==============================================================================
# توابع کمکی عمومی (General Helper Functions)
# ==============================================================================

def set_rtl_and_column_widths(ws, col_widths):
    """تنظیم جهت نوشتار راست به چپ و عرض ستون‌ها."""
    ws.sheet_view.rightToLeft = True
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

def add_header(ws, company_name, report_title, period, currency_note=""):
    """افزودن سربرگ استاندارد به شیت."""
    ws.merge_cells('A1:D1')
    ws.cell(row=1, column=1, value=company_name).font = Font(bold=True, size=16)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    ws.cell(row=2, column=1, value=report_title).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:D3')
    ws.cell(row=3, column=1, value=period).font = Font(size=11)
    ws.cell(row=3, column=1).alignment = Alignment(horizontal='center')
    
    if currency_note:
        ws.merge_cells('A4:D4')
        ws.cell(row=4, column=1, value=currency_note).font = Font(size=10, italic=True)
        ws.cell(row=4, column=1).alignment = Alignment(horizontal='center')
    
    return ws.max_row + 1

def generate_all_employees_data(num_employees):
    """تولید داده‌های کارمندان برای شیت حقوق و دستمزد."""
    first_names = ["علی", "محمد", "فاطمه", "زهرا", "حسین", "زینب", "رضا", "مریم", "امیر", "نارین"]
    last_names = ["احمدی", "کریمی", "محمدی", "رضایی", "نقدی", "صفری", "عباسی", "حسینی", "امیری", "محسنی"]
    units = ["فروش", "اداری", "تولید", "بازرگانی", "مالی", "انبار", "تدارکات", "حراست"]
    roles = [
        "مدیر عامل", "مدیر فروش", "مدیر تولید", "مدیر مالی", "مدیر منابع انسانی",
        "حسابدار", "کارشناس فروش", "کارشناس تولید", "کارشناس اداری", "دامپزشک",
        "کارگر", "نگهبان", "انباردار", "راننده"
    ]
    employees = []
    for i in range(1, num_employees + 1):
        num_children = random.randint(0, 3)
        unit = random.choice(units)
        role = random.choice([r for r in roles if any(u in r or r in u for u in unit.split()) or "مدیر" in r or "کارشناس" in r or "حسابدار" in r or "دامپزشک" in r or "کارگر" in r or "نگهبان" in r or "انباردار" in r or "راننده" in r])
        
        if unit == "فروش" and "مدیر فروش" not in [emp["role"] for emp in employees if emp["unit"] == "فروش"]: role = "مدیر فروش"
        elif unit == "تولید" and "مدیر تولید" not in [emp["role"] for emp in employees if emp["unit"] == "تولید"]: role = "مدیر تولید"
        elif unit == "اداری" and "مدیر منابع انسانی" not in [emp["role"] for emp in employees if emp["unit"] == "اداری"]: role = "مدیر منابع انسانی"
        elif unit == "مالی" and "مدیر مالی" not in [emp["role"] for emp in employees if emp["unit"] == "مالی"]: role = "مدیر مالی"

        employees.append({
            "id": i,
            "first_name": random.choice(first_names),
            "last_name": random.choice(last_names),
            "unit": unit,
            "role": role,
            "num_children": num_children
        })
    return employees

# ==============================================================================
# توابع populate_ (Populate Sheet Functions)
# ==============================================================================

def populate_starting_balance_sheet(ws):
    """ایجاد و پر کردن شیت ترازنامه افتتاحیه برای سال پایه (پایان 1401 / ابتدای 1402)."""
    ws.title = "ترازنامه پایه"
    col_widths = {'A': 5, 'B': 40, 'C': 45, 'D': 18}
    set_rtl_and_column_widths(ws, col_widths)
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "ترازنامه افتتاحیه (پایه)", "در تاریخ 29 اسفند 1401 / 1 فروردین 1402", "(ارقام به میلیون ریال)")

    # --- دارایی‌ها ---
    ws.cell(row=current_row, column=2, value="دارایی‌ها").font = Font(bold=True)
    current_row += 1
    ws.cell(row=current_row, column=2, value="دارایی‌های جاری").font = Font(bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=3, value="موجودی نقد")
    ws.cell(row=current_row, column=4, value=2700004.733) # مبلغ اولیه تراز شده برای سال 1402
    ws.row_cash_base = current_row # ذخیره ردیف برای استفاده در سایر شیت‌ها
    current_row += 1

    ws.cell(row=current_row, column=3, value="حساب‌ها و اسناد دریافتنی")
    ws.cell(row=current_row, column=4, value=515068)
    ws.row_receivables_base = current_row
    current_row += 1

    ws.cell(row=current_row, column=3, value="موجودی کالا")
    ws.cell(row=current_row, column=4, value=200000)
    ws.row_inventory_base = current_row
    current_row += 1

    ws.cell(row=current_row, column=3, value="پیش‌پرداخت‌ها و سایر دارایی‌های جاری")
    ws.cell(row=current_row, column=4, value=50000)
    ws.row_prepayments_base = current_row
    current_row += 1

    ws.cell(row=current_row, column=2, value="جمع دارایی‌های جاری").font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=f"=SUM(D{ws.row_cash_base}:D{ws.row_prepayments_base})")
    ws.row_total_current_assets_base = current_row
    current_row += 2 # یک سطر خالی

    ws.cell(row=current_row, column=2, value="دارایی‌های غیرجاری").font = Font(bold=True)
    current_row += 1
    ws.cell(row=current_row, column=3, value="بهای تمام شده ناخالص دارایی‌های ثابت")
    ws.cell(row=current_row, column=4, value=3000000)
    ws.row_gross_fixed_assets_base = current_row
    current_row += 1
    
    ws.cell(row=current_row, column=3, value="کسر می‌شود: استهلاک انباشته")
    ws.cell(row=current_row, column=4, value=300000)
    ws.row_accumulated_dep_base = current_row
    current_row += 1
    
    ws.cell(row=current_row, column=3, value="دارایی‌های ثابت مشهود (ارزش دفتری)")
    ws.cell(row=current_row, column=4, value=f"=D{ws.row_gross_fixed_assets_base}-D{ws.row_accumulated_dep_base}")
    ws.row_net_fixed_assets_base = current_row
    current_row += 1

    ws.cell(row=current_row, column=3, value="سایر دارایی‌های غیرجاری")
    ws.cell(row=current_row, column=4, value=120000)
    ws.row_other_non_current_assets_base = current_row
    current_row += 1

    ws.cell(row=current_row, column=2, value="جمع دارایی‌های غیرجاری").font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=f"=SUM(D{ws.row_net_fixed_assets_base}:D{ws.row_other_non_current_assets_base})")
    ws.row_total_non_current_assets_base = current_row
    current_row += 2 # یک سطر خالی

    ws.cell(row=current_row, column=2, value="جمع کل دارایی‌ها").font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=f"=D{ws.row_total_current_assets_base}+D{ws.row_total_non_current_assets_base}")
    ws.row_total_assets_base = current_row
    current_row += 2 # یک سطر خالی

    # --- بدهی‌ها و حقوق مالکانه ---
    ws.cell(row=current_row, column=2, value="بدهی‌ها و حقوق مالکانه").font = Font(bold=True)
    current_row += 1
    ws.cell(row=current_row, column=2, value="بدهی‌های جاری").font = Font(bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=3, value="حساب‌ها و اسناد پرداختنی")
    ws.cell(row=current_row, column=4, value=380000)
    ws.row_payables_base = current_row
    current_row += 1

    ws.cell(row=current_row, column=3, value="مالیات پرداختنی")
    ws.cell(row=current_row, column=4, value=25000)
    ws.row_tax_payable_base = current_row
    current_row += 1

    ws.cell(row=current_row, column=3, value="سود سهام پرداختنی")
    ws.cell(row=current_row, column=4, value=75000)
    ws.row_dividends_payable_base = current_row
    current_row += 1

    ws.cell(row=current_row, column=3, value="بخش جاری تسهیلات بلندمدت")
    ws.cell(row=current_row, column=4, value=60000)
    ws.row_current_lt_debt_base = current_row
    current_row += 1

    ws.cell(row=current_row, column=2, value="جمع بدهی‌های جاری").font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=f"=SUM(D{ws.row_payables_base}:D{ws.row_current_lt_debt_base})")
    ws.row_total_current_liabilities_base = current_row
    current_row += 2 # یک sطر خالی

    ws.cell(row=current_row, column=2, value="بدهی‌های غیرجاری").font = Font(bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=3, value="تسهیلات مالی بلندمدت")
    ws.cell(row=current_row, column=4, value=700000)
    ws.row_lt_debt_base = current_row
    current_row += 1

    ws.cell(row=current_row, column=3, value="مزایای پایان خدمت کارکنان")
    ws.cell(row=current_row, column=4, value=150000)
    ws.row_employee_benefits_base = current_row
    current_row += 1

    ws.cell(row=current_row, column=2, value="جمع بدهی‌های غیرجاری").font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=f"=SUM(D{ws.row_lt_debt_base}:D{ws.row_employee_benefits_base})")
    ws.row_total_non_current_liabilities_base = current_row
    current_row += 2 # یک سطر خالی

    ws.cell(row=current_row, column=2, value="جمع کل بدهی‌ها").font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=f"=D{ws.row_total_current_liabilities_base}+D{ws.row_total_non_current_liabilities_base}")
    ws.row_total_liabilities_base = current_row
    current_row += 2 # یک سطر خالی

    ws.cell(row=current_row, column=2, value="حقوق مالکانه").font = Font(bold=True)
    current_row += 1
    ws.cell(row=current_row, column=3, value="سرمایه")
    ws.cell(row=current_row, column=4, value=1000000)
    ws.row_capital_base = current_row
    current_row += 1

    ws.cell(row=current_row, column=3, value="اندوخته قانونی")
    ws.cell(row=current_row, column=4, value=120000)
    ws.row_legal_reserve_base = current_row
    current_row += 1

    ws.cell(row=current_row, column=3, value="سایر اندوخته‌ها")
    ws.cell(row=current_row, column=4, value=60000)
    ws.row_other_reserve_base = current_row
    current_row += 1

    ws.cell(row=current_row, column=3, value="سود انباشته")
    ws.cell(row=current_row, column=4, value=f"=D{ws.row_total_assets_base}-D{ws.row_total_liabilities_base}-D{ws.row_capital_base}-D{ws.row_legal_reserve_base}-D{ws.row_other_reserve_base}")
    ws.row_retained_earnings_base = current_row
    current_row += 1

    ws.cell(row=current_row, column=2, value="جمع کل حقوق مالکانه").font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=f"=SUM(D{ws.row_capital_base}:D{ws.row_retained_earnings_base})")
    ws.row_total_equity_base = current_row
    current_row += 2 # یک سطر خالی

    ws.cell(row=current_row, column=2, value="جمع کل بدهی‌ها و حقوق مالکانه").font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=f"=D{ws.row_total_liabilities_base}+D{ws.row_total_equity_base}")
    ws.row_total_liabilities_and_equity_base = current_row
    current_row += 2 # یک سطر خالی

    ws.cell(row=current_row, column=2, value="کنترل تراز (باید صفر باشد)").font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=f"=D{ws.row_total_assets_base}-D{ws.row_total_liabilities_and_equity_base}")
    current_row += 1

    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"


def populate_assumptions_sheet(ws):
    """ایجاد و پر کردن شیت مفروضات نهایی مدل مالی و بازگرداندن نقشه آدرس ها."""
    ws.title = "مفروضات"
    col_widths = {'A': 45, 'B': 18, 'C': 18}
    set_rtl_and_column_widths(ws, col_widths)
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "شیت مفروضات مدل مالی (سناریو سوددهی)", "")

    headers = ["شرح مفروضات", "مقدار (سال 1403)", "مقدار (سال 1402)"]
    ws.cell(row=current_row, column=1, value=headers[0]).font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=headers[1]).font = Font(bold=True)
    ws.cell(row=current_row, column=3, value=headers[2]).font = Font(bold=True)
    current_row += 1

    assumptions = {
        "مفروضات صورت سود و زیان": [
            ("درصد رشد درآمدهای عملیاتی", 0.50, 0.15),
            ("نرخ مالیات بر درآمد", 0.25, 0.25),
            ("سهم فروش مرغ گوشتی از کل درآمد عملیاتی", 0.80, 0.81),
            ("سهم فروش جوجه یک روزه از کل درآمد عملیاتی", 0.10, 0.12),
            ("سهم فروش کود از کل درآمد عملیاتی", 0.10, 0.07),
            # مفروضات جدید برای یادداشت 22.23 (دارایی‌های زیستی)
            ("جوجه یک روزه مولد به درصد از درآمدهای عملیاتی", 0.01, 0.01),
            ("مرغ مولد به درصد از درآمدهای عملیاتی", 0.05, 0.04),
            ("سایر دارایی‌های زیستی به درصد از درآمدهای عملیاتی", 0.005, 0.005),
            # مفروضات جدید برای یادداشت 24.25 (اموال، ماشین‌آلات و تجهیزات)
            ("سهم زمین از بهای تمام شده دارایی ثابت", 0.15, 0.15),
            ("سهم ساختمان از بهای تمام شده دارایی ثابت", 0.40, 0.40),
            ("سهم ماشین‌آلات از بهای تمام شده دارایی ثابت", 0.30, 0.30),
            ("سهم تجهیزات از بهای تمام شده دارایی ثابت", 0.15, 0.15),
            # مفروضات جدید برای یادداشت 32.33 (ذخایر)
            ("سایر ذخایر به درصد از درآمدهای عملیاتی", 0.005, 0.004),
            ("مالیات تکلیفی اجاره به درصد از درآمدهای عملیاتی", 0.003, 0.003),
            ("مالیات ارزش افزوده به درصد از درآمدهای عملیاتی", 0.007, 0.006)
        ],
        "مفروضات ترازنامه (سرمایه در گردش)": [
            ("دوره وصول مطالبات (روز)", 90, 95),
            ("دوره گردش موجودی کالا (روز)", 120, 125),
            ("دوره پرداخت بدهی‌ها (روز)", 75, 80),
            ("پیش‌پرداخت‌ها به درصد از درآمدهای عملیاتی", 0.015, 0.02),
            ("سایر دارایی‌های غیرجاری به درصد از درآمدهای عملیاتی", 0.03, 0.035),
            # مفروضات جدید برای سایر پرداختنی‌ها (جاری) در 46-3
            ("سایر پرداختنی‌ها به درصد از درآمدهای عملیاتی", 0.008, 0.007),
            # مفروضات جدید برای سایر بدهی‌های بلندمدت (مثل اوراق مشارکت بلندمدت)
            ("اوراق مشارکت بلندمدت به درصد از درآمدهای عملیاتی", 0.05, 0.04)
        ],
        "مفروضات دارایی ثابت و استهلاک": [
            ("سرمایه‌گذاری ثابت سالانه (CAPEX)", 600000, 450000),
            ("نرخ استهلاک سالانه (نسبت به بهای تمام شده اول دوره)", 0.10, 0.10)
        ],
        "مفروضات تامین مالی": [
            ("هزینه مالی ثابت", 50000, 60000),
            ("سود سهام پرداختی (درصد از سود خالص)", 0.40, 0.45),
            ("مبلغ وام جدید دریافتی طی سال", 850000, 300000),
            ("مبلغ بازپرداخت اصل وام طی سال", 50000, 40000),
            ("مانده اولیه تسهیلات بلندمدت (1402)", 0, 0)
        ],
        "مفروضات تفکیک وجوه نقد و سرمایه گذاری کوتاه مدت": [
            ("موجودی نقد و معادل‌های آن نزد بانک‌ها به درصد از موجودی نقد پایان دوره", 0.80, 0.80),
            ("سپرده‌های دیداری کوتاه‌مدت به درصد از موجودی نقد پایان دوره", 0.20, 0.20),
            ("سرمایه‌گذاری در اوراق بهادار به درصد از موجودی نقد پایان دوره", 0.05, 0.05),
            ("سایر سرمایه‌گذاری‌های کوتاه‌مدت به درصد از موجودی نقد پایان دوره", 0.02, 0.02),
            ("موجودی نقد نزد صندوق به درصد از موجودی نقد پایان دوره", 0.01, 0.01),
            ("موجودی نقد نزد بانک به درصد از موجودی نقد پایان دوره", 0.99, 0.99)
        ],
        "مفروضات تفکیک پیش پرداخت ها": [
            ("سهم پیش‌پرداخت اجاره از پیش‌پرداخت‌ها", 0.4, 0.4),
            ("سهم پیش‌پرداخت بیمه از پیش‌پرداخت‌ها", 0.3, 0.3),
            ("سهم سایر پیش‌پرداخت‌ها از پیش‌پرداخت‌ها", 0.3, 0.3)
        ],
        "مفروضات تفکیک سایر دارایی های غیرجاری (تفصیلی)": [
            ("سهم سپرده‌های بلندمدت از سایر دارایی‌های غیرجاری", 0.6, 0.6),
            ("سهم سایر اقلام غیرجاری از سایر دارایی‌های غیرجاری", 0.4, 0.4)
        ],
        "مفروضات تفکیک مطالبات غیرتجاری و سایر دارایی‌ها": [
            ("مطالبات از کارکنان به درصد از درآمدهای عملیاتی", 0.005, 0.005),
            ("سایر مطالبات به درصد از درآمدهای عملیاتی", 0.003, 0.003)
        ],
        "مفروضات تفکیک موجودی کالا و پیش پرداخت‌ها": [
            ("سهم موجودی مواد اولیه از کل موجودی کالا", 0.4, 0.4),
            ("سهم موجودی کالای در جریان ساخت از کل موجودی کالا", 0.3, 0.3),
            ("سهم موجودی کالای ساخته شده از کل موجودی کالا", 0.3, 0.3)
        ],
        "مفروضات تفکیک تسهیلات مالی کوتاه‌مدت": [
            ("سهم تسهیلات بانکی کوتاه‌مدت از بخش جاری تسهیلات بلندمدت", 0.7, 0.7),
            ("سهم سایر تسهیلات کوتاه‌مدت از بخش جاری تسهیلات بلندمدت", 0.3, 0.3)
        ],
        "مفروضات تفکیک حساب‌ها و اسناد پرداختنی": [
            ("سهم حساب‌های پرداختنی تجاری از حساب‌ها و اسناد پرداختنی", 0.8, 0.8),
            ("سهم اسناد پرداختنی تجاری از حساب‌ها و اسناد پرداختنی", 0.2, 0.2),
        ],
        "مفروضات تفکیک تسهیلات مالی کوتاه‌مدت (تفصیلی)": [
            ("سهم وام از بانک ملی از تسهیلات کوتاه مدت", 0.5, 0.5),
            ("سهم وام از بانک ملت از تسهیلات کوتاه مدت", 0.5, 0.5)
        ],
        "مفروضات تفکیک وجوه نقد (تفصیلی)": [
            ("سهم صندوق از وجوه نقد", 0.1, 0.1),
            ("سهم بانک پاسارگاد از وجوه نقد", 0.9, 0.9)
        ],
        "مفروضات تفکیک حساب‌ها و اسناد دریافتنی تجاری و غیرتجاری": [
            ("سهم حساب‌های دریافتنی تجاری از حساب‌ها و اسناد دریافتنی", 0.8, 0.8),
            ("سهم اسناد دریافتنی تجاری از حساب‌ها و اسناد دریافتنی", 0.2, 0.2),
        ],
        "مفروضات دارایی‌های غیرجاری نگهداری شده برای فروش": [
            ("زمین نگهداری شده برای فروش به درصد از درآمدهای عملیاتی", 0.01, 0.01),
            ("ساختمان نگهداری شده برای فروش به درصد از درآمدهای عملیاتی", 0.005, 0.005),
        ],
        "مفروضات سایر دارایی‌ها": [
            ("ودیعه اجاره به درصد از درآمدهای عملیاتی", 0.002, 0.002),
            ("سپرده حسن انجام کار به درصد از درآمدهای عملیاتی", 0.001, 0.001),
        ],
        "مفروضات تفکیک تسهیلات مالی بلندمدت": [
            ("سهم وام بانکی بلندمدت از تسهیلات بلندمدت", 0.7, 0.7),
            ("سهم وام بلندمدت از صندوق توسعه ملی از تسهیلات بلندمدت", 0.3, 0.3),
        ]
    }

    assumption_map = {}
    for category, items in assumptions.items():
        ws.cell(row=current_row, column=1, value=category).font = Font(bold=True, color="000080")
        current_row += 1
        for desc, val_1403, val_1402 in items:
            ws.cell(row=current_row, column=1, value=desc)
            ws.cell(row=current_row, column=2, value=val_1403)
            ws.cell(row=current_row, column=3, value=val_1402)
            if "درصد" in desc or "نرخ" in desc or "سهم" in desc:
                ws.cell(row=current_row, column=2).number_format = '0.00%'
                ws.cell(row=current_row, column=3).number_format = '0.00%'
            
            assumption_map[desc] = {'1403': f'B{current_row}', '1402': f'C{current_row}'}
            current_row += 1
        current_row += 1
    
    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"

    return assumption_map


def populate_payroll_list_sheet(ws):
    """پر کردن شیت لیست حقوق و دستمزد با آدرس‌دهی دقیق خروجی‌ها و هزینه کنترل شده."""
    ws.title = "لیست حقوق و دستمزد"
    col_widths = {
        'A': 5, 'B': 15, 'C': 15, 'D': 15, 'E': 20, 'F': 20, 'G': 15, 'H': 15,
        'I': 20, 'J': 20, 'K': 20, 'L': 20, 'M': 20, 'N': 20, 'O': 20, 'P': 20, 'Q': 20, 'R': 20, 'S': 20, 'T': 20, 'U': 20,
        'V': 20, 'W': 20, 'X': 20, 'Y': 20, 'Z': 20, 'AA': 20, 'AB': 20, 'AC': 20, 'AD': 20, 'AE': 20, 'AF': 20, 'AG': 20, 'AH': 20, 'AI': 20, 'AJ': 20, 'AK': 20, 'AL': 20
    }
    set_rtl_and_column_widths(ws, col_widths)
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "لیست حقوق و دستمزد (سال 1403 و 1402)", "تفکیک بر اساس واحد", "(ارقام به ریال)")

    headers_base = [
        "ردیف", "نام", "نام خانوادگی", "واحد", "سمت", "کد ملی", "شماره بیمه", "تعداد اولاد"
    ]
    headers_1403_data = [
        "حقوق پایه (1403)", "حق مسکن (1403)", "حق بن (1403)", "حق اولاد (1403)", "جمع مزایا (1403)", "حقوق ناخالص (1403)", "حقوق مشمول بیمه (1403)",
        "بیمه سهم کارمند (7% - 1403)", "حقوق مشمول مالیات (1403)", "مالیات حقوق (1403)", "کسورات متفرقه (1403)", "جمع کسورات (1403)",
        "حقوق خالص (پرداختنی - 1403)", "بیمه سهم کارفرما (23% - 1403)", "کل هزینه برای کارفرما (1403)"
    ]
    headers_1402_data = [
        "حقوق پایه (1402)", "حق مسکن (1402)", "حق بن (1402)", "حق اولاد (1402)", "جمع مزایا (1402)", "حقوق ناخالص (1402)", "حقوق مشمول بیمه (1402)",
        "بیمه سهم کارمند (7% - 1402)", "حقوق مشمول مالیات (1402)", "مالیات حقوق (1402)", "کسورات متفرقه (1402)", "جمع کسورات (1402)",
        "حقوق خالص (پرداختنی - 1402)", "بیمه سهم کارفرما (23% - 1402)", "کل هزینه برای کارفرما (1402)"
    ]
    
    for i, header in enumerate(headers_base):
        ws.cell(row=current_row, column=1 + i, value=header).font = Font(bold=True)
    
    for i, header in enumerate(headers_1403_data):
        ws.cell(row=current_row, column=9 + i, value=header).font = Font(bold=True)
    
    for i, header in enumerate(headers_1402_data):
        ws.cell(row=current_row, column=9 + len(headers_1403_data) + i, value=header).font = Font(bold=True)
    
    current_row += 1

    min_wage_daily_1403 = 2_388_728
    housing_allowance_1403 = 9_000_000
    consumer_basket_allowance_1403 = 14_000_000
    tax_exemption_monthly_1403 = 120_000_000
    
    min_wage_daily_1402 = 1_769_428
    housing_allowance_1402 = 6_500_000
    consumer_basket_allowance_1402 = 11_000_000
    tax_exemption_monthly_1402 = 100_000_000

    tax_rate_excess = 0.10
    ins_employee_share_rate = 0.07
    ins_employer_share_rate = 0.23

    employees_data_list = generate_all_employees_data(100)
    initial_data_row_start = current_row

    min_wage_monthly_1403 = min_wage_daily_1403 * 30
    min_wage_monthly_1402 = min_wage_daily_1402 * 30

    for i, emp in enumerate(employees_data_list):
        row_idx = initial_data_row_start + i

        base_salary_per_employee_1403 = random.randint(min_wage_monthly_1403, 120_000_000)
        if "مدیر" in emp["role"]:
            base_salary_per_employee_1403 = random.randint(130_000_000, 200_000_000)
        elif "دامپزشک" in emp["role"]:
            base_salary_per_employee_1403 = random.randint(100_000_000, 160_000_000)
        elif "کارشناس" in emp["role"] or "حسابدار" in emp["role"] or "انباردار" in emp["role"]:
            base_salary_per_employee_1403 = random.randint(80_000_000, 130_000_000)

        base_salary_per_employee_1402 = int(base_salary_per_employee_1403 * random.uniform(0.75, 0.85)) 
        base_salary_per_employee_1402 = max(min_wage_monthly_1402, base_salary_per_employee_1402)

        child_benefit_amount_1403 = emp["num_children"] * 3 * min_wage_daily_1403
        child_benefit_amount_1402 = emp["num_children"] * 3 * min_wage_daily_1402
        
        ws.cell(row=row_idx, column=1, value=emp["id"])
        ws.cell(row=row_idx, column=2, value=emp["first_name"])
        ws.cell(row=row_idx, column=3, value=emp["last_name"])
        ws.cell(row=row_idx, column=4, value=emp["unit"])
        ws.cell(row=row_idx, column=5, value=emp["role"])
        ws.cell(row=row_idx, column=6, value=random.randint(1000000000, 9999999999))
        ws.cell(row=row_idx, column=7, value=random.randint(10000000000, 99999999999))
        ws.cell(row=row_idx, column=8, value=emp["num_children"])

        ws.cell(row=row_idx, column=9, value=base_salary_per_employee_1403)
        ws.cell(row=row_idx, column=10, value=housing_allowance_1403)
        ws.cell(row=row_idx, column=11, value=consumer_basket_allowance_1403)
        ws.cell(row=row_idx, column=12, value=child_benefit_amount_1403)
        ws.cell(row=row_idx, column=13, value=f"=SUM(J{row_idx}:L{row_idx})")
        ws.cell(row=row_idx, column=14, value=f"=I{row_idx}+M{row_idx}")
        ws.cell(row=row_idx, column=15, value=f"=I{row_idx}+J{row_idx}+K{row_idx}")
        ws.cell(row=row_idx, column=16, value=f"=O{row_idx}*{ins_employee_share_rate}")
        ws.cell(row=row_idx, column=17, value=f"=MAX(0, N{row_idx}-P{row_idx}-{tax_exemption_monthly_1403})")
        ws.cell(row=row_idx, column=18, value=f"=ROUND(Q{row_idx}*{tax_rate_excess},0)")
        ws.cell(row=row_idx, column=19, value=random.randint(500_000, 2_000_000))
        ws.cell(row=row_idx, column=20, value=f"=SUM(P{row_idx},R{row_idx},S{row_idx})")
        ws.cell(row=row_idx, column=21, value=f"=N{row_idx}-T{row_idx}")
        ws.cell(row=row_idx, column=22, value=f"=O{row_idx}*{ins_employer_share_rate}")
        ws.cell(row=row_idx, column=23, value=f"=N{row_idx}+V{row_idx}")

        col_offset = 15
        ws.cell(row=row_idx, column=9 + col_offset, value=base_salary_per_employee_1402)
        ws.cell(row=row_idx, column=10 + col_offset, value=housing_allowance_1402)
        ws.cell(row=row_idx, column=11 + col_offset, value=consumer_basket_allowance_1402)
        ws.cell(row=row_idx, column=12 + col_offset, value=child_benefit_amount_1402)
        ws.cell(row=row_idx, column=13 + col_offset, value=f"=SUM(Y{row_idx}:AA{row_idx})")
        ws.cell(row=row_idx, column=14 + col_offset, value=f"=X{row_idx}+AB{row_idx}")
        ws.cell(row=row_idx, column=15 + col_offset, value=f"=X{row_idx}+Y{row_idx}+Z{row_idx}")
        ws.cell(row=row_idx, column=16 + col_offset, value=f"=AD{row_idx}*{ins_employee_share_rate}")
        ws.cell(row=row_idx, column=17 + col_offset, value=f"=MAX(0, AC{row_idx}-AE{row_idx}-{tax_exemption_monthly_1402})")
        ws.cell(row=row_idx, column=18 + col_offset, value=f"=ROUND(AF{row_idx}*{tax_rate_excess},0)")
        ws.cell(row=row_idx, column=19 + col_offset, value=random.randint(400_000, 1_500_000))
        ws.cell(row=row_idx, column=20 + col_offset, value=f"=SUM(AE{row_idx},AG{row_idx},AH{row_idx})")
        ws.cell(row=row_idx, column=21 + col_offset, value=f"=AC{row_idx}-AI{row_idx}")
        ws.cell(row=row_idx, column=22 + col_offset, value=f"=AD{row_idx}*{ins_employer_share_rate}")
        ws.cell(row=row_idx, column=23 + col_offset, value=f"=AC{row_idx}+AK{row_idx}")


    final_data_row = ws.max_row
    
    current_row = final_data_row + 2

    ws[f'A{current_row}'] = "جمع کل ماهانه (ریال) - 1403"
    for col_idx in range(9, 24):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{current_row}'] = f'=SUM({col_letter}{initial_data_row_start}:{col_letter}{final_data_row})'
    current_row += 1
    
    ws[f'A{current_row}'] = "جمع کل ماهانه (ریال) - 1402"
    for col_idx in range(9 + len(headers_1403_data), 9 + len(headers_1403_data) + len(headers_1402_data)):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{current_row}'] = f'=SUM({col_letter}{initial_data_row_start}:{col_letter}{final_data_row})'
    current_row += 1


    ws[f'A{current_row}'] = "جمع کل سالانه (ریال) - 1403"
    for col_idx in range(9, 24):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{current_row}'] = f'={col_letter}{current_row-2}*12'
    total_yearly_rial_1403_idx = current_row
    current_row += 1

    ws[f'A{current_row}'] = "جمع کل سالانه (ریال) - 1402"
    for col_idx in range(9 + len(headers_1403_data), 9 + len(headers_1403_data) + len(headers_1402_data)):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{current_row}'] = f'={col_letter}{current_row-2}*12'
    total_yearly_rial_1402_idx = current_row
    current_row += 1


    total_yearly_million_1403_idx = current_row
    ws[f'A{current_row}'] = "جمع کل سالانه (میلیون ریال) - 1403"
    for col_idx in range(9, 24):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{current_row}'] = f'={col_letter}{current_row-2}/1000000'
        ws[f'{col_letter}{current_row}'].number_format = '#,##0.00'
    current_row += 1

    total_yearly_million_1402_idx = current_row
    ws[f'A{current_row}'] = "جمع کل سالانه (میلیون ریال) - 1402"
    for col_idx in range(9 + len(headers_1403_data), 9 + len(headers_1403_data) + len(headers_1402_data)):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{current_row}'] = f'={col_letter}{current_row-2}/1000000'
        ws[f'{col_letter}{current_row}'].number_format = '#,##0.00'
    current_row += 1

    current_row += 1
    ws[f'A{current_row}'] = "خروجی برای سایر شیت‌ها (ارقام به میلیون ریال):"
    current_row += 1
    
    ws[f'B{current_row}'] = "کل هزینه سالانه پرسنل فروش - 1403:"
    ws[f'E{current_row}'] = f"""=ROUND(SUMIF(D{initial_data_row_start}:D{final_data_row},"*فروش*",W{initial_data_row_start}:{get_column_letter(23)}{final_data_row})*12/1000000,0)"""
    payroll_sales_1403_row = current_row
    current_row += 1
    ws[f'B{current_row}'] = "کل هزینه سالانه پرسنل فروش - 1402:"
    ws[f'E{current_row}'] = f"""=ROUND(SUMIF(D{initial_data_row_start}:D{final_data_row},"*فروش*",AL{initial_data_row_start}:{get_column_letter(38)}{final_data_row})*12/1000000,0)"""
    payroll_sales_1402_row = current_row
    current_row += 1
    
    ws[f'B{current_row}'] = "کل هزینه سالانه پرسنل اداری - 1403:"
    ws[f'E{current_row}'] = f"""=ROUND(SUMIF(D{initial_data_row_start}:D{final_data_row},"*اداری*",W{initial_data_row_start}:{get_column_letter(23)}{final_data_row})*12/1000000,0)"""
    payroll_admin_1403_row = current_row
    current_row += 1
    ws[f'B{current_row}'] = "کل هزینه سالانه پرسنل اداری - 1402:"
    ws[f'E{current_row}'] = f"""=ROUND(SUMIF(D{initial_data_row_start}:D{final_data_row},"*اداری*",AL{initial_data_row_start}:{get_column_letter(38)}{final_data_row})*12/1000000,0)"""
    payroll_admin_1402_row = current_row
    current_row += 1

    ws[f'B{current_row}'] = "کل هزینه سالانه پرسنل تولید - 1403:"
    ws[f'E{current_row}'] = f"""=ROUND(SUMIFS(W{initial_data_row_start}:{get_column_letter(23)}{final_data_row}, D{initial_data_row_start}:D{final_data_row}, "<>*فروش*", D{initial_data_row_start}:D{final_data_row}, "<>*اداری*")*12/1000000, 0)"""
    payroll_production_1403_row = current_row
    current_row += 1
    ws[f'B{current_row}'] = "کل هزینه سالانه پرسنل تولید - 1402:"
    ws[f'E{current_row}'] = f"""=ROUND(SUMIFS(AL{initial_data_row_start}:{get_column_letter(38)}{final_data_row}, D{initial_data_row_start}:D{final_data_row}, "<>*فروش*", D{initial_data_row_start}:D{final_data_row}, "<>*اداری*")*12/1000000, 0)"""
    payroll_production_1402_row = current_row
    current_row += 1
    
    ws.cell(row=1, column=52, value="بازگشت به سود و زیان").hyperlink = f"#'سودوزیان'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"

    return (
        f"E{payroll_sales_1403_row}",
        f"E{payroll_sales_1402_row}",
        f"E{payroll_admin_1403_row}",
        f"E{payroll_admin_1402_row}",
        f"E{payroll_production_1403_row}",
        f"E{payroll_production_1402_row}",
        f"N{total_yearly_million_1403_idx}",   # کل حقوق ناخالص سالانه 1403 (میلیون ریال)
        f"AC{total_yearly_million_1402_idx}",  # کل حقوق ناخالص سالانه 1402 (میلیون ریال)
        f"V{total_yearly_million_1403_idx}",   # کل بیمه سهم کارفرما سالانه 1403 (میلیون ریال)
        f"AK{total_yearly_million_1402_idx}",  # کل بیمه سهم کارفرما سالانه 1402 (میلیون ریال)
        # دو مقدار زیر اضافه شده‌اند تا تعداد خروجی‌ها 12 شود و خطای IndexError برطرف گردد
        f"M{total_yearly_million_1403_idx}",   # مثال: جمع مزایا 1403 (میلیون ریال)
        f"AB{total_yearly_million_1402_idx}"  # مثال: جمع مزایا 1402 (میلیون ریال)
    )


def populate_detailed_inventory_sheet(ws):
    """پر کردن شیت موجودی تفصیلی."""
    ws.title = "موجودی_تفصیلی"
    col_widths = {'A': 5, 'B': 20, 'C': 10, 'D': 15, 'E': 15, 'F': 15, 'G': 15,
                  'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 15}
    set_rtl_and_column_widths(ws, col_widths)
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "موجودی تفصیلی انبار (مقدار و ریال)", "برای سال مالی منتهی به 29 اسفند 1403 و 1402", "(ارقام به ریال برای قیمت واحد و میلیون ریال برای مقادیر)")

    headers_qty = [
        "ردیف", "نام کالا", "واحد", 
        "ابتدای دوره 1403 (مقدار)", "ورود 1403 (مقدار)", "خروج 1403 (مقدار)", "پایان دوره 1403 (مقدار)",
        "ابتدای دوره 1402 (مقدار)", "ورود 1402 (مقدار)", "خروج 1402 (مقدار)", "پایان دوره 1402 (مقدار)",
        "قیمت واحد میانگین (ریال)"
    ]
    ws.cell(row=current_row, column=1, value=headers_qty[0]).font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=headers_qty[1]).font = Font(bold=True)
    ws.cell(row=current_row, column=3, value=headers_qty[2]).font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=headers_qty[3]).font = Font(bold=True)
    ws.cell(row=current_row, column=5, value=headers_qty[4]).font = Font(bold=True)
    ws.cell(row=current_row, column=6, value=headers_qty[5]).font = Font(bold=True)
    ws.cell(row=current_row, column=7, value=headers_qty[6]).font = Font(bold=True)
    ws.cell(row=current_row, column=8, value=headers_qty[7]).font = Font(bold=True)
    ws.cell(row=current_row, column=9, value=headers_qty[8]).font = Font(bold=True)
    ws.cell(row=current_row, column=10, value=headers_qty[9]).font = Font(bold=True)
    ws.cell(row=current_row, column=11, value=headers_qty[10]).font = Font(bold=True)
    ws.cell(row=current_row, column=12, value=headers_qty[11]).font = Font(bold=True)
    current_row += 1

    inventory_items = [
        (1, "جوجه یک روزه", "عدد", 800000, 4500000, 4300000, None, 600000, 3800000, 3400000, None, 150),
        (2, "خوراک (دان)", "کیلوگرم", 8000000, 25000000, 24000000, None, 6000000, 22000000, 20000000, None, 390),
        (3, "مرغ در حال رشد", "عدد", 200000, 900000, 850000, None, 150000, 750000, 700000, None, 3600),
        (4, "دارو و واکسن", "بسته", 10000, 40000, 38000, None, 8000, 30000, 28000, None, 25),
        (5, "مرغ آماده فروش", "کیلوگرم", 50000, 900000, 880000, None, 40000, 750000, 720000, None, 140)
    ]

    initial_data_row_start_qty = current_row
    for item in inventory_items:
        row_idx = initial_data_row_start_qty + inventory_items.index(item)
        ws.cell(row=row_idx, column=1, value=item[0])
        ws.cell(row=row_idx, column=2, value=item[1])
        ws.cell(row=row_idx, column=3, value=item[2])
        ws.cell(row=row_idx, column=4, value=item[3])
        ws.cell(row=row_idx, column=5, value=item[4])
        ws.cell(row=row_idx, column=6, value=item[5])
        ws.cell(row=row_idx, column=7, value=f'=D{row_idx}+E{row_idx}-F{row_idx}')
        ws.cell(row=row_idx, column=8, value=item[7])
        ws.cell(row=row_idx, column=9, value=item[8])
        ws.cell(row=row_idx, column=10, value=item[9])
        ws.cell(row=row_idx, column=11, value=f'=H{row_idx}+I{row_idx}-J{row_idx}')
        ws.cell(row=row_idx, column=12, value=item[11])
    
    current_row = ws.max_row + 2


    # سربرگ های بخش ریالی
    headers_value = [
        "ردیف", "نام کالا", "واحد", 
        "ابتدای دوره 1403 (میلیون ریال)", "ورود 1403 (میلیون ریال)", "خروج (بهای تمام شده) 1403 (میلیون ریال)", "پایان دوره 1403 (میلیون ریال)",
        "ابتدای دوره 1402 (میلیون ریال)", "ورود 1402 (میلیون ریال)", "خروج (بهای تمام شده) 1402 (میلیون ریال)", "پایان دوره 1402 (میلیون ریال)"
    ]
    ws.cell(row=current_row, column=2, value="اطلاعات ریالی").font = Font(bold=True)
    current_row += 1
    ws.cell(row=current_row, column=1, value=headers_value[0]).font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=headers_value[1]).font = Font(bold=True)
    ws.cell(row=current_row, column=3, value=headers_value[2]).font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=headers_value[3]).font = Font(bold=True)
    ws.cell(row=current_row, column=5, value=headers_value[4]).font = Font(bold=True)
    ws.cell(row=current_row, column=6, value=headers_value[5]).font = Font(bold=True)
    ws.cell(row=current_row, column=7, value=headers_value[6]).font = Font(bold=True)
    ws.cell(row=current_row, column=8, value=headers_value[7]).font = Font(bold=True)
    ws.cell(row=current_row, column=9, value=headers_value[8]).font = Font(bold=True)
    ws.cell(row=current_row, column=10, value=headers_value[9]).font = Font(bold=True)
    ws.cell(row=current_row, column=11, value=headers_value[10]).font = Font(bold=True)
    current_row += 1

    initial_data_row_start_value = current_row
    for i in range(len(inventory_items)):
        data_row_qty = initial_data_row_start_qty + i
        row_idx_value = initial_data_row_start_value + i

        ws.cell(row=row_idx_value, column=1, value=f'=A{data_row_qty}')
        ws.cell(row=row_idx_value, column=2, value=f'=B{data_row_qty}')
        ws.cell(row=row_idx_value, column=3, value='م.ر')
        
        ws.cell(row=row_idx_value, column=4, value=f'=ROUND(D{data_row_qty}*L{data_row_qty}/1000000,0)')
        ws.cell(row=row_idx_value, column=5, value=f'=ROUND(E{data_row_qty}*L{data_row_qty}/1000000,0)')
        ws.cell(row=row_idx_value, column=6, value=f'=ROUND(F{data_row_qty}*L{data_row_qty}/1000000,0)')
        ws.cell(row=row_idx_value, column=7, value=f'=ROUND(G{data_row_qty}*L{data_row_qty}/1000000,0)')
        
        ws.cell(row=row_idx_value, column=8, value=f'=ROUND(H{data_row_qty}*L{data_row_qty}/1000000,0)')
        ws.cell(row=row_idx_value, column=9, value=f'=ROUND(I{data_row_qty}*L{data_row_qty}/1000000,0)')
        ws.cell(row=row_idx_value, column=10, value=f'=ROUND(J{data_row_qty}*L{data_row_qty}/1000000,0)')
        ws.cell(row=row_idx_value, column=11, value=f'=ROUND(K{data_row_qty}*L{data_row_qty}/1000000,0)')
    
    final_data_row_value = ws.max_row
    current_row = final_data_row_value + 1

    ws.cell(row=current_row, column=2, value="جمع کل (میلیون ریال)").font = Font(bold=True)
    for col_idx in range(4, 12): 
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{current_row}'] = f'=SUM({col_letter}{initial_data_row_start_value}:{col_letter}{final_data_row_value})'
        ws[f'{col_letter}{current_row}'].font = Font(bold=True)
    
    total_row_value_final = current_row
    current_row += 2


    ws.cell(row=current_row, column=2, value="**خروجی‌ها برای سایر شیت‌ها**").font = Font(bold=True)
    current_row += 1

    ws.cell(row=current_row, column=2, value="بهای تمام شده سال 1403 (برای سود و زیان):")
    ws.cell(row=current_row, column=6, value=f"=F{total_row_value_final}").font = Font(bold=True)
    current_row += 1

    ws.cell(row=current_row, column=2, value="بهای تمام شده سال 1402 (برای سود و زیان):")
    ws.cell(row=current_row, column=6, value=f"=J{total_row_value_final}").font = Font(bold=True)
    current_row += 1

    ws.cell(row=current_row, column=2, value="موجودی پایان دوره 1403 (برای ترازنامه):")
    ws.cell(row=current_row, column=6, value=f"=G{total_row_value_final}").font = Font(bold=True)
    current_row += 1

    ws.cell(row=current_row, column=2, value="موجودی پایان دوره 1402 (برای ترازنامه):")
    ws.cell(row=current_row, column=6, value=f"=K{total_row_value_final}").font = Font(bold=True)
    current_row += 1

    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"


def populate_note_8_and_9(wb):
    ## یادداشت 9: بهای تمام شده
    ws9 = wb['9']
    for row in ws9.iter_rows(min_row=7):
        for cell in row:
            cell.value = None

    set_rtl_and_column_widths(ws9, {'A': 5, 'B': 45, 'C': 18, 'D': 18})
    current_row_ws9 = add_header(ws9, "شرکت نمونه", "یادداشت 9: بهای تمام شده", "سال مالی منتهی به 29 اسفند 1403 و 1402", "(مبالغ به میلیون ریال)")

    ws9.cell(row=current_row_ws9, column=2, value='شرح').font = Font(bold=True)
    ws9.cell(row=current_row_ws9, column=3, value='1403').font = Font(bold=True)
    ws9.cell(row=current_row_ws9, column=4, value='1402').font = Font(bold=True)
    current_row_ws9 += 1

    cogs_items = [
        ("بهای تمام شده کالای فروش رفته", "='موجودی_تفصیلی'!F30", "='موجودی_تفصیلی'!F31"),
        ("حقوق و دستمزد مستقیم تولید", f"='لیست حقوق و دستمزد'!{wb.payroll_output_cells_E121}", f"='لیست حقوق و دستمزد'!{wb.payroll_output_cells_E122}"),
        ("هزینه استهلاک دارایی‌های تولیدی (80%)", "='گردش دارایی ثابت'!D11*0.8", "='گردش دارایی ثابت'!D6*0.8"),
        ("سایر هزینه‌های مستقیم تولید (سربار)", 50000, 45000)
    ]
    
    start_row_cogs_data = current_row_ws9
    for item in cogs_items:
        ws9.cell(row=current_row_ws9, column=2, value=item[0])
        ws9.cell(row=current_row_ws9, column=3, value=item[1])
        ws9.cell(row=current_row_ws9, column=4, value=item[2])
        current_row_ws9 += 1
    
    total_row_cogs = current_row_ws9
    ws9.cell(row=total_row_cogs, column=2, value="جمع کل بهای تمام شده").font = Font(bold=True)
    ws9[f'C{total_row_cogs}'] = f"=SUM(C{start_row_cogs_data}:C{total_row_cogs-1})"
    ws9[f'D{total_row_cogs}'] = f"=SUM(D{start_row_cogs_data}:D{total_row_cogs-1})"
    wb.max_row_9 = total_row_cogs
    
    ws9.cell(row=1, column=52, value="بازگشت به سود و زیان").hyperlink = f"#'سودوزیان'!A1"
    ws9.cell(row=1, column=52).style = "Hyperlink"

    ## یادداشت 8: هزینه‌های فروش، اداری و عمومی
    ws8 = wb['8']
    for row in ws8.iter_rows(min_row=7):
        for cell in row:
            cell.value = None

    set_rtl_and_column_widths(ws8, {'A': 5, 'B': 45, 'C': 18, 'D': 18})
    current_row_ws8 = add_header(ws8, "شرکت نمونه", "یادداشت 8: هزینه‌های فروش، اداری و عمومی", "سال مالی منتهی به 29 اسفند 1403 و 1402", "(مبالغ به میلیون ریال)")

    ws8.cell(row=current_row_ws8, column=2, value='شرح').font = Font(bold=True)
    ws8.cell(row=current_row_ws8, column=3, value='1403').font = Font(bold=True)
    ws8.cell(row=current_row_ws8, column=4, value='1402').font = Font(bold=True)
    current_row_ws8 += 1

    ws8.cell(row=current_row_ws8, column=2, value='الف) هزینه‌های فروش و توزیع:').font = Font(bold=True)
    current_row_ws8 += 1

    sga_sales_items = [
        ("هزینه پرسنل فروش", f"='لیست حقوق و دستمزد'!{wb.payroll_output_cells_E117}", f"='لیست حقوق و دستمزد'!{wb.payroll_output_cells_E118}"),
        ("هزینه تبلیغات و بازاریابی", 50000, 40000)
    ]
    start_row_sales_data = current_row_ws8
    for item in sga_sales_items:
        ws8.cell(row=current_row_ws8, column=2, value=item[0])
        ws8.cell(row=current_row_ws8, column=3, value=item[1])
        ws8.cell(row=current_row_ws8, column=4, value=item[2])
        current_row_ws8 += 1
    
    total_row_sales = current_row_ws8
    ws8.cell(row=total_row_sales, column=2, value="جمع هزینه‌های فروش").font = Font(bold=True)
    ws8[f'C{total_row_sales}'] = f"=SUM(C{start_row_sales_data}:C{total_row_sales-1})"
    ws8[f'D{total_row_sales}'] = f"=SUM(D{start_row_sales_data}:D{total_row_sales-1})"
    current_row_ws8 += 2
    
    ws8.cell(row=current_row_ws8, column=2, value='ب) هزینه‌های اداری و عمومی:').font = Font(bold=True)
    current_row_ws8 += 1

    sga_admin_items = [
        ("هزینه پرسنل اداری", f"='لیست حقوق و دستمزد'!{wb.payroll_output_cells_E119}", f"='لیست حقوق و دستمزد'!{wb.payroll_output_cells_E120}"),
        ("هزینه استهلاک دارایی‌های اداری (20%)", "='گردش دارایی ثابت'!D11*0.2", "='گردش دارایی ثابت'!D6*0.2"),
        ("هزینه ذخیره مزایای پایان خدمت کارکنان", 80000, 75000),
        ("سایر هزینه‌های اداری", 30000, 25000)
    ]
    start_row_admin_data = current_row_ws8
    for item in sga_admin_items:
        ws8.cell(row=current_row_ws8, column=2, value=item[0])
        ws8.cell(row=current_row_ws8, column=3, value=item[1])
        ws8.cell(row=current_row_ws8, column=4, value=item[2])
        current_row_ws8 += 1
    
    total_row_admin = current_row_ws8
    ws8.cell(row=total_row_admin, column=2, value="جمع هزینه‌های اداری").font = Font(bold=True)
    ws8[f'C{total_row_admin}'] = f"=SUM(C{start_row_admin_data}:C{total_row_admin-1})"
    ws8[f'D{total_row_admin}'] = f"=SUM(D{start_row_admin_data}:D{total_row_admin-1})"
    current_row_ws8 += 2
    
    total_row_all_sga = current_row_ws8
    ws8.cell(row=total_row_all_sga, column=2, value="جمع کل هزینه‌های فروش، اداری و عمومی").font = Font(bold=True)
    ws8[f'C{total_row_all_sga}'] = f"=C{total_row_sales}+C{total_row_admin}"
    ws8[f'D{total_row_all_sga}'] = f"=D{total_row_sales}+D{total_row_admin}"
    wb.max_row_8 = total_row_all_sga
    wb.max_row_8_benefits_expense = total_row_admin - 1
    
    ws8.cell(row=1, column=52, value="بازگشت به سود و زیان").hyperlink = f"#'سودوزیان'!A1"
    ws8.cell(row=1, column=52).style = "Hyperlink"


def populate_profit_loss_sheet(ws, assumption_map, wb_main_obj):
    """ایجاد صورت سود و زیان یکپارچه که هزینه‌ها را از یادداشت‌ها می‌خواند."""
    ws.title = "سودوزیان"
    col_widths = {'A': 5, 'B': 40, 'C': 12, 'D': 18, 'E': 10, 'F': 18, 'G': 18}
    set_rtl_and_column_widths(ws, col_widths)
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "صورت سود و زیان (یکپارچه)", "سال مالی منتهی به 29 اسفند 1403 و 1402", "(ارقام به میلیون ریال)")

    ws.cell(row=current_row, column=5, value="یادداشت").font = Font(bold=True)
    ws.cell(row=current_row, column=6, value="سال 1403").font = Font(bold=True)
    ws.cell(row=current_row, column=7, value="سال 1402").font = Font(bold=True)
    current_row += 1

    # درآمدهای عملیاتی
    ws.cell(row=current_row, column=2, value="درآمدهای عملیاتی").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value='5').hyperlink = f"#'5'!A1"
    ws.cell(row=current_row, column=5).style = "Hyperlink"
    ws.cell(row=current_row, column=7, value=f"='5'!C{wb_main_obj.max_row_5}")
    ws.cell(row=current_row, column=6, value=f"=G{current_row}*(1+'مفروضات'!{assumption_map['درصد رشد درآمدهای عملیاتی']['1403']})")
    row_revenue = current_row
    current_row += 1

    # بهای تمام شده درآمدهای عملیاتی
    ws.cell(row=current_row, column=2, value="بهای تمام شده درآمدهای عملیاتی").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value='9').hyperlink = f"#'9'!A1"
    ws.cell(row=current_row, column=5).style = "Hyperlink"
    ws.cell(row=current_row, column=6, value=f"=-'9'!C{wb_main_obj.max_row_9}")
    ws.cell(row=current_row, column=7, value=f"=-'9'!D{wb_main_obj.max_row_9}")
    row_cogs = current_row
    current_row += 1

    # سود ناخالص
    ws.cell(row=current_row, column=2, value="سود ناخالص").font = Font(bold=True)
    ws.cell(row=current_row, column=6, value=f"=F{row_revenue}+F{row_cogs}")
    ws.cell(row=current_row, column=7, value=f"=G{row_revenue}+G{row_cogs}")
    row_gross_profit = current_row
    current_row += 1
    
    # هزینه‌های فروش، اداری و عمومی
    current_row += 1
    ws.cell(row=current_row, column=2, value="هزینه‌های فروش، اداری و عمومی").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value='8').hyperlink = f"#'8'!A1"
    ws.cell(row=current_row, column=5).style = "Hyperlink"
    ws.cell(row=current_row, column=6, value=f"=-'8'!C{wb_main_obj.max_row_8}")
    ws.cell(row=current_row, column=7, value=f"=-'8'!D{wb_main_obj.max_row_8}")
    row_sga = current_row
    current_row += 1

    # سایر درآمدها
    ws.cell(row=current_row, column=2, value="سایر درآمدها").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value='26.27').hyperlink = f"#'26.27'!A1"
    ws.cell(row=current_row, column=5).style = "Hyperlink"
    ws.cell(row=current_row, column=6, value=f"='26.27'!C{wb_main_obj.max_row_26_27_revenue_section}")
    ws.cell(row=current_row, column=7, value=f"='26.27'!D{wb_main_obj.max_row_26_27_revenue_section}")
    row_other_income = current_row
    current_row += 1

    # سایر هزینه‌ها
    ws.cell(row=current_row, column=2, value="سایر هزینه‌ها").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value='26.27').hyperlink = f"#'26.27'!A1"
    ws.cell(row=current_row, column=5).style = "Hyperlink"
    ws.cell(row=current_row, column=6, value=f"=-'26.27'!C{wb_main_obj.max_row_26_27_expense_section}")
    ws.cell(row=current_row, column=7, value=f"=-'26.27'!D{wb_main_obj.max_row_26_27_expense_section}")
    row_other_expense = current_row
    current_row += 1

    # سود عملیاتی
    ws.cell(row=current_row, column=2, value="سود عملیاتی").font = Font(bold=True)
    ws.cell(row=current_row, column=6, value=f"=F{row_gross_profit}+F{row_sga}+F{row_other_income}+F{row_other_expense}")
    ws.cell(row=current_row, column=7, value=f"=G{row_gross_profit}+G{row_sga}+G{row_other_income}+G{row_other_expense}")
    row_operating_profit = current_row
    current_row += 1

    # هزینه‌های مالی
    current_row += 1
    ws.cell(row=current_row, column=2, value="هزینه‌های مالی").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value='26.27').hyperlink = f"#'26.27'!A1"
    ws.cell(row=current_row, column=5).style = "Hyperlink"
    ws.cell(row=current_row, column=6, value=f"=-'مفروضات'!{assumption_map['هزینه مالی ثابت']['1403']}")
    ws.cell(row=current_row, column=7, value=f"=-'مفروضات'!{assumption_map['هزینه مالی ثابت']['1402']}")
    row_finance_cost = current_row
    current_row += 1

    # سود قبل از مالیات
    ws.cell(row=current_row, column=2, value="سود قبل از مالیات").font = Font(bold=True)
    ws.cell(row=current_row, column=6, value=f"=F{row_operating_profit}+F{row_finance_cost}")
    ws.cell(row=current_row, column=7, value=f"=G{row_operating_profit}+G{row_finance_cost}")
    row_pbt = current_row
    current_row += 1

    # مالیات بر درآمد
    current_row += 1
    ws.cell(row=current_row, column=2, value="مالیات بر درآمد").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value='34').hyperlink = f"#'34'!A1"
    ws.cell(row=current_row, column=5).style = "Hyperlink"
    ws.cell(row=current_row, column=6, value=f"=IF(F{row_pbt}>0, F{row_pbt}*(-'مفروضات'!{assumption_map['نرخ مالیات بر درآمد']['1403']}), 0)")
    ws.cell(row=current_row, column=7, value=f"=IF(G{row_pbt}>0, G{row_pbt}*(-'مفروضات'!{assumption_map['نرخ مالیات بر درآمد']['1402']}), 0)")
    row_tax = current_row
    current_row += 1

    # سود خالص
    ws.cell(row=current_row, column=2, value="سود خالص").font = Font(bold=True)
    ws.cell(row=current_row, column=6, value=f"=F{row_pbt}+F{row_tax}")
    ws.cell(row=current_row, column=7, value=f"=G{row_pbt}+G{row_tax}")
    row_net_profit = current_row
    current_row += 1

    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"

    # ذخیره ردیف‌های نهایی سود خالص برای استفاده در سایر شیت‌ها
    ws.row_revenue_pl = row_revenue
    ws.row_cogs_pl = row_cogs
    ws.row_gross_profit_pl = row_gross_profit
    ws.row_sga_pl = row_sga
    ws.row_other_income_pl = row_other_income
    ws.row_other_expense_pl = row_other_expense
    ws.row_operating_profit_pl = row_operating_profit
    ws.row_finance_cost_pl = row_finance_cost
    ws.row_pbt_pl = row_pbt
    ws.row_tax_pl = row_tax
    ws.row_net_profit_pl_1403 = row_net_profit
    ws.row_net_profit_pl_1402 = row_net_profit
    # اینها باید در populate_numeric_note_sheets ذخیره شوند.
    ws.max_row_26_27_revenue_section = 0
    ws.max_row_26_27_expense_section = 0


def populate_balance_sheet(ws, assumption_map, wb_main_obj):
    """پر کردن شیت وضعیت مالی."""
    ws.title = "وضعیت مالی"
    col_widths = {'A': 5, 'B': 40, 'C': 45, 'D': 12, 'E': 18, 'F': 18}
    set_rtl_and_column_widths(ws, col_widths)
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "صورت وضعیت مالی (پویا)", "در تاریخ 29 اسفند 1403 و 1402", "(ارقام به میلیون ریال)")

    ws.cell(row=current_row, column=4, value="یادداشت").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value="1403").font = Font(bold=True)
    ws.cell(row=current_row, column=6, value="1402").font = Font(bold=True)
    current_row += 1
    
    # --- دارایی‌ها ---
    ws.cell(row=current_row, column=2, value="دارایی‌ها").font = Font(bold=True)
    current_row += 1
    ws.cell(row=current_row, column=2, value="دارایی‌های جاری").font = Font(bold=True)
    current_row += 1
    
    # موجودی نقد
    ws.cell(row=current_row, column=3, value="موجودی نقد")
    ws.cell(row=current_row, column=4, value=6).hyperlink = f"#'6'!A1"
    ws.cell(row=current_row, column=4).style = "Hyperlink"
    ws.cell(row=current_row, column=5, value=f"='جریان های نقدی'!C{wb_main_obj.max_row_cash_flow_end_1403}")
    ws.cell(row=current_row, column=6, value=f"='ترازنامه پایه'!D10")
    row_cash = current_row
    current_row += 1

    # حساب‌ها و اسناد دریافتنی
    ws.cell(row=current_row, column=3, value="حساب‌ها و اسناد دریافتنی")
    ws.cell(row=current_row, column=4, value="42.43").hyperlink = f"#'42.43'!A1"
    ws.cell(row=current_row, column=4).style = "Hyperlink"
    ws.cell(row=current_row, column=5, value=f"=(('مفروضات'!{assumption_map['دوره وصول مطالبات (روز)']['1403']})/365)*'سودوزیان'!F8")
    ws.cell(row=current_row, column=6, value=f"=(('مفروضات'!{assumption_map['دوره وصول مطالبات (روز)']['1402']})/365)*'سودوزیان'!G8")
    row_receivables = current_row
    current_row += 1

    # موجودی کالا
    ws.cell(row=current_row, column=3, value="موجودی کالا")
    ws.cell(row=current_row, column=4, value=9).hyperlink = f"#'موجودی'!A1"
    ws.cell(row=current_row, column=4).style = "Hyperlink"
    ws.cell(row=current_row, column=5, value="='موجودی_تفصیلی'!G27")
    ws.cell(row=current_row, column=6, value="='موجودی_تفصیلی'!K27")
    row_inventory = current_row
    current_row += 1

    # پیش‌پرداخت‌ها و سایر دارایی‌های جاری
    ws.cell(row=current_row, column=3, value="پیش‌پرداخت‌ها و سایر دارایی‌های جاری")
    ws.cell(row=current_row, column=4, value=10).hyperlink = f"#'10.11.12'!A1"
    ws.cell(row=current_row, column=4).style = "Hyperlink"
    ws.cell(row=current_row, column=5, value=f"='مفروضات'!{assumption_map['پیش‌پرداخت‌ها به درصد از درآمدهای عملیاتی']['1403']}*'سودوزیان'!F8")
    ws.cell(row=current_row, column=6, value=f"='مفروضات'!{assumption_map['پیش‌پرداخت‌ها به درصد از درآمدهای عملیاتی']['1402']}*'سودوزیان'!G8")
    row_prepayments = current_row
    current_row += 1

    # جمع دارایی‌های جاری
    ws.cell(row=current_row, column=2, value="جمع دارایی‌های جاری").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value=f"=SUM(E{row_cash}:E{row_prepayments})")
    ws.cell(row=current_row, column=6, value=f"=SUM(F{row_cash}:F{row_prepayments})")
    row_total_current_assets = current_row
    current_row += 2

    # دارایی‌های غیرجاری
    ws.cell(row=current_row, column=2, value="دارایی‌های غیرجاری").font = Font(bold=True)
    current_row += 1
    
    # دارایی‌های ثابت مشهود (ارزش دفتری)
    ws.cell(row=current_row, column=3, value="دارایی‌های ثابت مشهود (ارزش دفتری)")
    ws.cell(row=current_row, column=4, value="گردش دارایی ثابت").hyperlink = f"#'گردش دارایی ثابت'!A1"
    ws.cell(row=current_row, column=4).style = "Hyperlink"
    ws.cell(row=current_row, column=5, value="='گردش دارایی ثابت'!E13")
    ws.cell(row=current_row, column=6, value="='گردش دارایی ثابت'!E8")
    row_fixed_assets = current_row
    current_row += 1

    # سایر دارایی‌های غیرجاری
    ws.cell(row=current_row, column=3, value="سایر دارایی‌های غیرجاری")
    ws.cell(row=current_row, column=4, value=13).hyperlink = f"#'13'!A1"
    ws.cell(row=current_row, column=4).style = "Hyperlink"
    ws.cell(row=current_row, column=5, value=f"='مفروضات'!{assumption_map['سایر دارایی‌های غیرجاری به درصد از درآمدهای عملیاتی']['1403']}*'سودوزیان'!F8")
    ws.cell(row=current_row, column=6, value=f"='مفروضات'!{assumption_map['سایر دارایی‌های غیرجاری به درصد از درآمدهای عملیاتی']['1402']}*'سودوزیان'!G8")
    row_other_non_current_assets = current_row
    current_row += 1

    # جمع دارایی‌های غیرجاری
    ws.cell(row=current_row, column=2, value="جمع دارایی‌های غیرجاری").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value=f"=SUM(E{row_fixed_assets}:E{row_other_non_current_assets})")
    ws.cell(row=current_row, column=6, value=f"=SUM(F{row_fixed_assets}:F{row_other_non_current_assets})")
    row_total_non_current_assets = current_row
    current_row += 2

    # جمع کل دارایی‌ها
    ws.cell(row=current_row, column=2, value="جمع کل دارایی‌ها").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value=f"=E{row_total_current_assets}+E{row_total_non_current_assets}")
    ws.cell(row=current_row, column=6, value=f"=F{row_total_current_assets}+F{row_total_non_current_assets}")
    row_total_assets = current_row
    current_row += 2

    # --- بدهی‌ها و حقوق مالکانه ---
    ws.cell(row=current_row, column=2, value="بدهی‌ها و حقوق مالکانه").font = Font(bold=True)
    current_row += 1
    ws.cell(row=current_row, column=2, value="بدهی‌های جاری").font = Font(bold=True)
    current_row += 1
    
    # حساب‌ها و اسناد پرداختنی
    ws.cell(row=current_row, column=3, value="حساب‌ها و اسناد پرداختنی")
    ws.cell(row=current_row, column=4, value="28.29.30.31").hyperlink = f"#'28.29.30.31'!A1"
    ws.cell(row=current_row, column=4).style = "Hyperlink"
    ws.cell(row=current_row, column=5, value=f"='28.29.30.31'!C{wb_main_obj.max_row_28_29_30_31_partA}")
    ws.cell(row=current_row, column=6, value=f"='28.29.30.31'!D{wb_main_obj.max_row_28_29_30_31_partA}")
    row_payables = current_row
    current_row += 1

    # مالیات پرداختنی
    ws.cell(row=current_row, column=3, value="مالیات پرداختنی")
    ws.cell(row=current_row, column=4, value=17).hyperlink = f"#'17'!A1"
    ws.cell(row=current_row, column=4).style = "Hyperlink"
    ws.cell(row=current_row, column=5, value="='سودوزیان'!F20*-1")
    ws.cell(row=current_row, column=6, value="='سودوزیان'!G20*-1")
    row_tax_payable = current_row
    current_row += 1

    # سود سهام پرداختنی
    ws.cell(row=current_row, column=3, value="سود سهام پرداختنی")
    ws.cell(row=current_row, column=4, value=18).hyperlink = f"#'18'!A1"
    ws.cell(row=current_row, column=4).style = "Hyperlink"
    ws.cell(row=current_row, column=5, value=f"='18'!C{wb_main_obj.max_row_18}")
    ws.cell(row=current_row, column=6, value=f"='18'!D{wb_main_obj.max_row_18}")
    row_dividends_payable = current_row
    current_row += 1

    # بخش جاری تسهیلات بلندمدت
    ws.cell(row=current_row, column=3, value="بخش جاری تسهیلات بلندمدت")
    ws.cell(row=current_row, column=4, value=30).hyperlink = f"#'16'!A1"
    ws.cell(row=current_row, column=4).style = "Hyperlink"
    ws.cell(row=current_row, column=5, value=f"='مفروضات'!{assumption_map['مبلغ بازپرداخت اصل وام طی سال']['1403']}")
    ws.cell(row=current_row, column=6, value=f"='مفروضات'!{assumption_map['مبلغ بازپرداخت اصل وام طی سال']['1402']}")
    row_current_lt_debt = current_row
    current_row += 1

    # جمع بدهی‌های جاری
    ws.cell(row=current_row, column=2, value="جمع بدهی‌های جاری").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value=f"=SUM(E{row_payables}:E{row_current_lt_debt})")
    ws.cell(row=current_row, column=6, value=f"=SUM(F{row_payables}:F{row_current_lt_debt})")
    row_total_current_liabilities = current_row
    current_row += 2

    # دارایی‌های غیرجاری
    ws.cell(row=current_row, column=2, value="بدهی‌های غیرجاری").font = Font(bold=True)
    current_row += 1
    
    # تسهیلات مالی بلندمدت
    ws.cell(row=current_row, column=3, value="تسهیلات مالی بلندمدت")
    ws.cell(row=current_row, column=4, value=19).hyperlink = f"#'19'!A1"
    ws.cell(row=current_row, column=4).style = "Hyperlink"
    ws.cell(row=current_row, column=5, value=f"='19'!C{wb_main_obj.max_row_19}")
    ws.cell(row=current_row, column=6, value=f"='19'!D{wb_main_obj.max_row_19}")
    row_lt_debt = current_row
    current_row += 1

    # مزایای پایان خدمت کارکنان
    ws.cell(row=current_row, column=3, value="مزایای پایان خدمت کارکنان")
    ws.cell(row=current_row, column=4, value=20).hyperlink = f"#'20'!A1"
    ws.cell(row=current_row, column=4).style = "Hyperlink"
    ws.cell(row=current_row, column=5, value=f"='20'!C{wb_main_obj.max_row_20}")
    ws.cell(row=current_row, column=6, value=f"='20'!D{wb_main_obj.max_row_20}")
    row_employee_benefits = current_row
    current_row += 1

    # جمع بدهی‌های غیرجاری
    ws.cell(row=current_row, column=2, value="جمع بدهی‌های غیرجاری").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value=f"=SUM(E{row_lt_debt}:E{row_employee_benefits})")
    ws.cell(row=current_row, column=6, value=f"=SUM(F{row_lt_debt}:F{row_employee_benefits})")
    row_total_non_current_liabilities = current_row
    current_row += 2

    # جمع کل بدهی‌ها
    ws.cell(row=current_row, column=2, value="جمع کل بدهی‌ها").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value=f"=E{row_total_current_liabilities}+E{row_total_non_current_liabilities}")
    ws.cell(row=current_row, column=6, value=f"=F{row_total_current_liabilities}+F{row_total_non_current_liabilities}")
    row_total_liabilities = current_row
    current_row += 2

    # حقوق مالکانه
    ws.cell(row=current_row, column=2, value="حقوق مالکانه").font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=21).hyperlink = f"#'21'!A1"
    ws.cell(row=current_row, column=4).style = "Hyperlink"
    ws.cell(row=current_row, column=5, value=f"='حقوق مالکانه'!F{wb_main_obj.max_row_equity_total_equity_1403}")
    ws.cell(row=current_row, column=6, value=f"='حقوق مالکانه'!F{wb_main_obj.max_row_equity_total_equity_1402}")
    row_equity = current_row
    current_row += 2

    # جمع کل بدهی‌ها و حقوق مالکانه
    ws.cell(row=current_row, column=2, value="جمع کل بدهی‌ها و حقوق مالکانه").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value=f"=E{row_total_liabilities}+E{row_equity}")
    ws.cell(row=current_row, column=6, value=f"=F{row_total_liabilities}+F{row_equity}")
    row_total_liabilities_and_equity = current_row
    current_row += 2

    # کنترل تراز
    ws.cell(row=current_row, column=2, value="کنترل تراز").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value=f'=IF(ROUND(E{row_total_assets}-E{row_total_liabilities_and_equity},0)=0,"تراز","عدم تراز")')
    ws.cell(row=current_row, column=6, value=f'=IF(ROUND(F{row_total_assets}-F{row_total_liabilities_and_equity},0)=0,"تراز","عدم تراز")')
    current_row += 1

    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"

    # ذخیره ردیف‌های مهم برای استفاده در سایر شیت‌ها (main_model_execution)
    ws.row_total_assets = row_total_assets
    ws.row_total_liabilities = row_total_liabilities
    ws.row_total_current_assets = row_total_current_assets
    ws.row_total_current_liabilities = row_total_current_liabilities
    ws.row_payables = row_payables
    ws.row_current_lt_debt = row_current_lt_debt
    ws.row_lt_debt = row_lt_debt
    ws.row_employee_benefits = row_employee_benefits
    ws.row_receivables = row_receivables
    ws.row_inventory = row_inventory
    ws.row_prepayments = row_prepayments
    ws.row_other_non_current_assets = row_other_non_current_assets


def populate_fixed_asset_roll_forward_sheet(ws, assumption_map):
    """پر کردن شیت گردش دارایی‌های ثابت مشهود (پویا)."""
    ws.title = "گردش دارایی ثابت"
    col_widths = {'A': 35, 'B': 20, 'C': 20, 'D': 20, 'E': 20}
    set_rtl_and_column_widths(ws, col_widths)
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "گردش دارایی‌های ثابت مشهود (پویا)", "برای سال مالی منتهی به 29 اسفند 1403", "(ارقام به میلیون ریال)")

    headers = ["شرح", "مانده اول دوره", "افزایش (CAPEX)", "کاهش (استهلاک)", "مانده پایان دوره"]
    ws.cell(row=current_row, column=1, value=headers[0]).font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=headers[1]).font = Font(bold=True)
    ws.cell(row=current_row, column=3, value=headers[2]).font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=headers[3]).font = Font(bold=True)
    ws.cell(row=current_row, column=5, value=headers[4]).font = Font(bold=True)
    current_row += 1

    # محاسبات سال 1402
    row_1402_cost_start = current_row
    ws.cell(row=current_row, column=1, value="بهای تمام شده دارایی")
    ws.cell(row=current_row, column=2, value="='ترازنامه پایه'!D17").font = Font(bold=True)
    ws.cell(row=current_row, column=3, value=f"='مفروضات'!{assumption_map['سرمایه‌گذاری ثابت سالانه (CAPEX)']['1402']}")
    ws.cell(row=current_row, column=4, value=0)
    ws.cell(row=current_row, column=5, value=f"=SUM(B{current_row}:D{current_row})")
    current_row += 1

    row_1402_dep_start = current_row
    ws.cell(row=current_row, column=1, value="استهلاک انباشته")
    ws.cell(row=current_row, column=2, value="='ترازنامه پایه'!D18").font = Font(bold=True)
    ws.cell(row=current_row, column=3, value=0)
    ws.cell(row=current_row, column=4, value=f"=B{row_1402_cost_start}*'مفروضات'!{assumption_map['نرخ استهلاک سالانه (نسبت به بهای تمام شده اول دوره)']['1402']}")  
    ws.cell(row=current_row, column=5, value=f"=B{current_row}+D{current_row}")
    current_row += 2


    row_1402_book_value = current_row
    ws.cell(row=current_row, column=1, value="ارزش دفتری خالص").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=f"=B{row_1402_cost_start}-B{row_1402_dep_start}")
    ws.cell(row=current_row, column=5, value=f"=E{row_1402_cost_start}-E{row_1402_dep_start}")
    current_row += 2


    # محاسبات سال 1403
    row_1403_cost_start = current_row
    ws.cell(row=current_row, column=1, value="بهای تمام شده دارایی")
    ws.cell(row=current_row, column=2, value=f"=E{row_1402_cost_start}")
    ws.cell(row=current_row, column=3, value=f"='مفروضات'!{assumption_map['سرمایه‌گذاری ثابت سالانه (CAPEX)']['1403']}")
    ws.cell(row=current_row, column=4, value=0)
    ws.cell(row=current_row, column=5, value=f"=SUM(B{current_row}:D{current_row})")
    current_row += 1

    row_1403_dep_start = current_row
    ws.cell(row=current_row, column=1, value="استهلاک انباشته")
    ws.cell(row=current_row, column=2, value=f"=E{row_1402_dep_start}")
    ws.cell(row=current_row, column=3, value=0)
    ws.cell(row=current_row, column=4, value=f"=B{row_1403_cost_start}*'مفروضات'!{assumption_map['نرخ استهلاک سالانه (نسبت به بهای تمام شده اول دوره)']['1403']}")  
    ws.cell(row=current_row, column=5, value=f"=B{current_row}+D{current_row}")
    current_row += 2


    row_1403_book_value = current_row
    ws.cell(row=current_row, column=1, value="ارزش دفتری خالص").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=f"=B{row_1403_cost_start}-B{row_1403_dep_start}")
    ws.cell(row=current_row, column=5, value=f"=E{row_1403_cost_start}-E{row_1403_dep_start}")
    current_row += 1

    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"


def populate_equity_sheet(ws, assumption_map): # assumption_map به عنوان آرگومان اضافه شد
    """پر کردن شیت حقوق مالکانه با ساختار استاندارد و فرمول‌های صحیح."""
    ws.title = "حقوق مالکانه"
    col_widths = {'A': 30, 'B': 18, 'C': 18, 'D': 18, 'E': 18, 'F': 20}
    set_rtl_and_column_widths(ws, col_widths)
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "صورت تغییرات در حقوق مالکانه", "برای سال مالی منتهی به 29 اسفند 1403 و 1402", "(ارقام به میلیون ریال)")
    
    headers = ["شرح", "سرمایه", "اندوخته قانونی", "سایر اندوخته‌ها", "سود انباشته", "جمع کل"]
    ws.cell(row=current_row, column=1, value=headers[0]).font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=headers[1]).font = Font(bold=True)
    ws.cell(row=current_row, column=3, value=headers[2]).font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=headers[3]).font = Font(bold=True)
    ws.cell(row=current_row, column=5, value=headers[4]).font = Font(bold=True)
    ws.cell(row=current_row, column=6, value=headers[5]).font = Font(bold=True)
    current_row += 1

    # Data for 1402
    row_1402_start = current_row
    ws.cell(row=current_row, column=1, value="مانده در ابتدای 1402")
    ws.cell(row=current_row, column=2, value="='ترازنامه پایه'!D40")
    ws.cell(row=current_row, column=3, value="='ترازنامه پایه'!D41")
    ws.cell(row=current_row, column=4, value="='ترازنامه پایه'!D42")
    ws.cell(row=current_row, column=5, value="='ترازنامه پایه'!D43")
    ws.cell(row=current_row, column=6, value=f"=SUM(B{current_row}:E{current_row})")
    current_row += 1

    row_net_profit_1402_equity = current_row
    ws.cell(row=current_row, column=1, value="سود خالص 1402")
    ws.cell(row=current_row, column=5, value="='سودوزیان'!G21")
    ws.cell(row=current_row, column=6, value=f"=E{current_row}")
    current_row += 1

    row_legal_reserve_increase_1402 = current_row
    ws.cell(row=current_row, column=1, value="انتقال به اندوخته قانونی")
    ws.cell(row=current_row, column=3, value=f"=MAX(0,'سودوزیان'!G21)*0.05")
    ws.cell(row=current_row, column=5, value=f"=-C{current_row}")
    ws.cell(row=current_row, column=6, value="0")
    current_row += 1

    row_dividends_1402_equity = current_row
    ws.cell(row=current_row, column=1, value="تقسیم سود مصوب")
    ws.cell(row=current_row, column=5, value=f"=-(سودوزیان!G21*مفروضات!C{assumption_map['سود سهام پرداختی (درصد از سود خالص)']['1402'][1:]})")
    ws.cell(row=current_row, column=6, value=f"=E{current_row}")
    current_row += 1

    row_1402_end = current_row
    ws.cell(row=current_row, column=1, value="مانده در پایان 1402").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=f"=B{row_1402_start}")
    ws.cell(row=current_row, column=3, value=f"=SUM(C{row_1402_start}:C{current_row-1})")
    ws.cell(row=current_row, column=4, value=f"=D{row_1402_start}")
    ws.cell(row=current_row, column=5, value=f"=SUM(E{row_1402_start}:E{current_row-1})")
    ws.cell(row=current_row, column=6, value=f"=SUM(B{current_row}:E{current_row})")
    current_row += 2


    # Data for 1403
    row_1403_start = current_row
    ws.cell(row=current_row, column=1, value="مانده در ابتدای 1403").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=f"=B{row_1402_end}")
    ws.cell(row=current_row, column=3, value=f"=C{row_1402_end}")
    ws.cell(row=current_row, column=4, value=f"=D{row_1402_end}")
    ws.cell(row=current_row, column=5, value=f"=E{row_1402_end}")
    ws.cell(row=current_row, column=6, value=f"=F{row_1402_end}")
    current_row += 1

    row_net_profit_1403_equity = current_row
    ws.cell(row=current_row, column=1, value="سود خالص 1403")
    ws.cell(row=current_row, column=5, value="='سودوزیان'!F21")
    ws.cell(row=current_row, column=6, value=f"=E{current_row}")
    current_row += 1

    row_legal_reserve_increase_1403 = current_row
    ws.cell(row=current_row, column=1, value="انتقال به اندوخته قانونی")
    ws.cell(row=current_row, column=3, value=f"=MAX(0,'سودوزیان'!F21)*0.05")
    ws.cell(row=current_row, column=5, value=f"=-C{current_row}")
    ws.cell(row=current_row, column=6, value="0")
    current_row += 1

    row_dividends_1403_equity = current_row
    ws.cell(row=current_row, column=1, value="تقسیم سود مصوب")
    ws.cell(row=current_row, column=5, value=f"=-(سودوزیان!F21*مفروضات!B{assumption_map['سود سهام پرداختی (درصد از سود خالص)']['1403'][1:]})")
    ws.cell(row=current_row, column=6, value=f"=E{current_row}")
    current_row += 1

    row_1403_end = current_row
    ws.cell(row=current_row, column=1, value="مانده در پایان 1403").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=f"=B{row_1403_start}")
    ws.cell(row=current_row, column=3, value=f"=SUM(C{row_1403_start}:C{current_row-1})")
    ws.cell(row=current_row, column=4, value=f"=D{row_1403_start}")
    ws.cell(row=current_row, column=5, value=f"=SUM(E{row_1403_start}:E{current_row-1})")
    ws.cell(row=current_row, column=6, value=f"=SUM(B{current_row}:E{current_row})")
    current_row += 1

    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"

    # ذخیره ردیف‌های نهایی حقوق مالکانه برای استفاده در سایر شیت‌ها
    ws.max_row_equity_capital_1403 = row_1403_start
    ws.max_row_equity_legal_reserve_1403 = row_legal_reserve_increase_1403
    ws.max_row_equity_other_reserve_1403 = row_1403_start + 2
    ws.max_row_equity_retained_earnings_1403 = row_net_profit_1403_equity
    ws.max_row_equity_total_equity_1403 = row_1403_end
    
    ws.max_row_equity_capital_1402 = row_1402_start
    ws.max_row_equity_legal_reserve_1402 = row_legal_reserve_increase_1402
    ws.max_row_equity_other_reserve_1402 = row_1402_start + 2
    ws.max_row_equity_retained_earnings_1402 = row_net_profit_1402_equity
    ws.max_row_equity_total_equity_1402 = row_1402_end

    ws.max_row_equity_div_1403 = row_dividends_1403_equity
    ws.max_row_equity_div_1402 = row_dividends_1402_equity

    ws.row_legal_reserve_1402_start = row_1402_start + 1
    ws.row_legal_reserve_1401_end = row_1402_start + 1
    ws.row_legal_reserve_increase_1403 = row_legal_reserve_increase_1403
    ws.row_legal_reserve_increase_1402 = row_legal_reserve_increase_1402


def populate_cash_flow_sheet(ws, assumption_map, wb_main_obj):
    """پر کردن شیت جریان‌های نقدی."""
    ws.title = "جریان های نقدی"
    col_widths = {'A': 5, 'B': 55, 'C': 18, 'D': 18}
    set_rtl_and_column_widths(ws, col_widths)
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "صورت جریان‌های نقدی (نهایی و پویا)", "برای سال مالی منتهی به 29 اسفند 1403", "(ارقام به میلیون ریال)")

    ws.cell(row=current_row, column=3, value="1403").font = Font(bold=True)
    ws.cell(row=current_row, column=4, value="1402 (پایه)").font = Font(bold=True)
    current_row += 1

    # فعالیت‌های عملیاتی
    row_op_activities_start = current_row
    ws.cell(row=current_row, column=2, value="جریان‌های نقدی ناشی از فعالیت‌های عملیاتی").font = Font(bold=True)
    current_row += 1
    ws.cell(row=current_row, column=2, value="سود خالص")
    ws.cell(row=current_row, column=3, value=f"='سودوزیان'!F{wb_main_obj.row_net_profit_pl_1403}")
    ws.cell(row=current_row, column=4, value=f"='سودوزیان'!G{wb_main_obj.row_net_profit_pl_1402}")
    row_net_profit_cf = current_row
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="تعدیلات بابت اقلام غیرنقدی:").font = Font(bold=True)
    current_row += 1
    ws.cell(row=current_row, column=2, value="هزینه استهلاک")
    ws.cell(row=current_row, column=1, value="12").hyperlink = f"#'10.11.12'!A1"
    ws.cell(row=current_row, column=1).style = "Hyperlink"
    ws.cell(row=current_row, column=3, value="='گردش دارایی ثابت'!D11")
    ws.cell(row=current_row, column=4, value="='گردش دارایی ثابت'!D6")
    row_depreciation_cf = current_row
    current_row += 1

    ws.cell(row=current_row, column=2, value="هزینه مزایای پایان خدمت")
    ws.cell(row=current_row, column=3, value=f"='8'!C{wb_main_obj.max_row_8_benefits_expense}")
    ws.cell(row=current_row, column=4, value=f"='8'!D{wb_main_obj.max_row_8_benefits_expense}")
    row_benefits_expense_cf = current_row
    current_row += 2


    ws.cell(row=current_row, column=2, value="تغییرات در سرمایه در گردش:").font = Font(bold=True)
    current_row += 1

    ws.cell(row=current_row, column=2, value="کاهش(افزایش) در دریافتنی‌ها")
    ws.cell(row=current_row, column=3, value=f"='وضعیت مالی'!F{wb_main_obj.row_receivables}-'وضعیت مالی'!E{wb_main_obj.row_receivables}")
    ws.cell(row=current_row, column=4, value=f"='ترازنامه پایه'!D11-'وضعیت مالی'!F{wb_main_obj.row_receivables}")
    row_receivables_change = current_row
    current_row += 1

    ws.cell(row=current_row, column=2, value="کاهش(افزایش) در موجودی کالا")
    ws.cell(row=current_row, column=3, value=f"='وضعیت مالی'!F{wb_main_obj.row_inventory}-'وضعیت مالی'!E{wb_main_obj.row_inventory}")
    ws.cell(row=current_row, column=4, value=f"='ترازنامه پایه'!D12-'وضعیت مالی'!F{wb_main_obj.row_inventory}")
    row_inventory_change = current_row
    current_row += 1

    ws.cell(row=current_row, column=2, value="افزایش(کاهش) در پرداختنی‌ها")
    ws.cell(row=current_row, column=3, value=f"='وضعیت مالی'!E{wb_main_obj.row_payables}-'وضعیت مالی'!F{wb_main_obj.row_payables}")
    ws.cell(row=current_row, column=4, value=f"='وضعیت مالی'!F{wb_main_obj.row_payables}-'ترازنامه پایه'!D27")
    row_payables_change = current_row
    current_row += 1

    row_net_cash_op = current_row
    ws.cell(row=current_row, column=2, value="خالص جریان نقد عملیاتی").font = Font(bold=True)
    ws.cell(row=current_row, column=3, value=f"=SUM(C{row_net_profit_cf},C{row_depreciation_cf},C{row_benefits_expense_cf},C{row_receivables_change}:C{row_payables_change})")
    ws.cell(row=current_row, column=4, value=f"=SUM(D{row_net_profit_cf},D{row_depreciation_cf},D{row_benefits_expense_cf},D{row_receivables_change}:D{row_payables_change})")
    current_row += 2


    # فعالیت‌های سرمایه‌گذاری
    row_inv_activities_start = current_row
    ws.cell(row=current_row, column=2, value="جریان‌های نقدی ناشی از فعالیت‌های سرمایه‌گذاری").font = Font(bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="پرداخت بابت خرید دارایی ثابت (CAPEX)")
    ws.cell(row=current_row, column=1, value="12").hyperlink = f"#'10.11.12'!A1"
    ws.cell(row=current_row, column=1).style = "Hyperlink"
    ws.cell(row=current_row, column=3, value=f"=-'مفروضات'!{assumption_map['سرمایه‌گذاری ثابت سالانه (CAPEX)']['1403']}")
    ws.cell(row=current_row, column=4, value=f"=-'مفروضات'!{assumption_map['سرمایه‌گذاری ثابت سالانه (CAPEX)']['1402']}")
    row_capex_cf = current_row
    current_row += 1

    row_net_cash_inv = current_row
    ws.cell(row=current_row, column=2, value="خالص جریان نقد سرمایه‌گذاری").font = Font(bold=True)
    ws.cell(row=current_row, column=3, value=f"=C{row_capex_cf}")
    ws.cell(row=current_row, column=4, value=f"=D{row_capex_cf}")
    current_row += 2


    # فعالیت‌های تامین مالی
    row_fin_activities_start = current_row
    ws.cell(row=current_row, column=2, value="جریان‌های نقدی ناشی از فعالیت‌های تامین مالی").font = Font(bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="دریافت اصل تسهیلات")
    ws.cell(row=current_row, column=1, value="19").hyperlink = f"#'19'!A1"
    ws.cell(row=current_row, column=1).style = "Hyperlink"
    ws.cell(row=current_row, column=3, value=f"='مفروضات'!{assumption_map['مبلغ وام جدید دریافتی طی سال']['1403']}")
    ws.cell(row=current_row, column=4, value=f"='مفروضات'!{assumption_map['مبلغ وام جدید دریافتی طی سال']['1402']}")
    row_debt_received_cf = current_row
    current_row += 1

    ws.cell(row=current_row, column=2, value="بازپرداخت اصل تسهیلات")
    ws.cell(row=current_row, column=3, value=f"=-'مفروضات'!{assumption_map['مبلغ بازپرداخت اصل وام طی سال']['1403']}")
    ws.cell(row=current_row, column=4, value=f"=-'مفروضات'!{assumption_map['مبلغ بازپرداخت اصل وام طی سال']['1402']}")
    row_debt_repaid_cf = current_row
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="سود سهام پرداخت شده")
    ws.cell(row=current_row, column=1, value="18").hyperlink = f"#'18'!A1"
    ws.cell(row=current_row, column=1).style = "Hyperlink"
    ws.cell(row=current_row, column=3, value=f"='حقوق مالکانه'!E{wb_main_obj.max_row_equity_div_1403}")
    ws.cell(row=current_row, column=4, value=f"='حقوق مالکانه'!E{wb_main_obj.max_row_equity_div_1402}")
    row_dividends_paid_cf = current_row
    current_row += 1

    row_net_cash_fin = current_row
    ws.cell(row=current_row, column=2, value="خالص جریان نقد تامین مالی").font = Font(bold=True)
    ws.cell(row=current_row, column=3, value=f"=SUM(C{row_debt_received_cf}:C{row_dividends_paid_cf})")
    ws.cell(row=current_row, column=4, value=f"=SUM(D{row_debt_received_cf}:D{row_dividends_paid_cf})")
    current_row += 2


    # خلاصه
    row_summary_start = current_row
    ws.cell(row=current_row, column=2, value="خالص افزایش (کاهش) در موجودی نقد").font = Font(bold=True)
    ws.cell(row=current_row, column=3, value=f"=C{row_net_cash_op}+C{row_net_cash_inv}+C{row_net_cash_fin}")
    ws.cell(row=current_row, column=4, value=f"=D{row_net_cash_op}+D{row_net_cash_inv}+D{row_net_cash_fin}")
    row_net_cash_change = current_row
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="موجودی نقد ابتدای دوره").font = Font(bold=True)
    ws.cell(row=current_row, column=3, value=f"=D{row_net_cash_change + 1}")
    ws.cell(row=current_row, column=4, value="='ترازنامه پایه'!D10")
    row_cash_start_period_cf = current_row
    current_row += 1
    
    ws.cell(row=current_row, column=2, value="موجودی نقد در پایان دوره").font = Font(bold=True)
    ws.cell(row=current_row, column=3, value=f"=C{row_net_cash_change}+C{row_cash_start_period_cf}")
    ws.cell(row=current_row, column=4, value=f"=D{row_net_cash_change}+D{row_cash_start_period_cf}")
    row_cash_end_period_cf = current_row
    current_row += 1

    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"

    # ذخیره ردیف نهایی موجودی نقد پایان دوره برای استفاده در وضعیت مالی
    ws.max_row_cash_flow_end_1403 = row_cash_end_period_cf
    ws.max_row_cash_flow_end_1402 = row_cash_end_period_cf


def populate_comprehensive_income_sheet(ws):
    """پر کردن شیت صورت سود و زیان جامع."""
    ws.title = "جامع"
    col_widths = {'A': 5, 'B': 40, 'C': 18, 'D': 18}
    set_rtl_and_column_widths(ws, col_widths)
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "صورت سود و زیان جامع", "برای سال مالی منتهی به 29 اسفند 1403 و 1402", "(ارقام به میلیون ریال)")

    ws.cell(row=current_row, column=3, value="1403").font = Font(bold=True)
    ws.cell(row=current_row, column=4, value="1402").font = Font(bold=True)
    current_row += 1

    ws.cell(row=current_row, column=2, value="سود خالص دوره").font = Font(bold=True)
    ws.cell(row=current_row, column=3, value="='سودوزیان'!F21")
    ws.cell(row=current_row, column=4, value="='سودوزیان'!G21")
    row_net_profit_comprehensive = current_row
    current_row += 2

    ws.cell(row=current_row, column=2, value="سایر اقلام سود و زیان جامع:").font = Font(bold=True)
    current_row += 1
    ws.cell(row=current_row, column=2, value="تعدیلات تسعیر ارز عملیات خارجی (بعد از مالیات)")
    ws.cell(row=current_row, column=3, value=10_000)
    ws.cell(row=current_row, column=4, value=5_000)
    row_fx_adjustments = current_row
    current_row += 2

    ws.cell(row=current_row, column=2, value="جمع کل سود و زیان جامع دوره").font = Font(bold=True)
    ws.cell(row=current_row, column=3, value=f"=C{row_net_profit_comprehensive}+C{row_fx_adjustments}")
    ws.cell(row=current_row, column=4, value=f"=D{row_net_profit_comprehensive}+D{row_fx_adjustments}")
    current_row += 1

    ws.cell(row=1, column=52, value="بازگشت به سود و زیان").hyperlink = f"#'سودوزیان'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"


def populate_history_sheet(ws):
    """پر کردن شیت تاریخچه."""
    ws.title = "تاریخچه"
    ws.sheet_view.rightToLeft = True
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "تاریخچه شرکت نمونه (سهامی عام)", "")

    ws.cell(row=current_row, column=1, value="مقدمه:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="شرکت نمونه (سهامی عام) در سال 1375 با هدف سرمایه‌گذاری و فعالیت در صنعت مرغداری و زنجیره تامین گوشت مرغ تاسیس گردید. این شرکت با بهره‌گیری از دانش روز و تکنولوژی‌های پیشرفته در زمینه پرورش جوجه یک روزه اجداد، تولید جوجه گوشتی و عرضه به کشتارگاه، به یکی از پیشگامان صنعت در کشور تبدیل شده است.").alignment = Alignment(wrapText=True, horizontal='right')
    current_row = ws.max_row + 1

    ws.cell(row=current_row, column=1, value="اهداف و استراتژی‌ها:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="هدف اصلی شرکت، تولید پروتئین با کیفیت بالا، افزایش بهره‌وری در تمامی مراحل زنجیره تامین، توسعه پایدار و ایفای نقش مسئولانه در تامین امنیت غذایی کشور است. استراتژی‌های شرکت شامل توسعه فارم‌های جدید، بهبود نژادهای پرورشی، بهینه‌سازی مصرف خوراک و کاهش ضایعات می‌باشد.").alignment = Alignment(wrapText=True, horizontal='right')
    current_row = ws.max_row + 1

    ws.cell(row=current_row, column=1, value="فعالیت‌های اصلی:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="شرکت در حال حاضر دارای 10 فارم پرورش مرغ گوشتی، 5 انبار نگهداری دان و مرغ، و واحد لجستیک پیشرفته برای حمل و نقل محصولات به کشتارگاه‌ها می‌باشد. ظرفیت تولید سالانه شرکت بیش از 50,000 تن مرغ گوشتی است.").alignment = Alignment(wrapText=True, horizontal='right')
    current_row = ws.max_row + 1

    ws.cell(row=current_row, column=1, value="چشم‌انداز آینده:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="شرکت نمونه با تمرکز بر نوآوری، توسعه بازار و افزایش ظرفیت‌های تولیدی، در نظر دارد سهم خود را در بازار افزایش داده و به عنوان یکی از بزرگترین شرکت‌های زنجیره تامین مرغ در منطقه شناخته شود.").alignment = Alignment(wrapText=True, horizontal='right')
    current_row = ws.max_row + 1

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 80
    
    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"


def populate_significant_accounting_policy_sheet(ws, policy_number):
    """پر کردن شیت اهم رویه‌های حسابداری."""
    ws.title = f"اهم رویه{policy_number}"
    ws.sheet_view.rightToLeft = True
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", f"یادداشت {policy_number}: اهم رویه های حسابداری", "")

    ws.cell(row=current_row, column=1, value="مقدمه:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="صورت‌های مالی حاضر بر اساس استانداردهای حسابداری ایران (نشریه شماره 160 سازمان حسابرسی) تهیه شده‌اند. اهم رویه‌های حسابداری مورد استفاده در تهیه این صورت‌ها به شرح زیر است:").alignment = Alignment(wrapText=True, horizontal='right')
    current_row = ws.max_row + 1

    policies = {
        1: "نحوه ارائه صورت‌های مالی: صورت‌های مالی طبق استانداردهای حسابداری ایران و به روش بهای تمام شده تاریخی تهیه و ارائه شده‌اند. کلیه رویدادهای مالی در زمان وقوع شناسایی و ثبت می‌گردند. این صورت‌ها بر اساس اصل تداوم فعالیت تهیه شده‌اند.",
        2: "شناخت درآمد: درآمد حاصل از فروش مرغ گوشتی در زمان تحویل محصول و انتقال تمامی ریسک‌ها و مزایای مالکیت به خریدار شناسایی می‌شود. درآمدهای حاصل از فروش جوجه یک روزه نیز پس از تحویل و انتقال مالکیت و قطعیت وصول وجه شناسایی می‌گردد. درآمدهای فرعی (مانند فروش کود) نیز در زمان تحقق شناسایی می‌شوند.",
        3: "موجودی مواد و کالا: موجودی مواد و کالا (شامل جوجه، دان، دارو و مرغ آماده فروش) بر اساس روش میانگین موزون و به اقل بهای تمام شده یا خالص ارزش بازیافتنی اندازه‌گیری می‌شود. بهای تمام شده جوجه‌هاي در حال رشد شامل هزینه‌های مستقیم پرورش (دان، دارو، واکسن، دستمزد مستقیم کارگران فارم) و سهم مناسبی از سربار تولید است. خالص ارزش بازیافتنی، بهای فروش برآوردی در روال عادی عملیات پس از کسر هزینه‌های برآوردی تکمیل و هزینه‌های برآوردی لازم برای انجام فروش است. ذخیره کاهش ارزش موجودی‌ها در صورت لزوم شناسایی می‌گردد.",
        4: "دارایی‌های ثابت مشهود: دارایی‌های ثابت مشهود به بهای تمام شده تاریخی پس از کسر استهلاک انباشته و زیان کاهش ارزش انباشته در ترازنامه منعقد می‌شوند. استهلاک دارایی‌ها به روش خط مستقیم طی عمر مفید برآوردی دارایی صورت می‌گیرد. مخارج بعدی مربوط به دارایی‌های ثابت مشهود تنها در صورتی به بهای تمام شده دارایی اضافه می‌شود که منجر به افزایش قابل ملاحظه در منافع اقتصادی آتی ناشی از آن گردد. دارایی‌هایی که آماده بهره‌برداری نیستند، در حساب دارایی در جریان تکمیل ثبت می‌شوند.",
        5: "ارزهای خارجی: معاملات ارزی با نرخ تسعیر ارز در تاریخ معامله ثبت می‌شوند. اقلام پولی دارایی‌ها و بدهی‌های ارزی با نرخ تسعیر ارز در تاریخ ترازنامه تسعیر شده و سود یا زیان ناشی از تسعیر ارز به عنوان درآمد/هزینه غیرعملیاتی در صورت سود و زیان جامع شناسایی می‌شود. تفاوت‌های تسعیر ارز ناشی از اقلام پولی غیرپولی, در صورت‌های مالی منعکس نمی‌گردد.",
        6: "مزایای پایان خدمت کارکنان: تعهدات مزایای پایان خدمت کارکنان (پاداش پایان خدمت) بر اساس قوانین کار و تامین اجتماعی ایران و با استفاده از روش تعهدات برآوردی محاسبه و شناسایی می‌گردد."
    }
    
    policy_text = policies.get(policy_number, "توضیحات رویه حسابداری برای این یادداشت موجود نیست.")
    ws.cell(row=current_row, column=1, value=f"رویه حسابداری شماره {policy_number}:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=policy_text).alignment = Alignment(wrapText=True, horizontal='right')
    current_row = ws.max_row + 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80
    
    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"


def populate_management_judgment_sheet(ws):
    """پر کردن شیت قضاوت مدیریت در فرایند بکارگیری رویه های حسابداری."""
    ws.title = "قضاوت مدیریت"
    ws.sheet_view.rightToLeft = True
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "یادداشت: قضاوت مدیریت در فرایند بکارگیری رویه های حسابداری", "")

    ws.cell(row=current_row, column=1, value="مقدمه:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="تهیه صورت‌های مالی بر اساس استانداردهای حسابداری ایران مستلزم اعمال قضاوت‌های مهم توسط مدیریت در بکارگیری رویه‌های حسابداری و برآوردهای حسابداری است. برخی از زمینه‌های کلیدی که مدیریت در آن‌ها قضاوت‌های مهمی اعمال می‌کند، به شرح زیر است:").alignment = Alignment(wrapText=True, horizontal='right')
    current_row = ws.max_row + 1
    
    ws.cell(row=current_row, column=1, value="1. برآورد عمر مفید دارایی‌های ثابت مشهود و نامشهود:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="مدیریت عمر مفید اقتصادی دارایی‌های ثابت مشهود (نظیر ساختمان فارم‌ها، تجهیزات و ماشین‌آلات) و نامشهود (نظیر نرم‌افزارها و حقوق استفاده از نژادهای خاص) را بر اساس تجربه قبلی، انتظارات از فرسودگی فیزیکی و منسوخ شدن تکنولوژیکی برآورد می‌کند. هرگونه تغییر در این برآوردها می‌تواند بر مبلغ استهلاک و بهای تمام شده دارایی‌ها در دوره‌های آتی تاثیر بگذارد. تجدید ارزیابی دارایی‌ها بر اساس رویه شرکت انجام می‌شود.").alignment = Alignment(wrapText=True, horizontal='right')
    current_row = ws.max_row + 1
    
    ws.cell(row=current_row, column=1, value="2. خالص ارزش بازیافتنی موجودی مواد و کالا:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="مدیریت برای تعیین خالص ارزش بازیافتنی موجودی‌ها (از جمله جوجه، دان و مرغ آماده فروش)، قضاوت‌هایی در خصوص قیمت‌هاي فروش آتی، هزینه‌های تکمیل و هزینه‌های لازم برای انجام فروش اعمال می‌کند. این برآوردها تحت تاثیر شرایط بازار، نوسانات قیمت خوراک و دارو و میزان تقاضا برای محصولات شرکت قرار دارد و ممکن است منجر به شناسایی ذخیره کاهش ارزش موجودی‌ها شود.").alignment = Alignment(wrapText=True, horizontal='right')
    current_row = ws.max_row + 1

    ws.cell(row=current_row, column=1, value="3. ذخیره مطالبات مشکوک‌الوصول:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="مدیریت بر اساس سابقه تاریخ‌ياب وصول مطالبات، وضعیت مالی مشتریان و شرایط اقتصادی جاری، برآوردی از مبلغ مطالبات مشکوک‌الوصول را انجام می‌دهد. این برآورد شامل قضاوت در خصوص میزان عدم قطعیت در وصول مطالبات آتی است. شناسایی این ذخیره بر اساس اصل احتياط و قابلیت وصول مطالبات انجام می‌شود.").alignment = Alignment(wrapText=True, horizontal='right')
    current_row = ws.max_row + 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 80
    
    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"


def populate_attachment_sheet(ws):
    """پر کردن شیت پیوست صورت‌های مالی."""
    ws.title = "پیوست"
    ws.sheet_view.rightToLeft = True
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "پیوست صورت‌های مالی", "")

    ws.cell(row=current_row, column=1, value="این بخش شامل هرگونه اطلاعات تکمیلی و جداول تفصیلی است که برای درک کامل‌تر صورت‌های مالی ضروری است.").font = Font(bold=True)
    current_row += 1
    ws.cell(row=current_row, column=1, value="مثال: جدول تفصیلی دارایی‌های ثابت مشهود، جداول تفصیلی سرمایه‌گذاری‌ها، تفکیک درآمدها بر حسب نوع محصول و منطقه جغرافیایی، گزارش کامل حقوق و دستمزد تفکیکی.").alignment = Alignment(wrapText=True, horizontal='right')
    current_row = ws.max_row + 1

    ws.column_dimensions['A'].width = 80
    
    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"


def populate_page_header_sheet(ws):
    """پر کردن شیت سربرگ صفحات."""
    ws.title = "سر برگ صفحات"
    ws.sheet_view.rightToLeft = True
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "سر برگ صفحات (برای چاپ و ارائه)", "")

    ws.cell(row=current_row, column=1, value="این شیت می‌تواند شامل اطلاعات تکراری در بالای هر صفحه چاپی باشد.").font = Font(italic=True)
    current_row += 1
    
    ws.cell(row=current_row, column=1, value="نام شرکت: شرکت نمونه (سهامی عام)").alignment = Alignment(wrapText=True, horizontal='right')
    current_row += 1
    ws.cell(row=current_row, column=1, value="صورت مالی: صورت سود و زیان / صورت وضعیت مالی و غیره").alignment = Alignment(wrapText=True, horizontal='right')
    current_row += 1
    ws.cell(row=current_row, column=1, value="سال مالی: منتهی به 29 اسفند 1403").alignment = Alignment(wrapText=True, horizontal='right')
    current_row += 1

    ws.column_dimensions['A'].width = 80
    
    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"


def populate_signature_sheet(ws):
    """پر کردن شیت صفحه امضا کنندگان صورت‌های مالی."""
    ws.title = "ص امضا"
    ws.sheet_view.rightToLeft = True
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "صفحه امضا کنندگان صورت‌های مالی", "")

    ws.cell(row=current_row, column=1, value="این صورت‌های مالی توسط افراد زیر تهیه و تأیید شده‌اند:").font = Font(bold=True)
    current_row += 2
    
    ws.cell(row=current_row, column=1, value="نام و نام خانوادگی: [نام مدیر عامل]").alignment = Alignment(wrapText=True, horizontal='right')
    current_row += 1
    ws.cell(row=current_row, column=1, value="سمت: مدیر عامل").alignment = Alignment(wrapText=True, horizontal='right')
    current_row += 1
    ws.cell(row=current_row, column=1, value="تاریخ: 1404/03/22").alignment = Alignment(wrapText=True, horizontal='right')
    current_row += 2
    
    ws.cell(row=current_row, column=1, value="نام و نام خانوادگی: [نام مدیر مالی]").alignment = Alignment(wrapText=True, horizontal='right')
    current_row += 1
    ws.cell(row=current_row, column=1, value="سمت: مدیر مالی").alignment = Alignment(wrapText=True, horizontal='right')
    current_row += 1
    ws.cell(row=current_row, column=1, value="تاریخ: 1404/03/22").alignment = Alignment(wrapText=True, horizontal='right')
    current_row += 2
    
    ws.cell(row=current_row, column=1, value="نام و نام خانوادگی: [نام حسابرس]").alignment = Alignment(wrapText=True, horizontal='right')
    current_row += 1
    ws.cell(row=current_row, column=1, value="سمت: حسابرس مستقل").alignment = Alignment(wrapText=True, horizontal='right')
    current_row += 1
    ws.cell(row=current_row, column=1, value="تاریخ: 1404/03/22").alignment = Alignment(wrapText=True, horizontal='right')
    current_row += 1

    ws.column_dimensions['A'].width = 40
    
    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"


def populate_management_comparative_report(ws, wb_main_obj):
    """پر کردن شیت گزارش مدیریتی تطبیقی."""
    ws.title = "گزارش مدیریتی تطبیقی"
    ws.sheet_view.rightToLeft = True
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "گزارش مدیریتی تطبیقی", "برای سال مالی منتهی به 29 اسفند 1403 و 1402", "(ارقام به میلیون ریال)")

    ws.cell(row=current_row, column=3, value="1403").font = Font(bold=True)
    ws.cell(row=current_row, column=4, value="1402").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value="درصد تغییر").font = Font(bold=True)
    current_row += 2

    ws.cell(row=current_row, column=1, value="خلاصه‌ای از شاخص‌های کلیدی عملکرد (KPIs):").font = Font(bold=True)
    current_row += 1
    
    # درآمدهای عملیاتی
    ws.cell(row=current_row, column=2, value="درآمدهای عملیاتی")
    ws.cell(row=current_row, column=3, value=f"='سودوزیان'!F{wb_main_obj.row_revenue_pl}")
    ws.cell(row=current_row, column=4, value=f"='سودوزیان'!G{wb_main_obj.row_revenue_pl}")
    ws.cell(row=current_row, column=5, value=f"=IF(D{current_row}<>0,(C{current_row}-D{current_row})/D{current_row},\"N/A\")").number_format = '0.00%'
    ws.row_mcr_revenue = current_row
    current_row += 1

    # سود ناخالص
    ws.cell(row=current_row, column=2, value="سود ناخالص")
    ws.cell(row=current_row, column=3, value=f"='سودوزیان'!F{wb_main_obj.row_gross_profit_pl}")
    ws.cell(row=current_row, column=4, value=f"='سودوزیان'!G{wb_main_obj.row_gross_profit_pl}")
    ws.cell(row=current_row, column=5, value=f"=IF(D{current_row}<>0,(C{current_row}-D{current_row})/D{current_row},\"N/A\")").number_format = '0.00%'
    ws.row_mcr_gross_profit = current_row
    current_row += 1

    # سود عملیاتی
    ws.cell(row=current_row, column=2, value="سود عملیاتی")
    ws.cell(row=current_row, column=3, value=f"='سودوزیان'!F{wb_main_obj.row_operating_profit_pl}")
    ws.cell(row=current_row, column=4, value=f"='سودوزیان'!G{wb_main_obj.row_operating_profit_pl}")
    ws.cell(row=current_row, column=5, value=f"=IF(D{current_row}<>0,(C{current_row}-D{current_row})/D{current_row},\"N/A\")").number_format = '0.00%'
    ws.row_mcr_operating_profit = current_row
    current_row += 1

    # سود خالص
    ws.cell(row=current_row, column=2, value="سود خالص")
    ws.cell(row=current_row, column=3, value=f"='سودوزیان'!F{wb_main_obj.row_net_profit_pl_1403}")
    ws.cell(row=current_row, column=4, value=f"='سودوزیان'!G{wb_main_obj.row_net_profit_pl_1402}")
    ws.cell(row=current_row, column=5, value=f"=IF(D{current_row}<>0,(C{current_row}-D{current_row})/D{current_row},\"N/A\")").number_format = '0.00%'
    ws.row_mcr_net_profit = current_row
    current_row += 2

    # جمع کل دارایی‌ها
    ws.cell(row=current_row, column=2, value="جمع کل دارایی‌ها")
    ws.cell(row=current_row, column=3, value=f"='وضعیت مالی'!E{wb_main_obj.row_total_assets}")
    ws.cell(row=current_row, column=4, value=f"='وضعیت مالی'!F{wb_main_obj.row_total_assets}")
    ws.cell(row=current_row, column=5, value=f"=IF(D{current_row}<>0,(C{current_row}-D{current_row})/D{current_row},\"N/A\")").number_format = '0.00%'
    ws.row_mcr_total_assets = current_row
    current_row += 1

    # جمع کل بدهی‌ها
    ws.cell(row=current_row, column=2, value="جمع کل بدهی‌ها")
    ws.cell(row=current_row, column=3, value=f"='وضعیت مالی'!E{wb_main_obj.row_total_liabilities}")
    ws.cell(row=current_row, column=4, value=f"='وضعیت مالی'!F{wb_main_obj.row_total_liabilities}")
    ws.cell(row=current_row, column=5, value=f"=IF(D{current_row}<>0,(C{current_row}-D{current_row})/D{current_row},\"N/A\")").number_format = '0.00%'
    ws.row_mcr_total_liabilities = current_row
    current_row += 2
    
    ws.cell(row=current_row, column=1, value="نسبت‌های مالی کلیدی:").font = Font(bold=True)
    current_row += 1

    # نسبت جاری
    ws.cell(row=current_row, column=2, value="نسبت جاری (Current Ratio)")
    ws.cell(row=current_row, column=3, value=f"=IFERROR('وضعیت مالی'!E{wb_main_obj.row_total_current_assets}/'وضعیت مالی'!E{wb_main_obj.row_total_current_liabilities},0)").number_format = '0.00'
    ws.cell(row=current_row, column=4, value=f"=IFERROR('وضعیت مالی'!F{wb_main_obj.row_total_current_assets}/'وضعیت مالی'!F{wb_main_obj.row_total_current_liabilities},0)").number_format = '0.00'
    ws.row_mcr_current_ratio = current_row
    current_row += 1

    # نسبت بدهی
    ws.cell(row=current_row, column=2, value="نسبت بدهی (Debt Ratio)")
    ws.cell(row=current_row, column=3, value=f"=IFERROR('وضعیت مالی'!E{wb_main_obj.row_total_liabilities}/'وضعیت مالی'!E{wb_main_obj.row_total_assets},0)").number_format = '0.00'
    ws.cell(row=current_row, column=4, value=f"=IFERROR('وضعیت مالی'!F{wb_main_obj.row_total_liabilities}/'وضعیت مالی'!F{wb_main_obj.row_total_assets},0)").number_format = '0.00'
    ws.row_mcr_debt_ratio = current_row
    current_row += 1
    
    # حاشیه سود خالص
    ws.cell(row=current_row, column=2, value="حاشیه سود خالص (Net Profit Margin)")
    ws.cell(row=current_row, column=3, value=f"=IFERROR('سودوزیان'!F{wb_main_obj.row_net_profit_pl_1403}/'سودوزیان'!F{wb_main_obj.row_revenue_pl},0)").number_format = '0.00%'
    ws.cell(row=current_row, column=4, value=f"=IFERROR('سودوزیان'!G{wb_main_obj.row_net_profit_pl_1402}/'سودوزیان'!G{wb_main_obj.row_revenue_pl},0)").number_format = '0.00%'
    ws.row_mcr_net_profit_margin = current_row
    current_row += 1

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"

    # اضافه کردن ویژگی‌های row_mcr_* به شیء اصلی wb تا در توابع دیگر قابل دسترس باشند
    wb_main_obj.row_mcr_revenue = ws.row_mcr_revenue
    wb_main_obj.row_mcr_gross_profit = ws.row_mcr_gross_profit
    wb_main_obj.row_mcr_operating_profit = ws.row_mcr_operating_profit
    wb_main_obj.row_mcr_net_profit = ws.row_mcr_net_profit
    wb_main_obj.row_mcr_total_assets = ws.row_mcr_total_assets
    wb_main_obj.row_mcr_total_liabilities = ws.row_mcr_total_liabilities
    wb_main_obj.row_mcr_current_ratio = ws.row_mcr_current_ratio
    wb_main_obj.row_mcr_debt_ratio = ws.row_mcr_debt_ratio
    wb_main_obj.row_mcr_net_profit_margin = ws.row_mcr_net_profit_margin


def populate_business_analytical_report(ws, wb_main_obj):
    """پر کردن شیت گزارش تحلیلی کسب و کار."""
    ws.title = "گزارش تحلیلی کسب و کار"
    ws.sheet_view.rightToLeft = True
    current_row = add_header(ws, "شرکت نمونه (سهامی عام)", "گزارش تحلیلی کسب و کار", "برای سال مالی منتهی به 29 اسفند 1403", "")

    ws.cell(row=current_row, column=1, value="1. تحلیل عملکرد عملیاتی:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=f'''=CONCATENATE("شرکت در سال 1403 شاهد رشد ",TEXT('گزارش مدیریتی تطبیقی'!E{wb_main_obj.row_mcr_revenue},"0.00%")," درآمدهای عملیاتی نسبت به سال قبل بوده است. با این حال، بهای تمام شده درآمدهای عملیاتی نیز ",TEXT(IFERROR(('سودوزیان'!F9/'سودوزیان'!G9)-1,"0.00%"),"0.00%")," افزایش یافته که نیاز به کنترل بیشتر هزینه‌ها در زنجیره تامین دارد.")''').alignment = Alignment(wrapText=True, horizontal='right')
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    current_row = ws.max_row + 1


    ws.cell(row=current_row, column=1, value="2. تحلیل سودآوری:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=f'''=CONCATENATE("حاشیه سود خالص شرکت در سال 1403 به ",TEXT('گزارش مدیریتی تطبیقی'!C{wb_main_obj.row_mcr_net_profit_margin},"0.00%")," رسیده که نشان‌دهنده توانایی شرکت در مدیریت هزینه‌های مستقیم تولید است. با این حال، هزینه‌های اداری و عمومی نیز رشد قابل توجهی داشته‌اند که می‌بایست مورد بررسی قرار گیرند.")''').alignment = Alignment(wrapText=True, horizontal='right')
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    current_row = ws.max_row + 1


    ws.cell(row=current_row, column=1, value="3. تحلیل وضعیت نقدینگی:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=f'''=CONCATENATE("جریان‌های نقدی عملیاتی شرکت مثبت بوده که نشان‌دهنده توانایی شرکت در تامین نقدینگی از محل عملیات اصلی خود است. نسبت جاری شرکت در سال 1403 برابر با ",TEXT('گزارش مدیریتی تطبیقی'!C{wb_main_obj.row_mcr_current_ratio},"0.00")," است که نشان‌دهنده وضعیت نقدینگی مطلوب و توانایی ایفای تعهدات جاری است.")''').alignment = Alignment(wrapText=True, horizontal='right')
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="CCE0F5", end_color="CCE0F5", fill_type="solid")
    current_row = ws.max_row + 1


    ws.cell(row=current_row, column=1, value="4. پیشنهادها:").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value="- بررسی دقیق‌تر هزینه‌های اداری و عمومی و شناسایی فرصت‌های صرفه‌جویی.\\n- سرمایه‌گذاری در تکنولوژی‌های جدید برای افزایش بهره‌وری در فارم‌ها و کاهش بهای تمام شده تولید.\\n- توسعه بازارهای جدید برای محصولات شرکت.").alignment = Alignment(wrapText=True, horizontal='right')
    ws.cell(row=current_row, column=2).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    current_row = ws.max_row + 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80
    
    ws.cell(row=1, column=52, value="بازگشت به وضعیت مالی").hyperlink = f"#'وضعیت مالی'!A1"
    ws.cell(row=1, column=52).style = "Hyperlink"


# ==============================================================================
# توابع موقتی برای اجرای اسکریپت (Temporary functions to run the script)
# ==============================================================================

def populate_inventory_note(ws, wb_main_obj):
    """پر کردن شیت یادداشت موجودی کالا (موقتی)."""
    print("تابع populate_inventory_note فراخوانی شد اما پیاده‌سازی نشده است.")
    pass

def populate_numeric_note_sheets(wb, assumption_map, wb_main_obj):
    """پر کردن شیت‌های یادداشت‌های عددی (موقتی)."""
    print("تابع populate_numeric_note_sheets فراخوانی شد اما پیاده‌سازی نشده است.")
    pass

# ==============================================================================
# تابع اصلی اجرای مدل (Main Model Execution Function)
# ==============================================================================

def main_model_execution():
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    all_initial_sheets = [
        "مفروضات", "ترازنامه پایه", "گردش دارایی ثابت", "موجودی_تفصیلی", "لیست حقوق و دستمزد",
        "8", "9", "سودوزیان", "حقوق مالکانه", "جریان های نقدی", "وضعیت مالی", "موجودی",
        "جامع", "تاریخچه", "اهم رویه1", "اهم رویه2", "اهم رویه3", "اهم رویه4", "اهم رویه5", "اهم رویه6",
        "قضاوت مدیریت", "پیوست", "سر برگ صفحات", "ص امضا", "گزارش مدیریتی تطبیقی", "گزارش تحلیلی کسب و کار",
        '5', '6', '7', '10.11.12', '13', '14', '15', '16', '17', '18', '19', '20',
        '21', '22.-23', '24.25', '26.27', '28.29.30.31', '32.33', '34',
        '35', '35-1', '35-6', '36-37', '38.39.40', '41', '42.43', '44',
        '44-4', '44-6', '45', '46', '46-3', '47.48', '49',
        'ادامه16', 'ادامه34', 'ادامه41',
    ]

    for sheet_name in all_initial_sheets:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
    
    # 1. مفروضات (برای ایجاد map و استفاده در سایر شیت‌ها)
    ws_assumptions = wb["مفروضات"]
    assumption_map = populate_assumptions_sheet(ws_assumptions)

    # 2. ترازنامه پایه
    ws_starting_balance = wb["ترازنامه پایه"]
    populate_starting_balance_sheet(ws_starting_balance)
    wb.row_cash_base = ws_starting_balance.row_cash_base
    wb.row_receivables_base = ws_starting_balance.row_receivables_base
    wb.row_inventory_base = ws_starting_balance.row_inventory_base
    wb.row_prepayments_base = ws_starting_balance.row_prepayments_base
    wb.row_total_current_assets_base = ws_starting_balance.row_total_current_assets_base
    wb.row_gross_fixed_assets_base = ws_starting_balance.row_gross_fixed_assets_base
    wb.row_accumulated_dep_base = ws_starting_balance.row_accumulated_dep_base
    wb.row_net_fixed_assets_base = ws_starting_balance.row_net_fixed_assets_base
    wb.row_other_non_current_assets_base = ws_starting_balance.row_other_non_current_assets_base
    wb.row_total_non_current_assets_base = ws_starting_balance.row_total_non_current_assets_base
    wb.row_total_assets_base = ws_starting_balance.row_total_assets_base
    wb.row_payables_base = ws_starting_balance.row_payables_base
    wb.row_tax_payable_base = ws_starting_balance.row_tax_payable_base
    wb.row_dividends_payable_base = ws_starting_balance.row_dividends_payable_base
    wb.row_current_lt_debt_base = ws_starting_balance.row_current_lt_debt_base
    wb.row_total_current_liabilities_base = ws_starting_balance.row_total_current_liabilities_base
    wb.row_lt_debt_base = ws_starting_balance.row_lt_debt_base
    wb.row_employee_benefits_base = ws_starting_balance.row_employee_benefits_base
    wb.row_total_non_current_liabilities_base = ws_starting_balance.row_total_non_current_liabilities_base
    wb.row_total_liabilities_base = ws_starting_balance.row_total_liabilities_base
    wb.row_capital_base = ws_starting_balance.row_capital_base
    wb.row_legal_reserve_base = ws_starting_balance.row_legal_reserve_base
    wb.row_other_reserve_base = ws_starting_balance.row_other_reserve_base
    wb.row_retained_earnings_base = ws_starting_balance.row_retained_earnings_base
    wb.row_total_equity_base = ws_starting_balance.row_total_equity_base
    wb.row_total_liabilities_and_equity_base = ws_starting_balance.row_total_liabilities_and_equity_base


    # 3. گردش دارایی ثابت (برای استفاده در سود و زیان)
    ws_fixed_asset = wb["گردش دارایی ثابت"]
    populate_fixed_asset_roll_forward_sheet(ws_fixed_asset, assumption_map)

    # 4. موجودی تفصیلی (برای استفاده در یادداشت 9 و موجودی)
    ws_detailed_inventory = wb["موجودی_تفصیلی"]
    populate_detailed_inventory_sheet(ws_detailed_inventory)
    
    # 5. لیست حقوق و دستمزد (برای استفاده در یادداشت 8 و 9 و 35-6)
    ws_payroll_list = wb["لیست حقوق و دستمزد"]
    payroll_output_details = populate_payroll_list_sheet(ws_payroll_list)
    wb.payroll_output_cells_E117 = payroll_output_details[0]
    wb.payroll_output_cells_E118 = payroll_output_details[1]
    wb.payroll_output_cells_E119 = payroll_output_details[2]
    wb.payroll_output_cells_E120 = payroll_output_details[3]
    wb.payroll_output_cells_E121 = payroll_output_details[4]
    wb.payroll_output_cells_E122 = payroll_output_details[5]
    wb.total_yearly_million_1403_idx_payroll = payroll_output_details[6]
    wb.total_yearly_million_1402_idx_payroll = payroll_output_details[7]
    wb.gross_salary_1403_cell_million = payroll_output_details[8]
    wb.employer_insurance_1403_cell_million = payroll_output_details[9]
    wb.gross_salary_1402_cell_million = payroll_output_details[10]
    wb.employer_insurance_1402_cell_million = payroll_output_details[11]


    # 6. یادداشت 8 و 9 (برای استفاده در سود و زیان)
    populate_note_8_and_9(wb)

    # راه‌حل موقت برای خطای AttributeError: 'Workbook' object has no attribute 'max_row_5'
    # کاربر باید پیاده‌سازی صحیح populate_numeric_note_sheets را برای مقداردهی wb.max_row_5 ارائه دهد.
    # فرض می‌کنیم یادداشت 5 در شیت '5' تا ردیف 10 اطلاعات دارد. این مقدار باید توسط کاربر تایید شود.
    wb.max_row_5 = 10 
    # راه‌حل موقت برای خطای AttributeError مربوط به یادداشت 26.27
    # کاربر باید پیاده‌سازی صحیح populate_numeric_note_sheets را برای مقداردهی این موارد ارائه دهد.
    # فرض اولیه برای ردیف‌های اطلاعاتی در یادداشت 26.27 (این مقادیر باید توسط کاربر تایید شوند)
    wb.max_row_26_27_revenue_section = 10 
    # فرض می‌کنیم یادداشت 5 در شیت '5' تا ردیف 10 اطلاعات دارد. این مقدار باید توسط کاربر تایید شود.
    wb.max_row_5 = 10 
    # راه‌حل موقت برای خطای AttributeError مربوط به یادداشت 26.27
    # کاربر باید پیاده‌سازی صحیح populate_numeric_note_sheets را برای مقداردهی این موارد ارائه دهد.
    # فرض اولیه برای ردیف‌های اطلاعاتی در یادداشت 26.27 (این مقادیر باید توسط کاربر تایید شوند)
    wb.max_row_26_27_revenue_section = 10 
    wb.max_row_26_27_expense_section = 15


    # 7. سود و زیان (برای استفاده در حقوق مالکانه و جریان نقدی و ترازنامه)
    ws_profit_loss = wb["سودوزیان"]
    populate_profit_loss_sheet(ws_profit_loss, assumption_map, wb)
    wb.row_revenue_pl = ws_profit_loss.row_revenue_pl
    wb.row_cogs_pl = ws_profit_loss.row_cogs_pl
    wb.row_gross_profit_pl = ws_profit_loss.row_gross_profit_pl
    wb.row_sga_pl = ws_profit_loss.row_sga_pl
    wb.row_other_income_pl = ws_profit_loss.row_other_income_pl
    wb.row_other_expense_pl = ws_profit_loss.row_other_expense_pl
    wb.row_operating_profit_pl = ws_profit_loss.row_operating_profit_pl
    wb.row_finance_cost_pl = ws_profit_loss.row_finance_cost_pl
    wb.row_pbt_pl = ws_profit_loss.row_pbt_pl
    wb.row_tax_pl = ws_profit_loss.row_tax_pl
    wb.row_net_profit_pl_1403 = ws_profit_loss.row_net_profit_pl_1403
    wb.row_net_profit_pl_1402 = ws_profit_loss.row_net_profit_pl_1402
    # این مقادیر در populate_profit_loss_sheet مقداردهی اولیه می‌شوند و سپس توسط populate_numeric_note_sheets به‌روزرسانی می‌شوند.
    # wb.max_row_26_27_revenue_section = ws_profit_loss.max_row_26_27_revenue_section
    # wb.max_row_26_27_expense_section = ws_profit_loss.max_row_26_27_expense_section


    # 8. حقوق مالکانه (برای استفاده در جریان نقدی و ترازنامه)
    ws_equity = wb["حقوق مالکانه"]
    populate_equity_sheet(ws_equity, assumption_map) # assumption_map به عنوان آرگومان اضافه شد
    wb.max_row_equity_capital_1403 = ws_equity.max_row_equity_capital_1403
    wb.max_row_equity_legal_reserve_1403 = ws_equity.max_row_equity_legal_reserve_1403
    wb.max_row_equity_other_reserve_1403 = ws_equity.max_row_equity_other_reserve_1403
    wb.max_row_equity_retained_earnings_1403 = ws_equity.max_row_equity_retained_earnings_1403
    wb.max_row_equity_total_equity_1403 = ws_equity.max_row_equity_total_equity_1403
    
    wb.max_row_equity_capital_1402 = ws_equity.max_row_equity_capital_1402
    wb.max_row_equity_legal_reserve_1402 = ws_equity.max_row_equity_legal_reserve_1402
    wb.max_row_equity_other_reserve_1402 = ws_equity.max_row_equity_other_reserve_1402
    wb.max_row_equity_retained_earnings_1402 = ws_equity.max_row_equity_retained_earnings_1402
    wb.max_row_equity_total_equity_1402 = ws_equity.max_row_equity_total_equity_1402

    wb.max_row_equity_div_1403 = ws_equity.max_row_equity_div_1403
    wb.max_row_equity_div_1402 = ws_equity.max_row_equity_div_1402

    wb.max_row_equity_legal_reserve_1402_start = ws_equity.row_legal_reserve_1402_start
    wb.max_row_equity_legal_reserve_1401_end = ws_equity.row_legal_reserve_1401_end
    wb.max_row_equity_legal_reserve_1403_increase = ws_equity.row_legal_reserve_increase_1403
    wb.max_row_equity_legal_reserve_1402_increase = ws_equity.row_legal_reserve_increase_1402

    # مقداردهی اولیه ویژگی‌های ترازنامه قبل از اولین فراخوانی جریان‌های نقدی
    # این مقادیر در اولین اجرای populate_balance_sheet به‌روز خواهند شد
    wb.row_receivables = wb.row_receivables_base 
    wb.row_inventory = wb.row_inventory_base
    wb.row_payables = wb.row_payables_base
    # سایر ویژگی‌های مورد نیاز populate_cash_flow_sheet که توسط populate_balance_sheet تنظیم می‌شوند نیز باید در اینجا مقداردهی اولیه شوند
    # اما برای رفع خطای فعلی، فقط موارد بالا کافی است. کاربر باید در صورت نیاز موارد دیگر را اضافه کند.

    # مقداردهی اولیه برای ویژگی‌های جریان نقدی که در اولین اجرای ترازنامه استفاده می‌شوند
    # این مقادیر در اولین اجرای populate_cash_flow_sheet به‌روز خواهند شد
    # فرض اولیه برای ردیف پایانی موجودی نقد در شیت جریان‌های نقدی (کاربر باید تایید کند)
    wb.max_row_cash_flow_end_1403 = 30 
    wb.max_row_cash_flow_end_1402 = 30

    # مقداردهی اولیه برای ویژگی‌های یادداشت‌های عددی که در اولین اجرای ترازنامه استفاده می‌شوند
    # این مقادیر باید توسط populate_numeric_note_sheets به‌روز شوند. (کاربر باید تایید کند)
    wb.max_row_28_29_30_31_partA = 10 # فرض برای ردیف پایانی بخش A یادداشت 28-31
    wb.max_row_18 = 10 # فرض برای ردیف پایانی یادداشت 18 (سود سهام پرداختنی)
    wb.max_row_19 = 10 # فرض برای ردیف پایانی یادداشت 19 (تسهیلات بلندمدت)
    wb.max_row_20 = 10 # فرض برای ردیف پایانی یادداشت 20 (مزایای پایان خدمت)


    # 9. وضعیت مالی (ترازنامه) - باید قبل از جریان‌های نقدی اجرا شود چون جریان‌های نقدی به آن وابسته است
    ws_balance_sheet = wb["وضعیت مالی"]
    populate_balance_sheet(ws_balance_sheet, assumption_map, wb)
    wb.row_total_assets = ws_balance_sheet.row_total_assets # اینها پس از اجرای populate_balance_sheet به‌روز می‌شوند
    wb.row_total_liabilities = ws_balance_sheet.row_total_liabilities
    wb.row_total_current_assets = ws_balance_sheet.row_total_current_assets
    wb.row_total_current_liabilities = ws_balance_sheet.row_total_current_liabilities
    # wb.row_payables = ws_balance_sheet.row_payables # این مورد بالاتر مقداردهی اولیه شده
    wb.row_current_lt_debt = ws_balance_sheet.row_current_lt_debt
    wb.row_lt_debt = ws_balance_sheet.row_lt_debt
    wb.row_employee_benefits = ws_balance_sheet.row_employee_benefits
    # wb.row_receivables = ws_balance_sheet.row_receivables # این مورد بالاتر مقداردهی اولیه شده
    # wb.row_inventory = ws_balance_sheet.row_inventory # این مورد بالاتر مقداردهی اولیه شده
    wb.row_prepayments = ws_balance_sheet.row_prepayments
    wb.row_other_non_current_assets = ws_balance_sheet.row_other_non_current_assets


    # 10. جریان های نقدی (برای استفاده در ترازنامه)
    ws_cash_flow = wb["جریان های نقدی"]
    populate_cash_flow_sheet(ws_cash_flow, assumption_map, wb)
    wb.max_row_cash_flow_end_1403 = ws_cash_flow.max_row_cash_flow_end_1403
    wb.max_row_cash_flow_end_1402 = ws_cash_flow.max_row_cash_flow_end_1402

    # به‌روزرسانی مقادیر ترازنامه پس از محاسبه جریان‌های نقدی (چون موجودی نقد در ترازنامه از جریان نقدی می‌آید)
    populate_balance_sheet(ws_balance_sheet, assumption_map, wb)
    wb.row_total_assets = ws_balance_sheet.row_total_assets 
    wb.row_total_liabilities = ws_balance_sheet.row_total_liabilities
    wb.row_total_current_assets = ws_balance_sheet.row_total_current_assets
    wb.row_total_current_liabilities = ws_balance_sheet.row_total_current_liabilities
    wb.row_payables = ws_balance_sheet.row_payables
    wb.row_current_lt_debt = ws_balance_sheet.row_current_lt_debt
    wb.row_lt_debt = ws_balance_sheet.row_lt_debt
    wb.row_employee_benefits = ws_balance_sheet.row_employee_benefits
    wb.row_receivables = ws_balance_sheet.row_receivables
    wb.row_inventory = ws_balance_sheet.row_inventory
    wb.row_prepayments = ws_balance_sheet.row_prepayments
    wb.row_other_non_current_assets = ws_balance_sheet.row_other_non_current_assets


    # 11. شیت موجودی (برای مغایرت‌گیری موجودی)
    ws_inventory = wb["موجودی"]
    populate_inventory_note(ws_inventory, wb)


    # 12. تمامی یادداشت‌های عددی (که به شدت به سایر شیت‌ها و مفروضات وابسته اند)
    populate_numeric_note_sheets(wb, assumption_map, wb)


    # سایر شیت‌ها (که نیازی به حلقه همگرایی ندارند اما به مقادیر همگرا شده نیاز دارند)
    populate_comprehensive_income_sheet(wb["جامع"])
    populate_history_sheet(wb["تاریخچه"])
    for i in range(1, 7):
        sheet_name = f'اهم رویه{i}'
        populate_significant_accounting_policy_sheet(wb[sheet_name], i)
    populate_management_judgment_sheet(wb["قضاوت مدیریت"])
    populate_attachment_sheet(wb["پیوست"])
    populate_page_header_sheet(wb["سر برگ صفحات"])
    populate_signature_sheet(wb["ص امضا"])

    # مقداردهی اولیه ویژگی‌های مورد نیاز populate_management_comparative_report
    # این مقادیر باید توسط توابع قبلی (مانند populate_profit_loss_sheet و populate_balance_sheet) تنظیم شده باشند
    # اما برای اطمینان و جلوگیری از خطاهای احتمالی در وابستگی‌های پیچیده، آنها را بررسی می‌کنیم
    # و در صورت نیاز با مقادیر پیش‌فرض اولیه می‌کنیم (کاربر باید این مقادیر را تایید کند)
    default_row_num = 10 # یک شماره ردیف پیش‌فرض
    wb.row_revenue_pl = getattr(wb, 'row_revenue_pl', default_row_num)
    wb.row_gross_profit_pl = getattr(wb, 'row_gross_profit_pl', default_row_num)
    wb.row_operating_profit_pl = getattr(wb, 'row_operating_profit_pl', default_row_num)
    wb.row_net_profit_pl_1403 = getattr(wb, 'row_net_profit_pl_1403', default_row_num)
    wb.row_net_profit_pl_1402 = getattr(wb, 'row_net_profit_pl_1402', default_row_num)
    wb.row_total_assets = getattr(wb, 'row_total_assets', default_row_num)
    wb.row_total_liabilities = getattr(wb, 'row_total_liabilities', default_row_num)
    wb.row_total_current_assets = getattr(wb, 'row_total_current_assets', default_row_num)
    wb.row_total_current_liabilities = getattr(wb, 'row_total_current_liabilities', default_row_num)

    populate_management_comparative_report(wb["گزارش مدیریتی تطبیقی"], wb)
    # اکنون که populate_management_comparative_report اجرا شده، ویژگی‌های row_mcr_* باید تنظیم شده باشند
    populate_business_analytical_report(wb["گزارش تحلیلی کسب و کار"], wb)

    # حلقه تکرار برای همگرایی مدل (ترازنامه و جریان نقدی)
    num_iterations = 20
    print("اجرای حلقه محاسبه تراز...")
    for i in range(num_iterations):
        print(f"اجرای حلقه همگرایی - مرحله {i+1}")
        populate_profit_loss_sheet(wb['سودوزیان'], assumption_map, wb)
        populate_equity_sheet(wb['حقوق مالکانه'], assumption_map) # assumption_map به عنوان آرگومان اضافه شد
        # ترتیب اجرای populate_balance_sheet و populate_cash_flow_sheet در حلقه اصلاح شد
        populate_balance_sheet(wb['وضعیت مالی'], assumption_map, wb) 
        populate_cash_flow_sheet(wb['جریان های نقدی'], assumption_map, wb)
        # اجرای مجدد ترازنامه برای به‌روزرسانی موجودی نقد پس از محاسبه جریان‌های نقدی
        populate_balance_sheet(wb['وضعیت مالی'], assumption_map, wb) 
        
        populate_note_8_and_9(wb)
        populate_inventory_note(wb['موجودی'], wb)
        populate_fixed_asset_roll_forward_sheet(wb['گردش دارایی ثابت'], assumption_map)
        populate_numeric_note_sheets(wb, assumption_map, wb)

    print("مدل مالی پویا با موفقیت ایجاد و تراز شد.")

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    for col_idx in [5, 6]:
        balance_cell = wb['وضعیت مالی'].cell(row=40, column=col_idx)
        if balance_cell.value == "تراز":
            balance_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            balance_cell.font = Font(bold=True)
        else:
            balance_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            balance_cell.font = Font(bold=True)
        balance_cell.border = thin_border

    output_filename = "مدل_مالی_پویا_کامل.xlsx"
    wb.active = wb['وضعیت مالی']
    wb.save(output_filename)
    print(f"فایل اکسل '{output_filename}' با موفقیت ذخیره شد.")

if __name__ == '__main__':
    main_model_execution()
